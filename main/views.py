from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, authenticate
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, time, timedelta
from .models import ChecklistBase, ChecklistDynamicValue, ChecksheetContentConfig, DTPMVerificationHistory, DailyVerificationStatus, ErrorPreventionCheck, ErrorPreventionCheckHistory, ErrorPreventionMechanism, ErrorPreventionMechanismHistory, ErrorPreventionMechanismStatus, SubgroupEntry, SubgroupFrequencyConfig, Verification, Shift, User
from .forms import ChecklistBaseForm, DailyVerificationWorkflowForm, ErrorPreventionCheckForm, ErrorPreventionFilterForm, ErrorPreventionStatusFormSet, ErrorPreventionWorkflowForm, FTQRecordEditForm, SubgroupEntryForm, SubgroupVerificationForm, VerificationForm, ConcernForm, UserRegistrationForm
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
from django.utils import timezone
from django.db.models import Count, Q, Avg
from datetime import timedelta
import json
from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from datetime import timedelta
from itertools import chain
import json
from django.shortcuts import render
from django.db import transaction

from django.utils import timezone
from datetime import timedelta, datetime, time


# Authentication Views
def register_user(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Registration successful!')
            return redirect('dashboard')
        else:
            print(form.errors)  # Debugging: Prints errors in console
            messages.error(request, 'Registration failed. Please correct the errors below.')
    else:
        form = UserRegistrationForm()
    return render(request, 'main/register.html', {'form': form})

def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'main/login.html')


# Updated utility functions for views.py
def get_current_shift_type():
    """Get current shift type based on time with proper priority"""
    current_datetime = timezone.localtime(timezone.now())
    current_hour = current_datetime.hour
    current_minute = current_datetime.minute
    current_time_minutes = current_hour * 60 + current_minute
    
    # Define shift time ranges - ORDER MATTERS!
    # More specific shifts first, then broader ones
    
    if current_time_minutes >= (23 * 60 + 30) or current_time_minutes < (6 * 60 + 30):  # 11:30 PM to 6:30 AM
        return 'C'
    elif (18 * 60 + 30) <= current_time_minutes < (23 * 60 + 30):  # 6:30 PM to 11:30 PM
        return 'S2'  # S2 evening portion
    elif (15 * 60) <= current_time_minutes < (23 * 60 + 30):  # 3:00 PM to 11:30 PM
        return 'B'
    elif (8 * 60 + 30) <= current_time_minutes < (17 * 60):  # 8:30 AM to 5:00 PM
        return 'G'
    elif (6 * 60 + 30) <= current_time_minutes < (15 * 60):  # 6:30 AM to 3:00 PM
        return 'A'
    elif (6 * 60 + 30) <= current_time_minutes < (18 * 60 + 30):  # 6:30 AM to 6:30 PM
        return 'S1'
    else:  # 6:30 PM to 6:30 AM (next day)
        return 'S2'
    
def get_current_shift(user):
    """Get or create the current shift based on date and time"""
    current_date = timezone.now().date()
    shift_type = get_current_shift_type()
    
    # Try to find an existing shift for today and this shift type
    try:
        shift = Shift.objects.get(date=current_date, shift_type=shift_type)
    except Shift.DoesNotExist:
        # Create a new shift if none exists
        shift_supervisor = User.objects.filter(user_type='shift_supervisor').first()
        quality_supervisor = User.objects.filter(user_type='quality_supervisor').first()
        
        if not shift_supervisor or not quality_supervisor:
            raise ValueError("Cannot create shift: No supervisors available")
        
        shift = Shift.objects.create(
            date=current_date,
            shift_type=shift_type,
            operator=user,
            shift_supervisor=shift_supervisor,
            quality_supervisor=quality_supervisor
        )
    
    return shift

def get_shift_start_time(shift_type, date):
    """Get the start time for a given shift type and date"""
    from django.utils import timezone
    import datetime
    
    # Define shift start times
    shift_start_times = {
        'S1': datetime.time(6, 30),   # 6:30 AM
        'A': datetime.time(6, 30),    # 6:30 AM  
        'G': datetime.time(8, 30),    # 8:30 AM
        'B': datetime.time(15, 0),    # 3:00 PM
        'C': datetime.time(23, 30),   # 11:30 PM
        'S2': datetime.time(18, 30),  # 6:30 PM
    }
    
    start_time = shift_start_times.get(shift_type)
    if not start_time:
        return None
    
    # For night shifts that cross midnight, adjust the date
    if shift_type == 'C' and timezone.localtime(timezone.now()).time() < datetime.time(12, 0):
        # If it's before noon and C shift, the shift started yesterday
        date = date - datetime.timedelta(days=1)
    
    # Combine date and time and make it timezone aware
    shift_start_datetime = timezone.make_aware(
        datetime.datetime.combine(date, start_time),
        timezone.get_current_timezone()
    )
    
    return shift_start_datetime

def get_expected_subgroup_times(shift_type, date_param, max_subgroups=6):
    """Get list of expected subgroup times for a shift"""
    shift_start = get_shift_start_time(shift_type, date_param)
    if not shift_start:
        return []
    
    expected_times = []
    for i in range(max_subgroups):
        expected_time = shift_start + timedelta(hours=2 * i)
        expected_times.append(expected_time)
    
    return expected_times


def check_time_gap_shift_based(checklist):
    """
    Check if operator can add a new subgroup based on shift timing and frequency config
    Returns: (can_add, next_allowed_time, available_slots)
    """
    if not checklist or not checklist.verification_status or not checklist.verification_status.shift:
        return False, None, 0
    
    current_time = timezone.localtime(timezone.now())
    current_date = current_time.date()
    shift_type = checklist.verification_status.shift.shift_type
    
    # Get frequency configuration - UPDATED
    frequency_hours = 2  # Default fallback
    max_subgroups = 6    # Default fallback
    
    if checklist.frequency_config and checklist.frequency_config.is_active:
        frequency_hours = checklist.frequency_config.frequency_hours
        max_subgroups = checklist.frequency_config.max_subgroups
    else:
        # Try to get config based on model if not linked
        try:
            config = SubgroupFrequencyConfig.objects.get(
                model_name=checklist.selected_model,
                is_active=True
            )
            frequency_hours = config.frequency_hours
            max_subgroups = config.max_subgroups
            
            # Auto-link the config to the checklist for future use
            if not checklist.frequency_config:
                checklist.frequency_config = config
                checklist.save(update_fields=['frequency_config'])
        except SubgroupFrequencyConfig.DoesNotExist:
            pass  # Use defaults
    
    # Get expected subgroup times with dynamic frequency - UPDATED
    expected_times = get_expected_subgroup_times(shift_type, current_date, checklist)
    
    if not expected_times:
        return False, None, 0
    
    # Get existing subgroups
    existing_subgroups = checklist.subgroup_entries.count()
    
    # Check if we've reached maximum subgroups - UPDATED to use config
    if existing_subgroups >= max_subgroups:
        return False, None, 0
    
    # Find how many subgroups the operator should be able to add based on current time
    available_slots = 0
    next_allowed_time = None
    
    for i, expected_time in enumerate(expected_times):
        if current_time >= expected_time:
            available_slots = i + 1
        else:
            if next_allowed_time is None:
                next_allowed_time = expected_time
            break
    
    # Operator can add subgroups if they have available slots
    can_add = available_slots > existing_subgroups
    
    # If they can't add now, find the next time they can - UPDATED
    if not can_add and existing_subgroups < max_subgroups:
        if existing_subgroups < len(expected_times):
            next_allowed_time = expected_times[existing_subgroups]
    
    return can_add, next_allowed_time, available_slots


def get_expected_subgroup_times(shift_type, date_param, checklist=None):
    """Get list of expected subgroup times for a shift based on frequency config"""
    shift_start = get_shift_start_time(shift_type, date_param)
    if not shift_start:
        return []
    
    # Get frequency configuration - UPDATED
    frequency_hours = 2  # Default
    max_subgroups = 6    # Default
    
    if checklist:
        if checklist.frequency_config and checklist.frequency_config.is_active:
            frequency_hours = checklist.frequency_config.frequency_hours
            max_subgroups = checklist.frequency_config.max_subgroups
        else:
            # Try to get config based on model
            try:
                config = SubgroupFrequencyConfig.objects.get(
                    model_name=checklist.selected_model,
                    is_active=True
                )
                frequency_hours = config.frequency_hours
                max_subgroups = config.max_subgroups
            except SubgroupFrequencyConfig.DoesNotExist:
                pass
    
    expected_times = []
    for i in range(max_subgroups):
        expected_time = shift_start + timedelta(hours=frequency_hours * i)
        expected_times.append(expected_time)
    
    return expected_times


# Keep this for backward compatibility, but it's deprecated
def check_time_gap(last_subgroup, frequency_hours=2):
    """
    DEPRECATED: Use check_time_gap_shift_based instead
    Check if enough time has passed since last subgroup
    """
    if not last_subgroup:
        return True
    
    time_difference = timezone.now() - last_subgroup.timestamp
    return time_difference >= timedelta(hours=frequency_hours)


 
# Main Views
@login_required
def dashboard(request):
    """Route to appropriate dashboard based on user type"""
    try:
        if request.user.user_type == 'operator':
            return operator_dashboard(request)
        elif request.user.user_type == 'shift_supervisor':
            return supervisor_dashboard(request)
        elif request.user.user_type == 'quality_supervisor':
            return quality_dashboard(request)
        else:
            # Default to operator dashboard if user_type is not set
            return operator_dashboard(request)
    except Exception as e:
        messages.error(request, f"Error loading dashboard: {str(e)}")
        return render(request, 'main/error.html', {'error': str(e)})

from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import timedelta

from datetime import datetime, timedelta
from django.utils import timezone



# *********************************************************************





from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q, Count, Prefetch
from django.utils import timezone
from datetime import timedelta
from .models import (
    Checksheet, ChecksheetSection, ChecksheetField,
    ChecksheetResponse, ChecksheetFieldResponse,
    DailyVerificationStatus, Shift, User,
    ChecklistBase, FTQRecord, ErrorPreventionCheck, 
    DTPMChecklistFMA03New, SubgroupFrequencyConfig
)


@login_required
def operator_dashboard(request):
    current_datetime = timezone.localtime(timezone.now())
    current_date = current_datetime.date()
    current_time = current_datetime.time()
    
    # Get active verification status
    active_verification = DailyVerificationStatus.objects.filter(
        created_by=request.user,
        date=current_date
    ).select_related('shift').first()
    
    # Get active checklist
    active_checklist = None
    if active_verification:
        active_checklist = ChecklistBase.objects.filter(
            verification_status=active_verification
        ).select_related('frequency_config').prefetch_related('subgroup_entries').first()

    # Get frequency configuration
    frequency_hours = 2  # Default
    max_subgroups = 6    # Default
    
    if active_checklist:
        if active_checklist.frequency_config and active_checklist.frequency_config.is_active:
            frequency_hours = active_checklist.frequency_config.frequency_hours
            max_subgroups = active_checklist.frequency_config.max_subgroups
        else:
            try:
                config = SubgroupFrequencyConfig.objects.get(
                    model_name=active_checklist.selected_model,
                    is_active=True
                )
                frequency_hours = config.frequency_hours
                max_subgroups = config.max_subgroups
            except SubgroupFrequencyConfig.DoesNotExist:
                pass

    # Calculate next subgroup time
    next_subgroup_time = None
    time_remaining = None
    time_remaining_formatted = None
    can_add_subgroup = False
    available_slots = 0
    expected_schedule = []
    
    if active_checklist:
        can_add_subgroup, next_allowed_time, available_slots = check_time_gap_shift_based(active_checklist)
        
        if next_allowed_time and not can_add_subgroup:
            next_subgroup_time = next_allowed_time
            time_remaining = next_allowed_time - current_datetime
            
            if time_remaining.total_seconds() > 0:
                hours, remainder = divmod(int(time_remaining.total_seconds()), 3600)
                minutes, _ = divmod(remainder, 60)
                
                time_parts = []
                if hours > 0:
                    time_parts.append(f"{hours} hour{'s' if hours != 1 else ''}")
                if minutes > 0 or not time_parts:
                    time_parts.append(f"{minutes} min{'s' if minutes != 1 else ''}")
                    
                time_remaining_formatted = " ".join(time_parts)
        
        if active_verification and active_verification.shift:
            expected_times = get_expected_subgroup_times(
                active_verification.shift.shift_type, 
                current_date,
                active_checklist
            )
            for i, expected_time in enumerate(expected_times, 1):
                expected_schedule.append({
                    'subgroup_number': i,
                    'expected_time': expected_time,
                    'is_available': current_datetime >= expected_time,
                    'is_completed': active_checklist.subgroup_entries.filter(subgroup_number=i).exists()
                })

    # ============ CHECKSHEET INTEGRATION ============
    
    # Get active checksheets
    active_checksheets = Checksheet.objects.filter(is_active=True).annotate(
        total_sections=Count('sections', filter=Q(sections__is_active=True)),
        total_fields=Count('sections__fields', filter=Q(sections__is_active=True, sections__fields__is_active=True))
    ).order_by('name')
    
    # Get today's checksheet responses for this user
    today_checksheet_responses = ChecksheetResponse.objects.filter(
        filled_by=request.user,
        created_at__date=current_date
    ).select_related('checksheet').order_by('-created_at')
    
    # Get pending checksheet responses (draft or submitted but not approved)
    pending_checksheet_responses = ChecksheetResponse.objects.filter(
        filled_by=request.user,
        status__in=['draft', 'submitted', 'supervisor_approved']
    ).exclude(
        created_at__date=current_date
    ).select_related('checksheet').order_by('-created_at')[:5]
    
    # Check if there are any checksheets that need to be filled today
    checksheets_need_filling = []
    for checksheet in active_checksheets:
        # Check if already filled today
        already_filled = today_checksheet_responses.filter(checksheet=checksheet).exists()
        if not already_filled:
            checksheets_need_filling.append(checksheet)
    
    # Checksheet statistics
    checksheet_stats = {
        'total_active': active_checksheets.count(),
        'filled_today': today_checksheet_responses.count(),
        'pending': pending_checksheet_responses.count(),
        'need_filling': len(checksheets_need_filling),
        'completion_rate': 0
    }
    
    if checksheet_stats['total_active'] > 0:
        checksheet_stats['completion_rate'] = int(
            (checksheet_stats['filled_today'] / checksheet_stats['total_active']) * 100
        )
    
    # Recent checksheet responses (last 7 days)
    recent_checksheet_responses = ChecksheetResponse.objects.filter(
        filled_by=request.user,
        created_at__gte=current_datetime - timedelta(days=7)
    ).select_related('checksheet').order_by('-created_at')[:10]

    # ============ END CHECKSHEET INTEGRATION ============

    # Debug info
    debug_info = {
        'verification_exists': bool(active_verification),
        'verification_status': active_verification.status if active_verification else 'None',
        'verification_shift': active_verification.shift.shift_type if active_verification and active_verification.shift else 'None',
        'checklist_exists': bool(active_checklist),
        'checklist_status': active_checklist.status if active_checklist else 'None',
        'checklist_shift': active_checklist.shift if active_checklist else 'None',
        'subgroup_count': active_checklist.subgroup_entries.count() if active_checklist else 0,
        'user': request.user.username,
        'current_date': current_date,
        'can_add_subgroup': can_add_subgroup,
        'available_slots': available_slots,
        'current_time': current_datetime.strftime('%H:%M:%S'),
        'shift_type': active_verification.shift.shift_type if active_verification and active_verification.shift else None,
        'frequency_hours': frequency_hours,
        'max_subgroups': max_subgroups,
    }
    
    if active_checklist and active_verification and active_verification.shift:
        expected_times = get_expected_subgroup_times(
            active_verification.shift.shift_type, 
            current_date,
            active_checklist
        )
        debug_info['expected_subgroup_times'] = [
            t.strftime('%H:%M') for t in expected_times
        ]
    
    # Current shift determination
    current_shift_type = get_current_shift_type()
    SHIFT_CHOICES = [
        ('S1', 'S1 - 6:30 AM to 6:30 PM'),
        ('A', 'A - 6:30 AM to 3:00 PM'),
        ('G', 'G - 8:30 AM to 5:00 PM'),
        ('B', 'B - 3:00 PM to 11:30 PM'),
        ('C', 'C - 11:30 PM to 6:30 AM'),
        ('S2', 'S2 - 6:30 PM to 6:30 AM'),
    ]
    current_shift_display = dict(SHIFT_CHOICES).get(current_shift_type, current_shift_type)
    
    debug_info.update({
        'current_shift_type': current_shift_type,
        'current_shift_display': current_shift_display
    })
    
    # Get recent FTQ records for this operator
    recent_ftq_records = FTQRecord.objects.filter(
        created_by=request.user,
        date__gte=current_date - timedelta(days=7)
    ).select_related('verification_status').order_by('-date', '-created_at')[:5]
    
    # Get today's FTQ record if it exists
    today_ftq_query = FTQRecord.objects.filter(
        created_by=request.user,
        date=current_date
    )
    
    today_ftq = None
    if active_verification:
        today_ftq = today_ftq_query.filter(verification_status=active_verification).first()
        if not today_ftq and active_verification.shift:
            today_ftq = today_ftq_query.filter(shift_type=active_verification.shift.shift_type).first()
    
    if not today_ftq:
        today_ftq = today_ftq_query.filter(shift_type=current_shift_type).first()
    
    # Calculate FTQ percentage for display
    ftq_percentage = None
    if today_ftq and today_ftq.total_inspected > 0:
        ftq_percentage = ((today_ftq.total_inspected - today_ftq.total_defects) / today_ftq.total_inspected) * 100
    
    # Get today's EP check if it exists
    today_ep_check = None
    if active_verification:
        today_ep_check = ErrorPreventionCheck.objects.filter(
            verification_status=active_verification,
            date=current_date,
        ).select_related('verification_status').first()
    
    # Get recent EP checks for this operator
    recent_ep_checks = ErrorPreventionCheck.objects.filter(
        operator=request.user,
        date__gte=current_date - timedelta(days=7)
    ).select_related('verification_status').order_by('-date', '-created_at')[:5]
    
    # EP Check status summary
    ep_check_stats = {
        'ok_count': 0,
        'ng_count': 0,
        'na_count': 0,
        'total_mechanisms': 0,
        'current_model': None,
        'shift': None
    }
    
    if today_ep_check:
        mechanism_statuses = today_ep_check.mechanism_statuses.all()
        ep_check_stats['total_mechanisms'] = mechanism_statuses.count()
        ep_check_stats['ok_count'] = mechanism_statuses.filter(status='OK').count()
        ep_check_stats['ng_count'] = mechanism_statuses.filter(status='NG').count()
        ep_check_stats['na_count'] = mechanism_statuses.filter(is_not_applicable=True).count()
        ep_check_stats['current_model'] = today_ep_check.current_model
        ep_check_stats['shift'] = today_ep_check.shift
    
    # Get today's DTPM checklist if it exists
    today_dtpm_checklist = None
    if active_verification:
        today_dtpm_checklist = DTPMChecklistFMA03New.objects.filter(
            verification_status=active_verification,
            date=current_date,
        ).select_related('verification_status').first()
    
    # Get recent DTPM checklists for this operator
    recent_dtpm_checklists = DTPMChecklistFMA03New.objects.filter(
        operator=request.user,
        date__gte=current_date - timedelta(days=7)
    ).select_related('verification_status').order_by('-date', '-created_at')[:5]
    
    # DTPM Checklist status summary
    dtpm_checklist_stats = {
        'ok_count': 0,
        'ng_count': 0,
        'total_checkpoints': 0,
        'current_model': None,
        'shift': None
    }
    
    if today_dtpm_checklist:
        checkpoint_results = today_dtpm_checklist.check_results.all()
        dtpm_checklist_stats['total_checkpoints'] = checkpoint_results.count()
        dtpm_checklist_stats['ok_count'] = checkpoint_results.filter(status='OK').count()
        dtpm_checklist_stats['ng_count'] = checkpoint_results.filter(status='NG').count()
        dtpm_checklist_stats['current_model'] = today_dtpm_checklist.current_model
        dtpm_checklist_stats['shift'] = today_dtpm_checklist.checklist_shift
    
    # Check if should show workflow modal automatically
    show_workflow_modal = active_verification and not (today_ftq and today_ep_check and today_dtpm_checklist)
    
    # Enhanced can_create_new logic
    can_create_new = False
    if not active_verification:
        can_create_new = True
    elif not active_checklist:
        can_create_new = True
    
    context = {
        'current_date': current_date,
        'current_time': current_time,
        'current_shift': current_shift_display,
        'current_shift_type': current_shift_type,
        'active_verification': active_verification,
        'active_checklist': active_checklist,
        'can_create_new': can_create_new,
        'can_add_subgroup': can_add_subgroup,
        'next_subgroup_time': next_subgroup_time,
        'time_remaining': time_remaining,
        'time_remaining_formatted': time_remaining_formatted,
        'available_slots': available_slots,
        'expected_schedule': expected_schedule,
        'recent_ftq_records': recent_ftq_records,
        'today_ftq': today_ftq,
        'ftq_percentage': ftq_percentage,
        'today_ep_check': today_ep_check,
        'recent_ep_checks': recent_ep_checks,
        'ep_check_stats': ep_check_stats,
        'today_dtpm_checklist': today_dtpm_checklist,
        'recent_dtpm_checklists': recent_dtpm_checklists,
        'dtpm_checklist_stats': dtpm_checklist_stats,
        'show_workflow_modal': show_workflow_modal,
        'debug_info': debug_info,
        'frequency_hours': frequency_hours,
        'max_subgroups': max_subgroups,
        'current_subgroup_count': active_checklist.subgroup_entries.count() if active_checklist else 0,
        
        # Checksheet context
        'active_checksheets': active_checksheets,
        'today_checksheet_responses': today_checksheet_responses,
        'pending_checksheet_responses': pending_checksheet_responses,
        'checksheets_need_filling': checksheets_need_filling,
        'checksheet_stats': checksheet_stats,
        'recent_checksheet_responses': recent_checksheet_responses,
    }
    
    return render(request, 'main/operator_dashboard.html', context)





@login_required
@user_passes_test(lambda u: u.user_type == 'operator')


@login_required
@user_passes_test(lambda u: u.user_type == 'operator')
def create_checklist(request):
    """
    Handles the creation of an initial checklist without custom parameters.

    This view manages the following:
    - Renders the form for a new checklist.
    - Prevents creation if a pending checklist already exists for the user on the current day.
    - Fetches and passes dynamic configuration data (measurement frequencies, model-specific parameters)
      to the frontend.
    - On successful submission, creates a Shift, a DailyVerificationStatus, the ChecklistBase instance,
      and saves all standard and dynamic parameter values within a single database transaction.
    """
    current_datetime = timezone.now()
    current_date = current_datetime.date()
    suggested_shift_type = get_current_shift_type()

    # Prevent creating a new checklist if one is already in progress for the day
    existing_verification = DailyVerificationStatus.objects.filter(
        created_by=request.user,
        date=current_date,
        status__in=['pending', 'in_progress']
    ).first()

    if existing_verification:
        messages.error(request, 'A verification workflow already exists for today. Please complete or cancel the existing one.')
        return redirect('operator_dashboard')

    # Prepare data for the frontend JavaScript
    frequency_configs = {
        config.model_name: {
            'frequency_hours': config.frequency_hours,
            'max_subgroups': config.max_subgroups
        }
        for config in SubgroupFrequencyConfig.objects.filter(is_active=True)
    }

    dynamic_parameters = {}
    for model_choice in ChecklistBase.MODEL_CHOICES:
        model_name = model_choice[0]
        params = ChecksheetContentConfig.objects.filter(
            model_name=model_name, is_active=True
        ).order_by('order')
        dynamic_parameters[model_name] = [
            {
                'id': p.id,
                'parameter_name': p.parameter_name,
                'parameter_name_hindi': p.parameter_name_hindi,
                'measurement_type': p.measurement_type,
                'min_value': p.min_value,
                'max_value': p.max_value,
                'unit': p.unit,
                'requires_comment_if_nok': p.requires_comment_if_nok,
                'help_text': f"Range: {p.min_value} - {p.max_value} {p.unit}" if p.min_value is not None else ""
            } for p in params
        ]

    if request.method == 'POST':
        form = ChecklistBaseForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                with transaction.atomic():
                    selected_shift_type = form.cleaned_data.get('new_shift', suggested_shift_type)
                    selected_model = form.cleaned_data.get('selected_model')

                    # 1. Create Shift record
                    shift = Shift.objects.create(
                        date=current_date,
                        shift_type=selected_shift_type,
                        operator=request.user,
                        shift_supervisor=User.objects.filter(user_type='shift_supervisor').first(),
                        quality_supervisor=User.objects.filter(user_type='quality_supervisor').first()
                    )

                    # 2. Create Verification Status record
                    verification_status = DailyVerificationStatus.objects.create(
                        date=current_date,
                        shift=shift,
                        status='pending',
                        created_by=request.user
                    )

                    # 3. Get or create the frequency configuration
                    frequency_config, _ = SubgroupFrequencyConfig.objects.get_or_create(
                        model_name=selected_model,
                        defaults={'frequency_hours': 2, 'max_subgroups': 6, 'is_active': True}
                    )

                    # 4. Save the main checklist instance
                    checklist = form.save(commit=False)
                    checklist.verification_status = verification_status
                    checklist.frequency_config = frequency_config
                    checklist.save()

                    # 5. Save dynamic field values from admin-configured parameters
                    for key in request.POST:
                        if key.startswith('dynamic_') and not key.endswith('_comment'):
                            param_id_str = key.replace('dynamic_', '')
                            if param_id_str.isdigit():
                                try:
                                    param = ChecksheetContentConfig.objects.get(id=int(param_id_str))
                                    value = request.POST.get(key)
                                    comment = request.POST.get(f'{key}_comment', '')
                                    if value:
                                        ChecklistDynamicValue.objects.create(
                                            checklist=checklist, parameter=param, value=value, comment=comment
                                        )
                                except ChecksheetContentConfig.DoesNotExist:
                                    continue
                    
                    # 6. Provide feedback and redirect
                    messages.success(request, f'Checklist created successfully for {selected_shift_type} shift!')
                    if hasattr(form, 'warnings'):
                        for warning in form.warnings:
                            messages.warning(request, warning)
                    
                    return redirect('add_subgroup', checklist_id=checklist.id)

            except Exception as e:
                messages.error(request, f'An unexpected error occurred: {e}')
    else:
        form = ChecklistBaseForm(user=request.user, initial={'new_shift': suggested_shift_type})

    context = {
        'form': form,
        'frequency_configs_json': json.dumps(frequency_configs),
        'dynamic_parameters_json': json.dumps(dynamic_parameters)
    }
    return render(request, 'main/create_checklist.html', context)
 
 
 
# Updated add_subgroup view
@login_required
@user_passes_test(lambda u: u.user_type == 'operator')
def add_subgroup(request, checklist_id):
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    
    # Verify ownership and status
    if checklist.verification_status.created_by != request.user:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': 'You do not have permission to modify this checklist'}, status=403)
        messages.error(request, 'You do not have permission to modify this checklist')
        return redirect('operator_dashboard')
    
    if checklist.status != 'pending':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': 'Cannot add subgroups to a verified checklist'}, status=400)
        messages.error(request, 'Cannot add subgroups to a verified checklist')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    # Use shift-based timing check
    can_add, next_allowed_time, available_slots = check_time_gap_shift_based(checklist)
    
    # Get max subgroups from frequency config
    max_subgroups = 6
    if checklist.frequency_config:
        max_subgroups = checklist.frequency_config.max_subgroups
    
    if not can_add:
        error_msg = "Cannot add subgroup at this time based on shift schedule."
        if next_allowed_time:
            time_str = timezone.localtime(next_allowed_time).strftime('%H:%M')
            error_msg += f" Next subgroup can be added at {time_str}."
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    # Get current subgroup number
    current_subgroup = checklist.subgroup_entries.count() + 1
    
    if current_subgroup > max_subgroups:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': f'Maximum number of subgroups ({max_subgroups}) reached'}, status=400)
        messages.error(request, f'Maximum number of subgroups ({max_subgroups}) reached')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    # Show timing information
    timing_info = None
    if available_slots > checklist.subgroup_entries.count():
        slots_available = available_slots - checklist.subgroup_entries.count()
        timing_info = f'You can add {slots_available} more subgroup(s) based on shift timing'
        
        if checklist.verification_status and checklist.verification_status.shift:
            current_date = timezone.localtime(timezone.now()).date()
            expected_times = get_expected_subgroup_times(
                checklist.verification_status.shift.shift_type, 
                current_date,
                checklist
            )
            time_schedule = []
            for i, expected_time in enumerate(expected_times[:max_subgroups], 1):
                time_schedule.append(f"Subgroup {i}: {timezone.localtime(expected_time).strftime('%H:%M')}")
            
            if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                messages.info(request, f"Shift schedule: {', '.join(time_schedule)}")
    
    if request.method == 'POST':
        form = SubgroupEntryForm(request.POST, checklist=checklist)
        if form.is_valid():
            try:
                subgroup = form.save(commit=False)
                subgroup.checklist = checklist
                subgroup.subgroup_number = current_subgroup
                subgroup._editing_user = request.user
                subgroup.save()
                
                # Collect warnings and comments
                warnings = []
                comment_summary = []
                
                # Check UV vacuum test readings
                for i in range(1, 6):
                    field_name = f'uv_vacuum_test_{i}'
                    comment_field = f'uv_vacuum_test_{i}_comment'
                    
                    if form.cleaned_data.get(field_name):
                        uv_vacuum = float(form.cleaned_data[field_name])
                        comment = form.cleaned_data.get(comment_field, '').strip()
                        
                        if not (-43 <= uv_vacuum <= -35):
                            warning_msg = f'UV vacuum test {i}: {uv_vacuum} kPa is outside range (-43 to -35 kPa)'
                            if comment:
                                warning_msg += f' - Comment: {comment}'
                            warnings.append(warning_msg)
                        elif comment:
                            comment_summary.append(f'UV vacuum test {i} comment: {comment}')
                
                # Check UV flow value readings
                for i in range(1, 6):
                    field_name = f'uv_flow_value_{i}'
                    comment_field = f'uv_flow_value_{i}_comment'
                    
                    if form.cleaned_data.get(field_name):
                        uv_flow = float(form.cleaned_data[field_name])
                        comment = form.cleaned_data.get(comment_field, '').strip()
                        
                        if not (30 <= uv_flow <= 40):
                            warning_msg = f'UV flow value {i}: {uv_flow} LPM is outside range (30-40 LPM)'
                            if comment:
                                warning_msg += f' - Comment: {comment}'
                            warnings.append(warning_msg)
                        elif comment:
                            comment_summary.append(f'UV flow value {i} comment: {comment}')
                
                # Check NOK values - Umbrella valve
                for i in range(1, 6):
                    field_name = f'umbrella_valve_assembly_{i}'
                    comment_field = f'umbrella_valve_assembly_{i}_comment'
                    
                    if form.cleaned_data.get(field_name) == 'NOK':
                        comment = form.cleaned_data.get(comment_field, '').strip()
                        warning_msg = f'Umbrella valve {i} marked as NOK'
                        if comment:
                            warning_msg += f' - Comment: {comment}'
                        warnings.append(warning_msg)
                
                # Check NOK values - UV clip
                for i in range(1, 6):
                    field_name = f'uv_clip_pressing_{i}'
                    comment_field = f'uv_clip_pressing_{i}_comment'
                    
                    if form.cleaned_data.get(field_name) == 'NOK':
                        comment = form.cleaned_data.get(comment_field, '').strip()
                        warning_msg = f'UV clip pressing {i} marked as NOK'
                        if comment:
                            warning_msg += f' - Comment: {comment}'
                        warnings.append(warning_msg)
                
                # Check workstation cleanliness
                if form.cleaned_data.get('workstation_clean') == 'No':
                    comment = form.cleaned_data.get('workstation_clean_comment', '').strip()
                    warning_msg = 'Workstation marked as not clean'
                    if comment:
                        warning_msg += f' - Comment: {comment}'
                    warnings.append(warning_msg)
                
                # Check bin contamination
                for i in range(1, 6):
                    field_name = f'bin_contamination_check_{i}'
                    comment_field = f'bin_contamination_check_{i}_comment'
                    
                    if form.cleaned_data.get(field_name) == 'No':
                        comment = form.cleaned_data.get(comment_field, '').strip()
                        warning_msg = f'Bin contamination check {i} marked as No'
                        if comment:
                            warning_msg += f' - Comment: {comment}'
                        warnings.append(warning_msg)
                
                # Summary information
                summary_info = [
                    f'UV Vacuum Test Average: {subgroup.uv_vacuum_average:.2f} kPa',
                    f'UV Flow Value Average: {subgroup.uv_flow_average:.2f} LPM',
                    f'Umbrella Valve OK Count: {subgroup.umbrella_valve_ok_count}/5',
                    f'UV Clip OK Count: {subgroup.uv_clip_ok_count}/5',
                    f'Workstation Clean Status: {subgroup.workstation_status}',
                    f'Bin Contamination Check Yes Count: {subgroup.bin_contamination_yes_count}/5'
                ]
                
                if subgroup.is_after_maintenance:
                    summary_info.append('⚠️ This subgroup is marked as after maintenance/ME activity')
                
                if subgroup.requires_nok_approval:
                    summary_info.append('⚠️ This subgroup contains NOK/No entries and requires supervisor approval')
                
                if timing_info:
                    summary_info.append(timing_info)
                
                # For AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'status': 'success',
                        'message': f'Subgroup {current_subgroup} added successfully with 21 readings',
                        'subgroup_id': subgroup.id,
                        'subgroup_number': current_subgroup,
                        'warnings': warnings,
                        'comments': comment_summary,
                        'summary': summary_info,
                        'available_slots': available_slots,
                        'can_add_more': (available_slots > current_subgroup),
                        'requires_approval': subgroup.requires_nok_approval
                    })
                
                # For regular requests
                for warning in warnings:
                    messages.warning(request, warning)
                
                for comment in comment_summary:
                    messages.info(request, comment)
                
                for info in summary_info:
                    messages.info(request, info)
                
                if available_slots > current_subgroup:
                    remaining_slots = available_slots - current_subgroup
                    messages.success(request, f'Subgroup {current_subgroup} added successfully! You can add {remaining_slots} more subgroup(s) based on shift timing.')
                else:
                    messages.success(request, f'Subgroup {current_subgroup} added successfully!')
                
                return redirect('dashboard')
                
            except Exception as e:
                error_message = str(e)
                import traceback
                print(f"Error saving subgroup: {traceback.format_exc()}")

                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Error saving measurements: {error_message}'
                    }, status=500)
                    
                messages.error(request, f'Error saving measurements: {error_message}')
                return redirect('checklist_detail', checklist_id=checklist.id)
        else:
            error_message = "Form validation failed - please check all required fields"
            print(f"Form errors: {form.errors}")
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = []
                for field, error_list in form.errors.items():
                    for error in error_list:
                        errors.append(f"{field}: {error}")
                return JsonResponse({
                    'status': 'error', 
                    'message': error_message,
                    'errors': errors
                }, status=400)
                
            messages.error(request, error_message)
    else:
        form = SubgroupEntryForm(checklist=checklist)
        
        if timing_info:
            messages.info(request, timing_info)
            
        if checklist.verification_status and checklist.verification_status.shift:
            current_date = timezone.localtime(timezone.now()).date()
            expected_times = get_expected_subgroup_times(
                checklist.verification_status.shift.shift_type, 
                current_date,
                checklist
            )
            schedule_info = []
            current_time = timezone.localtime(timezone.now())
            
            for i, expected_time in enumerate(expected_times[:max_subgroups], 1):
                time_str = timezone.localtime(expected_time).strftime('%H:%M')
                status = ""
                if checklist.subgroup_entries.filter(subgroup_number=i).exists():
                    status = " ✓"
                elif current_time >= expected_time:
                    status = " (available)"
                else:
                    status = " (future)"
                schedule_info.append(f"Subgroup {i}: {time_str}{status}")
            
            messages.info(request, f"Shift schedule: {' | '.join(schedule_info)}")
    
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'status': 'error',
            'message': 'GET method not supported for AJAX requests to this endpoint'
        }, status=405)
    
    frequency_info = {
        'frequency_hours': checklist.frequency_config.frequency_hours if checklist.frequency_config else 2,
        'max_subgroups': max_subgroups
    }
    
    return render(request, 'main/add_subgroup.html', {
        'form': form,
        'checklist': checklist,
        'current_subgroup': current_subgroup,
        'available_slots': available_slots,
        'can_add_more': (available_slots > current_subgroup) if available_slots else False,
        'max_subgroups': max_subgroups,
        'frequency_info': frequency_info
    })    
    
    
@login_required
def validate_subgroup(request):
    """API endpoint for validating subgroup data"""
    if request.method == 'POST':
        data = request.POST
        warnings = []
        
        try:
            # Check UV vacuum test
            uv_vacuum = float(data.get('uv_vacuum_test', 0))
            if not (-43 <= uv_vacuum <= -35):
                warnings.append({
                    'field': 'uv_vacuum_test',
                    'message': f'UV vacuum test value {uv_vacuum} kPa is outside recommended range (-43 to -35 kPa)',
                    'value': uv_vacuum,
                    'recommended_range': {'min': -43, 'max': -35}
                })
            
            # Check UV flow value
            uv_flow = float(data.get('uv_flow_value', 0))
            if not (30 <= uv_flow <= 40):
                warnings.append({
                    'field': 'uv_flow_value',
                    'message': f'UV flow value {uv_flow} LPM is outside recommended range (30-40 LPM)',
                    'value': uv_flow,
                    'recommended_range': {'min': 30, 'max': 40}
                })
            
            return JsonResponse({
                'has_warnings': len(warnings) > 0,
                'warnings': warnings,
                'is_valid': True  # Always valid as we're only showing warnings
            })
            
        except (TypeError, ValueError) as e:
            return JsonResponse({
                'is_valid': False,
                'errors': ['Please enter valid numerical values']
            })
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def add_concern(request, checklist_id):
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    
    # Only allow adding concerns if checklist is still pending
    if checklist.status not in ['pending', 'supervisor_approved']:
        messages.error(request, 'Cannot add concerns to a completed checklist')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    if request.method == 'POST':
        form = ConcernForm(request.POST)
        if form.is_valid():
            concern = form.save(commit=False)
            concern.checklist = checklist
            concern.save()
            messages.success(request, 'Concern added successfully')
            return redirect('checklist_detail', checklist_id=checklist.id)
    else:
        form = ConcernForm()
    
    return render(request, 'main/add_concern.html', {
        'form': form,
        'checklist': checklist
    })
    
# views.py - Updated edit_subgroup view for 5 readings

@login_required
@user_passes_test(lambda u: u.user_type == 'operator')
def edit_subgroup(request, checklist_id, subgroup_id):
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    subgroup = get_object_or_404(SubgroupEntry, id=subgroup_id, checklist=checklist)
    
    # Verify ownership and status
    if checklist.verification_status.created_by != request.user:
        messages.error(request, 'You do not have permission to modify this subgroup')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    if checklist.status != 'pending':
        messages.error(request, 'Cannot edit subgroups of a verified checklist')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    if request.method == 'POST':
        form = SubgroupEntryForm(request.POST, instance=subgroup, checklist=checklist)
        if form.is_valid():
            try:
                subgroup._editing_user = request.user
                updated_subgroup = form.save()
                
                # Collect warnings and comments
                warnings = []
                comment_summary = []
                
                # Check UV vacuum test readings
                for i in range(1, 6):
                    field_name = f'uv_vacuum_test_{i}'
                    comment_field = f'uv_vacuum_test_{i}_comment'
                    
                    value = getattr(updated_subgroup, field_name, None)
                    comment = form.cleaned_data.get(comment_field, '').strip()
                    
                    if value is not None and not (-43 <= value <= -35):
                        warning_msg = f'UV vacuum test {i}: {value} kPa is outside range (-43 to -35 kPa)'
                        if comment:
                            warning_msg += f' - Comment: {comment}'
                        warnings.append(warning_msg)
                    elif comment:
                        comment_summary.append(f'UV vacuum test {i} comment: {comment}')
                
                # Check UV flow value readings
                # Check UV flow value readings
                for i in range(1, 6):
                    field_name = f'uv_flow_value_{i}'
                    comment_field = f'uv_flow_value_{i}_comment'
                    
                    value = getattr(updated_subgroup, field_name, None)
                    comment = form.cleaned_data.get(comment_field, '').strip()
                    
                    if value is not None and not (30 <= value <= 40):
                        warning_msg = f'UV flow value {i}: {value} LPM is outside range (30-40 LPM)'
                        if comment:
                            warning_msg += f' - Comment: {comment}'
                        warnings.append(warning_msg)
                    elif comment:
                        comment_summary.append(f'UV flow value {i} comment: {comment}')
                
                # Check NOK values - Umbrella valve
                for i in range(1, 6):
                    field_name = f'umbrella_valve_assembly_{i}'
                    comment_field = f'umbrella_valve_assembly_{i}_comment'
                    
                    if getattr(updated_subgroup, field_name, None) == 'NOK':
                        comment = form.cleaned_data.get(comment_field, '').strip()
                        warning_msg = f'Umbrella valve {i} marked as NOK'
                        if comment:
                            warning_msg += f' - Comment: {comment}'
                        warnings.append(warning_msg)
                
                # Check NOK values - UV clip
                for i in range(1, 6):
                    field_name = f'uv_clip_pressing_{i}'
                    comment_field = f'uv_clip_pressing_{i}_comment'
                    
                    if getattr(updated_subgroup, field_name, None) == 'NOK':
                        comment = form.cleaned_data.get(comment_field, '').strip()
                        warning_msg = f'UV clip pressing {i} marked as NOK'
                        if comment:
                            warning_msg += f' - Comment: {comment}'
                        warnings.append(warning_msg)
                
                # Check workstation cleanliness
                if updated_subgroup.workstation_clean == 'No':
                    comment = form.cleaned_data.get('workstation_clean_comment', '').strip()
                    warning_msg = 'Workstation marked as not clean'
                    if comment:
                        warning_msg += f' - Comment: {comment}'
                    warnings.append(warning_msg)
                
                # Check bin contamination
                for i in range(1, 6):
                    field_name = f'bin_contamination_check_{i}'
                    comment_field = f'bin_contamination_check_{i}_comment'
                    
                    if getattr(updated_subgroup, field_name, None) == 'No':
                        comment = form.cleaned_data.get(comment_field, '').strip()
                        warning_msg = f'Bin contamination check {i} marked as No'
                        if comment:
                            warning_msg += f' - Comment: {comment}'
                        warnings.append(warning_msg)
                
                # Calculate and show summary statistics
                summary_info = []
                if updated_subgroup.uv_vacuum_average > 0:
                    summary_info.append(f'UV Vacuum Test Average: {updated_subgroup.uv_vacuum_average:.2f} kPa')
                if updated_subgroup.uv_flow_average > 0:
                    summary_info.append(f'UV Flow Value Average: {updated_subgroup.uv_flow_average:.2f} LPM')
                
                summary_info.extend([
                    f'Umbrella Valve OK Count: {updated_subgroup.umbrella_valve_ok_count}/5',
                    f'UV Clip OK Count: {updated_subgroup.uv_clip_ok_count}/5',
                    f'Workstation Clean Status: {updated_subgroup.workstation_status}',
                    f'Bin Contamination Check Yes Count: {updated_subgroup.bin_contamination_yes_count}/5'
                ])
                
                # Show all warnings and comments
                for warning in warnings:
                    messages.warning(request, warning)
                
                for comment in comment_summary:
                    messages.info(request, comment)
                
                for info in summary_info:
                    messages.info(request, info)
                
                # Check completion status
                if updated_subgroup.is_complete:
                    messages.success(request, f'Subgroup {subgroup.subgroup_number} updated successfully - All 21 readings complete!')
                else:
                    completion_pct = updated_subgroup.completion_percentage
                    messages.success(request, f'Subgroup {subgroup.subgroup_number} updated successfully - {completion_pct:.1f}% complete ({updated_subgroup.total_readings_count}/21 readings)')
                
                return redirect('checklist_detail', checklist_id=checklist.id)
                
            except Exception as e:
                error_message = str(e)
                import traceback
                print(f"Error updating subgroup: {traceback.format_exc()}")
                messages.error(request, f'Error updating subgroup: {error_message}')
        else:
            # Form validation failed
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            
            messages.error(request, f'Form validation failed: {"; ".join(error_messages)}')
    else:
        form = SubgroupEntryForm(instance=subgroup, checklist=checklist)
    
    # Get edit history for this subgroup
    edit_history = subgroup.edit_history.all()[:10]  # Last 10 changes
    
    # Add current subgroup statistics to context
    context = {
        'form': form,
        'checklist': checklist,
        'subgroup': subgroup,
        'edit_history': edit_history,
        'subgroup_stats': {
            'completion_percentage': subgroup.completion_percentage,
            'total_readings': subgroup.total_readings_count,
            'uv_vacuum_average': subgroup.uv_vacuum_average if subgroup.uv_vacuum_average > 0 else None,
            'uv_flow_average': subgroup.uv_flow_average if subgroup.uv_flow_average > 0 else None,
            'umbrella_ok_count': subgroup.umbrella_valve_ok_count,
            'uv_clip_ok_count': subgroup.uv_clip_ok_count,
            'workstation_status': subgroup.workstation_status,
            'bin_yes_count': subgroup.bin_contamination_yes_count,
        }
    }
    
    return render(request, 'main/edit_subgroup.html', context)                
                
                
                
@login_required
@user_passes_test(lambda u: u.user_type == 'shift_supervisor')
def supervisor_dashboard(request):
    # Get current date/time information
    current_datetime = timezone.localtime(timezone.now())
    current_date = current_datetime.date()
    current_time = current_datetime.time()
    
    # Updated shift determination to match your new shift system
    def get_current_shift_type():
        current_hour = current_datetime.hour
        current_minute = current_datetime.minute
        current_time_minutes = current_hour * 60 + current_minute
        
        if (6 * 60 + 30) <= current_time_minutes < (15 * 60):  # 6:30 AM to 3:00 PM
            return 'A'
        elif (8 * 60 + 30) <= current_time_minutes < (17 * 60):  # 8:30 AM to 5:00 PM
            return 'G'
        elif (15 * 60) <= current_time_minutes < (23 * 60 + 30):  # 3:00 PM to 11:30 PM
            return 'B'
        elif current_time_minutes >= (23 * 60 + 30) or current_time_minutes < (6 * 60 + 30):  # 11:30 PM to 6:30 AM
            return 'C'
        else:
            if 6 <= current_hour < 18:  # 6 AM to 6 PM
                return 'S1'
            else:  # 6 PM to 6 AM
                return 'S2'
    
    current_shift_type = get_current_shift_type()
    
    # Get shift choices for display
    SHIFT_CHOICES = [
        ('S1', 'S1 - 6:30 AM to 6:30 PM'),
        ('A', 'A - 6:30 AM to 3:00 PM'),
        ('G', 'G - 8:30 AM to 5:00 PM'),
        ('B', 'B - 3:00 PM to 11:30 PM'),
        ('C', 'C - 11:30 PM to 6:30 AM'),
        ('S2', 'S2 - 6:30 PM to 6:30 AM'),
    ]
    current_shift_display = dict(SHIFT_CHOICES).get(current_shift_type, current_shift_type)
    
    # Get all pending subgroups that need verification
    pending_verifications = SubgroupEntry.objects.filter(
        verifications__isnull=True,  # No verification yet
        checklist__status='pending'  # Only check active checklists
    ).select_related(
        'checklist',
        'checklist__verification_status',
        'checklist__verification_status__shift'
    ).order_by('-timestamp')[:10]  # Most recent first
    
    # Get recent verifications by supervisors with better error handling
    try:
        all_verified_entries = SubgroupVerification.objects.filter(
            verifier_type__in=['supervisor', 'quality']  # Get both types
        ).select_related(
            'subgroup',
            'verified_by',
            'subgroup__checklist',
            'subgroup__checklist__verification_status'
        ).order_by('-verified_at')
        
        # Count verifications for today before taking a slice
        verified_today_count = all_verified_entries.filter(verified_at__date=current_date).count()
        
        # Now take the slice for display
        verified_entries = all_verified_entries[:10]
        
    except Exception as e:
        # Handle case where no verifications exist
        verified_entries = []
        verified_today_count = 0
    
    # Get recent FTQ records with model and shift info
    recent_ftq_records = FTQRecord.objects.select_related(
        'verification_status',
        'verification_status__shift'
    ).order_by('-date', '-created_at')[:5]
    
    # Get today's FTQ record if it exists
    today_ftq = FTQRecord.objects.filter(
        date=current_date
    ).select_related('verification_status').first()
    
    # Calculate FTQ percentage for display
    ftq_percentage = 0
    if today_ftq and today_ftq.total_inspected > 0:
        ftq_percentage = ((today_ftq.total_inspected - today_ftq.total_defects) / today_ftq.total_inspected) * 100
    
    # Get DTPM checklists with auto-populated model info
    recent_dtpm_checklists = DTPMChecklistFMA03New.objects.select_related(
        'verification_status',
        'shift'
    ).order_by('-date', '-created_at')[:5]
    
    # Get today's DTPM checklist if it exists
    today_dtpm_checklist = DTPMChecklistFMA03New.objects.filter(
        date=current_date
    ).select_related('verification_status').first()
    
    # DTPM Checklist status summary (count checkpoint statuses by OK/NG)
    dtpm_checklist_stats = {
        'ok_count': 0,
        'ng_count': 0,
        'total_checkpoints': 0,
        'current_model': None,
        'shift': None
    }
    
    if today_dtpm_checklist:
        # Get checkpoint results for this DTPM checklist
        checkpoint_results = today_dtpm_checklist.check_results.all()
        dtpm_checklist_stats['total_checkpoints'] = checkpoint_results.count()
        dtpm_checklist_stats['ok_count'] = checkpoint_results.filter(status='OK').count()
        dtpm_checklist_stats['ng_count'] = checkpoint_results.filter(status='NG').count()
        dtpm_checklist_stats['current_model'] = today_dtpm_checklist.current_model
        dtpm_checklist_stats['shift'] = today_dtpm_checklist.checklist_shift
    
    # Get EP checks with model and shift info
    recent_ep_checks = ErrorPreventionCheck.objects.select_related(
        'verification_status',
        'verification_status__shift'
    ).order_by('-date', '-created_at')[:5]
    
    # Get today's EP check if it exists
    today_ep_check = ErrorPreventionCheck.objects.filter(
        date=current_date
    ).select_related('verification_status').first()
    
    # EP Check status summary
    ep_check_stats = {
        'ok_count': 0,
        'ng_count': 0,
        'na_count': 0,
        'total_mechanisms': 0,
        'current_model': None,
        'shift': None
    }
    
    if today_ep_check:
        # Get mechanism statuses for this EP check
        mechanism_statuses = today_ep_check.mechanism_statuses.all()
        ep_check_stats['total_mechanisms'] = mechanism_statuses.count()
        ep_check_stats['ok_count'] = mechanism_statuses.filter(status='OK').count()
        ep_check_stats['ng_count'] = mechanism_statuses.filter(status='NG').count()
        ep_check_stats['na_count'] = mechanism_statuses.filter(is_not_applicable=True).count()
        ep_check_stats['current_model'] = today_ep_check.current_model
        ep_check_stats['shift'] = today_ep_check.shift
    
    # Get recent daily verification statuses for overview
    recent_verifications = DailyVerificationStatus.objects.filter(
        date__gte=current_date - timedelta(days=7)
    ).select_related(
        'shift',
        'created_by'
    ).order_by('-date', '-created_at')[:10]
    
    # Workflow completion summary for today
    today_workflows = DailyVerificationStatus.objects.filter(
        date=current_date
    ).select_related('shift')
    
    workflow_summary = {
        'total_workflows': today_workflows.count(),
        'completed_workflows': 0,
        'pending_workflows': 0,
        'models_in_use': set(),
        'shifts_active': set()
    }
    
    for workflow in today_workflows:
        if workflow.workflow_completion_status['is_complete']:
            workflow_summary['completed_workflows'] += 1
        else:
            workflow_summary['pending_workflows'] += 1
        
        # Get model from checklist if available
        if workflow.current_model_from_checklist:
            workflow_summary['models_in_use'].add(workflow.current_model_from_checklist)
        
        workflow_summary['shifts_active'].add(workflow.shift.shift_type)
    
    # Convert sets to lists for template
    workflow_summary['models_in_use'] = list(workflow_summary['models_in_use'])
    workflow_summary['shifts_active'] = list(workflow_summary['shifts_active'])
    
    context = {
        'current_date': current_date,
        'current_time': current_time,
        'current_shift': current_shift_display,
        'current_shift_type': current_shift_type,
        'pending_verifications': pending_verifications,
        'verified_entries': verified_entries,
        'verification_summary': {
            'pending_count': pending_verifications.count(),
            'verified_today': verified_today_count
        },
        'recent_ftq_records': recent_ftq_records,
        'today_ftq': today_ftq,
        'ftq_percentage': ftq_percentage or 0,
        'recent_dtpm_checklists': recent_dtpm_checklists,
        'today_dtpm_checklist': today_dtpm_checklist,
        'dtpm_checklist_stats': dtpm_checklist_stats,
        'recent_ep_checks': recent_ep_checks,
        'today_ep_check': today_ep_check,
        'ep_check_stats': ep_check_stats,
        'recent_verifications': recent_verifications,
        'workflow_summary': workflow_summary
    }
    
    return render(request, 'main/supervisor_dashboard.html', context)

def calculate_average_verification_time(entries):
    """Calculate average time taken for verification"""
    verified_entries = entries.filter(status='supervisor_approved')
    if not verified_entries:
        return 0
        
    total_time = timedelta()
    count = 0
    
    for entry in verified_entries:
        if entry.created_at and entry.supervisor_verified_at:
            time_diff = entry.supervisor_verified_at - entry.created_at
            total_time += time_diff
            count += 1
            
    if count == 0:
        return 0
        
    average_seconds = total_time.total_seconds() / count
    return round(average_seconds / 60)  # Return in minutes

def calculate_completion_rate(entries):
    """Calculate the rate of checklists that have all 6 subgroups"""
    total = entries.count()
    if total == 0:
        return 0
        
    completed = entries.annotate(
        subgroup_count=Count('subgroup_entries')
    ).filter(subgroup_count=6).count()
    
    return round((completed / total) * 100)

@login_required
@user_passes_test(lambda u: u.user_type == 'shift_supervisor')
def verify_checklist(request, checklist_id):
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    
    if request.method == 'POST':
        form = VerificationForm(request.POST)
        if form.is_valid():
            verification = form.save(commit=False)
            verification.checklist = checklist
            verification.supervisor = request.user
            verification.save()
            
            checklist.status = 'supervisor_approved'
            checklist.save()
            
            messages.success(request, 'Checklist verified successfully')
            return redirect('supervisor_dashboard')
    else:
        form = VerificationForm()
    
    return render(request, 'main/verify_checklist.html', {
        'form': form,
        'checklist': checklist
    })


def calculate_model_stats(entries):
    """Calculate statistics for each model"""
    model_stats = []
    for model_code, model_name in ChecklistBase.MODEL_CHOICES:
        model_entries = entries.filter(selected_model=model_code)
        model_total = model_entries.count()
        
        if model_total > 0:
            stats = {
                'model': model_code,
                'total': model_total,
                'approved': model_entries.filter(status='quality_approved').count(),
                'rejected': model_entries.filter(status='rejected').count()
            }
            stats['rate'] = round((stats['approved'] / model_total) * 100)
            model_stats.append(stats)
    
    return model_stats

def calculate_trend_data(current_date):
    """Calculate 7-day trend data"""
    trend_dates = []
    trend_rates = []
    
    for i in range(7, -1, -1):
        date = current_date - timedelta(days=i)
        entries = ChecklistBase.objects.filter(
            shift__date=date,
            status__in=['quality_approved', 'rejected']
        )
        total = entries.count()
        if total > 0:
            approved = entries.filter(status='quality_approved').count()
            rate = round((approved / total) * 100)
        else:
            rate = 0
        trend_dates.append(date.strftime('%Y-%m-%d'))
        trend_rates.append(rate)
    
    return {
        'dates': trend_dates,
        'rates': trend_rates
    }

def calculate_average_verification_time(verifications):
    """Calculate average time taken for quality verification"""
    total_time = timedelta()
    count = 0
    
    for verification in verifications:
        if hasattr(verification, 'verifications'):
            for v in verification.verifications.all():
                if hasattr(v, 'supervisor_verified_at') and v.quality_supervisor_verified_at:
                    time_diff = v.quality_supervisor_verified_at - v.supervisor_verified_at
                    total_time += time_diff
                    count += 1
            
    if count == 0:
        return 0
        
    return round(total_time.total_seconds() / count / 60)  # Return in minutes

def calculate_quality_stats(entries):
    """Calculate quality statistics"""
    stats = {
        'approved': entries.filter(status='quality_approved').count(),
        'rejected': entries.filter(status='rejected').count(),
        'pending': entries.filter(status__in=['pending', 'supervisor_approved']).count()
    }
    
    total_processed = stats['approved'] + stats['rejected']
    stats['approval_rate'] = round((stats['approved'] / total_processed * 100) if total_processed > 0 else 0)
    
    return stats

def process_measurements(checklists):
    """Process and validate measurements for a list of checklists"""
    processed_entries = []
    
    for checklist in checklists:
        entry_data = {
            'id': checklist.id,
            'created_at': checklist.created_at,
            'shift': checklist.shift,
            'selected_model': checklist.selected_model,
            'subgroup_count': checklist.subgroup_entries.count(),
            'all_measurements_ok': True,
            'measurement_issues': [],
            'critical_issues': []
        }

        # Process subgroups
        for subgroup in checklist.subgroup_entries.all():
            # Validate UV Vacuum Test
            if not (-43 <= subgroup.uv_vacuum_test <= -35):
                entry_data['measurement_issues'].append(
                    f"Subgroup {subgroup.subgroup_number}: UV vacuum test out of range ({subgroup.uv_vacuum_test})"
                )
                entry_data['all_measurements_ok'] = False

            # Validate UV Flow Value
            if not (30 <= subgroup.uv_flow_value <= 40):
                entry_data['measurement_issues'].append(
                    f"Subgroup {subgroup.subgroup_number}: UV flow value out of range ({subgroup.uv_flow_value})"
                )
                entry_data['all_measurements_ok'] = False

            # Check other critical values
            if not ((4.5 <= checklist.line_pressure <= 5.5) and 
                   (11 <= checklist.uv_flow_input_pressure <= 15) and
                   (0.25 <= checklist.test_pressure_vacuum <= 0.3)):
                entry_data['critical_issues'].append({
                    'severity': 'critical',
                    'message': 'Critical measurements out of range',
                    'measurements': {
                        'line_pressure': checklist.line_pressure,
                        'uv_flow_input_pressure': checklist.uv_flow_input_pressure,
                        'test_pressure_vacuum': checklist.test_pressure_vacuum
                    }
                })

        processed_entries.append(entry_data)

    return processed_entries

def validate_base_measurements(entry):
    """Validate base measurements"""
    measurements = {
        'line_pressure_ok': False,
        'uv_flow_input_ok': False,
        'test_pressure_ok': False
    }
    
    try:
        # Line Pressure validation
        if hasattr(entry, 'line_pressure'):
            line_pressure = float(entry.line_pressure)
            measurements['line_pressure_ok'] = 4.5 <= line_pressure <= 5.5
            if not measurements['line_pressure_ok']:
                entry.critical_issues.append(f"Line pressure critical: {line_pressure}")
        
        # UV Flow Input validation
        if hasattr(entry, 'uv_flow_input_pressure'):
            uv_pressure = float(entry.uv_flow_input_pressure)
            measurements['uv_flow_input_ok'] = 11 <= uv_pressure <= 15
            if not measurements['uv_flow_input_ok']:
                entry.critical_issues.append(f"UV flow input pressure critical: {uv_pressure}")
        
        # Test Pressure validation
        if hasattr(entry, 'test_pressure_vacuum'):
            test_pressure = float(entry.test_pressure_vacuum)
            measurements['test_pressure_ok'] = 0.25 <= test_pressure <= 0.3
            if not measurements['test_pressure_ok']:
                entry.critical_issues.append(f"Test pressure critical: {test_pressure}")
    except (ValueError, TypeError):
        entry.critical_issues.append("Invalid measurement values")
    
    return measurements

def validate_subgroup_measurements(entry):
    """Validate subgroup measurements"""
    subgroup_validations = []
    
    try:
        for subgroup in entry.subgroup_entries.all():
            validation = {
                'subgroup_number': subgroup.subgroup_number,
                'uv_vacuum_test_ok': False,
                'uv_flow_value_ok': False,
                'all_ok': True
            }
            
            # UV Vacuum Test validation
            try:
                uv_vacuum = float(subgroup.uv_vacuum_test)
                validation['uv_vacuum_test_ok'] = -43 <= uv_vacuum <= -35
                if not validation['uv_vacuum_test_ok']:
                    entry.measurement_issues.append(
                        f"Subgroup {subgroup.subgroup_number}: UV vacuum test out of range ({uv_vacuum})"
                    )
            except (ValueError, TypeError):
                validation['all_ok'] = False
            
            # UV Flow Value validation
            try:
                uv_flow = float(subgroup.uv_flow_value)
                validation['uv_flow_value_ok'] = 30 <= uv_flow <= 40
                if not validation['uv_flow_value_ok']:
                    entry.measurement_issues.append(
                        f"Subgroup {subgroup.subgroup_number}: UV flow value out of range ({uv_flow})"
                    )
            except (ValueError, TypeError):
                validation['all_ok'] = False
            
            subgroup_validations.append(validation)
    except Exception as e:
        entry.measurement_issues.append(f"Error processing subgroups: {str(e)}")
    
    return subgroup_validations

def get_critical_issues(in_progress, pending):
    """Get critical issues from both in-progress and pending checklists"""
    critical_issues = []
    
    for entries in [in_progress, pending]:
        for entry in entries:
            if entry['critical_issues']:
                critical_issues.append({
                    'checklist_id': entry['id'],
                    'operator': entry['original_entry'].shift.operator.username,
                    'model': entry['original_entry'].selected_model,
                    'issues': entry['critical_issues']
                })
    

@login_required
@user_passes_test(lambda u: u.user_type == 'quality_supervisor')
def quality_dashboard(request):
    current_datetime = timezone.now()
    current_date = current_datetime.date()
    current_time = current_datetime.time()
    is_day_shift = (8 <= current_time.hour < 20)
    current_shift = 'day' if is_day_shift else 'night'
    current_shift_display = 'Day Shift (8 AM - 8 PM)' if is_day_shift else 'Night Shift (8 PM - 8 AM)'

    # Get all subgroups that need quality verification (where supervisor has verified)
    pending_subgroups = SubgroupEntry.objects.filter(
        verification_status='supervisor_verified',
        # Include entries from the last 7 days
        checklist__created_at__date__gte=current_date - timedelta(days=7),
        checklist__created_at__date__lte=current_date
    ).select_related(
        'checklist'
    ).prefetch_related(
        'verifications'
    ).order_by('-timestamp')

    # Get recent quality verifications for today and yesterday
    recent_verifications = SubgroupVerification.objects.filter(
        verifier_type='quality',
        verified_at__date__gte=current_date - timedelta(days=1)
    ).select_related(
        'verified_by',
        'subgroup__checklist'
    ).order_by('-verified_at')

    # Calculate stats
    verification_stats = {
        'pending_count': pending_subgroups.count(),
        'approved_today': recent_verifications.filter(
            status='quality_verified',
            verified_at__date=current_date
        ).count(),
        'rejected_today': recent_verifications.filter(
            status='rejected',
            verified_at__date=current_date
        ).count()
    }

    # Get FTQ data (similar to supervisor dashboard)
    today_ftq = None
    ftq_percentage = 0
    recent_ftq_records = []
    
    try:
        today_ftq = FTQRecord.objects.filter(
            date=current_date
        ).first()
        
        if today_ftq and today_ftq.total_inspected > 0:
            ftq_percentage = ((today_ftq.total_inspected - today_ftq.total_defects) / today_ftq.total_inspected) * 100
        
        recent_ftq_records = FTQRecord.objects.filter(
            date__gte=current_date - timedelta(days=7)
        ).order_by('-date', '-id')[:5]
    except:
        # Handle case where FTQ model doesn't exist or has different structure
        pass

    # Get EP Check data - FIXED VERSION
    today_ep_check = None
    ep_check_stats = {'ok_count': 0, 'ng_count': 0, 'na_count': 0}
    recent_ep_checks = []
    
    try:
        today_ep_check = ErrorPreventionCheck.objects.filter(
            date=current_date
        ).select_related('verification_status').first()
        
        if today_ep_check:
            # Get mechanism statuses for this EP check - CORRECT FIELD NAMES
            mechanism_statuses = ErrorPreventionMechanismStatus.objects.filter(
                ep_check=today_ep_check
            )
            
            # Calculate EP stats based on the actual model structure
            ep_check_stats = {
                'ok_count': mechanism_statuses.filter(status='OK').count(),
                'ng_count': mechanism_statuses.filter(status='NG').count(),
                'na_count': mechanism_statuses.filter(is_not_applicable=True).count(),
            }
        
        recent_ep_checks = ErrorPreventionCheck.objects.filter(
            date__gte=current_date - timedelta(days=7)
        ).order_by('-date')[:5]
    except Exception as e:
        print(f"EP Check error: {e}")
        pass

    # Get DTPM data - FIXED VERSION
    today_dtpm_checklist = None
    dtpm_checklist_stats = {'ok_count': 0, 'ng_count': 0, 'total_checkpoints': 0}
    recent_dtpm_checklists = []
    
    try:
        today_dtpm_checklist = DTPMChecklistFMA03New.objects.filter(
            date=current_date
        ).select_related('verification_status').first()
        
        if today_dtpm_checklist:
            # Get checkpoint results for this DTPM checklist - CORRECT FIELD NAMES
            checkpoint_results = DTPMCheckResultNew.objects.filter(
                checklist=today_dtpm_checklist
            )
            
            # Calculate DTPM stats based on the actual model structure
            dtpm_checklist_stats = {
                'ok_count': checkpoint_results.filter(status='OK').count(),
                'ng_count': checkpoint_results.filter(status='NG').count(),
                'total_checkpoints': checkpoint_results.count(),
            }
        
        recent_dtpm_checklists = DTPMChecklistFMA03New.objects.filter(
            date__gte=current_date - timedelta(days=7)
        ).order_by('-date')[:5]
    except Exception as e:
        print(f"DTPM error: {e}")
        pass

    context = {
        'current_date': current_date,
        'current_time': current_time,
        'current_shift': current_shift_display,
        'pending_verifications': pending_subgroups,
        'verified_entries': recent_verifications,
        'verification_stats': verification_stats,
        'verification_summary': verification_stats,  # For compatibility with template
        
        # FTQ data
        'today_ftq': today_ftq,
        'ftq_percentage': ftq_percentage,
        'recent_ftq_records': recent_ftq_records,
        
        # EP Check data
        'today_ep_check': today_ep_check,
        'ep_check_stats': ep_check_stats,
        'recent_ep_checks': recent_ep_checks,
        
        # DTPM data
        'today_dtpm_checklist': today_dtpm_checklist,
        'dtpm_checklist_stats': dtpm_checklist_stats,
        'recent_dtpm_checklists': recent_dtpm_checklists,
    }
    
    return render(request, 'main/quality_dashboard.html', context)

@login_required
def reports_dashboard(request):
    current_date = timezone.now().date()
    start_of_week = current_date - timedelta(days=current_date.weekday())
    start_of_month = current_date.replace(day=1)

    # Get statistics
    daily_stats = ChecklistBase.objects.filter(
        shift__date=current_date
    ).aggregate(
        total=Count('id'),
        approved=Count('id', filter=Q(status='quality_approved')),
        rejected=Count('id', filter=Q(status='rejected')),
        pending=Count('id', filter=Q(status='pending'))
    )

    weekly_stats = ChecklistBase.objects.filter(
        shift__date__gte=start_of_week,
        shift__date__lte=current_date
    ).aggregate(
        total=Count('id'),
        approved=Count('id', filter=Q(status='quality_approved')),
        rejected=Count('id', filter=Q(status='rejected'))
    )

    monthly_stats = ChecklistBase.objects.filter(
        shift__date__gte=start_of_month,
        shift__date__lte=current_date
    ).aggregate(
        total=Count('id'),
        approved=Count('id', filter=Q(status='quality_approved')),
        rejected=Count('id', filter=Q(status='rejected'))
    )

    # Get model-wise statistics
    model_stats = ChecklistBase.objects.values('selected_model').annotate(
        total=Count('id'),
        approved=Count('id', filter=Q(status='quality_approved')),
        rejected=Count('id', filter=Q(status='rejected'))
    )

    # Calculate approval rates
    for stats in [daily_stats, weekly_stats, monthly_stats]:
        total_verified = stats['approved'] + stats['rejected']
        stats['approval_rate'] = (
            round((stats['approved'] / total_verified) * 100)
            if total_verified > 0 else 0
        )

    # Get recent activity
    recent_activity = ChecklistBase.objects.filter(
        status__in=['quality_approved', 'rejected']
    ).order_by('-created_at')[:10]

    context = {
        'daily_stats': daily_stats,
        'weekly_stats': weekly_stats,
        'monthly_stats': monthly_stats,
        'model_stats': model_stats,
        'recent_activity': recent_activity,
        'current_date': current_date
    }

    return render(request, 'main/reports/reports_dashboard.html', context)


from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def export_checklist_excel(request, checklist_id):
    import os
    from django.conf import settings
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    import logging
    
    # Set up logging to debug missing values
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    
    # Add stream handler if not already present
    if not logger.handlers:
        handler = logging.StreamHandler()
        logger.addHandler(handler)
    
    checklist = get_object_or_404(
        ChecklistBase.objects.select_related(
            'shift__operator',
            'shift__shift_supervisor',
            'shift__quality_supervisor'
        ).prefetch_related(
            'subgroup_entries__verifications',
        ), 
        id=checklist_id
    )
    
    # Path to your template file - adjust this to the actual location
    template_path = os.path.join(settings.STATIC_ROOT, 'templates', 'new demo.xlsx')
    
    # Load the template workbook instead of creating a new one
    wb = load_workbook(template_path)
    
    # Clean external links to prevent validation warnings
    if hasattr(wb, 'external_links') and wb.external_links:
        wb.external_links = []
        logger.info("Removed external links from workbook")
    
    ws = wb.active  # Or use wb["Sheet Name"] if you know the exact sheet name
    
    # Define styles for pass/fail values
    pass_style = {
        'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'font': Font(color='006100'),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    
    fail_style = {
        'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'font': Font(color='9C0006'),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    
    # Helper function to safely set cell value with additional debugging
    def safe_set_cell_value(ws, coord, value, field_name=None):
        try:
            cell = ws[coord]
            original_value = cell.value
            if isinstance(cell, MergedCell):
                # Find the merged range that contains this cell
                for merged_range in ws.merged_cells.ranges:
                    if coord in merged_range:
                        # Set the value on the top-left cell of the merged range
                        anchor_cell = merged_range.start_cell
                        anchor_coord = anchor_cell.coordinate
                        original_value = ws[anchor_coord].value  # Store original for debug
                        ws[anchor_coord] = value
                        logger.info(f"SET MERGED CELL {anchor_coord} (from {coord}) for {field_name}: '{original_value}' → '{value}'")
                        return anchor_cell
                logger.warning(f"MERGED CELL {coord} for {field_name} not found in any range")
                return None
            else:
                cell.value = value
                logger.info(f"SET CELL {coord} for {field_name}: '{original_value}' → '{value}'")
                return cell
        except Exception as e:
            logger.error(f"ERROR SETTING {coord} for {field_name}: {e}")
            return None
    
    # Get ordered subgroups
    subgroups = list(checklist.subgroup_entries.all().order_by('subgroup_number'))
    logger.info(f"SUBGROUPS: {len(subgroups)}")
    
    # HEADER INFO
    safe_set_cell_value(ws, 'B5', checklist.shift.date.strftime('%Y-%m-%d'), "Date")
    safe_set_cell_value(ws, 'E5', checklist.shift.get_shift_type_display(), "Shift")
    
    # RECORD TIME - Row 16
    for sg_index, sg in enumerate(subgroups):
        col_start = 2 + sg_index * 5  # Adjust column calculation (0-based index)
        time_str = sg.timestamp.strftime('%H:%M')
        for i in range(5):
            col = get_column_letter(col_start + i)
            safe_set_cell_value(ws, f'{col}16', time_str, "Record Time")
    
    # DIRECT ROW ASSIGNMENTS FOR ALL FIELDS
    
    # MACHINE SETTINGS section - single values
    row_mapping = {
        'Program selection on HMI': {'row': 20, 'field': 'selected_model', 'pass_fail': False},
        'Line pressure': {'row': 21, 'field': 'line_pressure', 'pass_fail': False},
        'O-ring conditon': {'row': 22, 'field': 'oring_condition', 'pass_fail': True},
        'UV Flow input Test Pressure': {'row': 23, 'field': 'uv_flow_input_pressure', 'pass_fail': False},
        'Master Verification for LVDT': {'row': 26, 'field': 'master_verification_lvdt', 'pass_fail': True},
        'Good and Bad master verification': {'row': 27, 'field': 'good_bad_master_verification', 'pass_fail': True},
        'Test Pressure for Vacumm generation': {'row': 28, 'field': 'test_pressure_vacuum', 'pass_fail': False},
        'Tool Alignmnet': {'row': 30, 'field': 'tool_alignment', 'pass_fail': True},
        'Top Tool': {'row': 31, 'field': 'top_tool_id', 'pass_fail': False},
        'Bottom Tool': {'row': 32, 'field': 'bottom_tool_id', 'pass_fail': False},
        'UV Assy Stage': {'row': 33, 'field': 'uv_assy_stage_id', 'pass_fail': False},
        'Retainer Part no': {'row': 35, 'field': 'retainer_part_no', 'pass_fail': False},
        'UV Clip Part No': {'row': 36, 'field': 'uv_clip_part_no', 'pass_fail': False},
        'Umbrella Part No': {'row': 37, 'field': 'umbrella_part_no', 'pass_fail': False},
        'Retainer ID lubrication': {'row': 38, 'field': 'retainer_id_lubrication', 'pass_fail': True},
    }
    
    # Set all single-value fields with pass/fail styling where needed
    for field_name, config in row_mapping.items():
        row = config['row']
        field = config['field']
        is_pass_fail = config['pass_fail']
        
        value = getattr(checklist, field, 'N/A')
        
        # Special handling for None or null values
        if value is None:
            value = 'N/A'
        
        # Set the value
        cell = safe_set_cell_value(ws, f'B{row}', value, field_name)
        
        # Apply pass/fail styling if needed
        if cell and is_pass_fail and value in ['OK', 'NG']:
            try:
                style = pass_style if value == 'OK' else fail_style
                cell.fill = style['fill']
                cell.font = style['font']
                cell.alignment = style['alignment']
            except Exception as e:
                logger.warning(f"Error styling {field_name}: {e}")
    
    # MEASUREMENT FIELDS - per subgroup with validation
    measurement_fields = {
        'UV Vacuum Test': {'row': 24, 'field': 'uv_vacuum_test', 'validation': lambda x: -43 <= x <= -35},
        'UV Flow Value': {'row': 25, 'field': 'uv_flow_value', 'validation': lambda x: 30 <= x <= 40},
    }
    
    for field_name, config in measurement_fields.items():
        row = config['row']
        field = config['field']
        validation = config['validation']
        
        for sg_index, sg in enumerate(subgroups):
            col_start = 2 + sg_index * 5
            value = getattr(sg, field, 0)
            
            # Ensure value is numeric
            if value is None:
                value = 0
                
            for i in range(6):
                col = get_column_letter(col_start + i)
                cell = safe_set_cell_value(ws, f'{col}{row}', value, f"{field_name} sample {i+1}")
                
                if cell:
                    try:
                        style = pass_style if validation(value) else fail_style
                        cell.fill = style['fill']
                        cell.font = style['font']
                        cell.alignment = style['alignment']
                    except Exception as e:
                        logger.warning(f"Error styling {field_name}: {e}")
    
    # REPEATED CHECKS - per subgroup
    repeated_fields = {
        'Umbrella Valve Assembly in Retainer': {'row': 29, 'field': 'umbrella_valve_assembly'},
        'UV Clip pressing': {'row': 39, 'field': 'uv_clip_pressing'},
        'All workstations are clean': {'row': 41, 'field': 'workstation_clean'},
        'All Error proofing verification': {'row': 42, 'field': 'error_proofing_verification'},
        'Station Operator confirmation': {'row': 43, 'field': 'bin_contamination_check'},
    }
    
    for field_name, config in repeated_fields.items():
        row = config['row']
        field = config['field']
        
        for sg_index, sg in enumerate(subgroups):
            col_start = 2 + sg_index * 6
            value = getattr(sg, field, 'N/A')
            
            # Handle None values
            if value is None:
                value = 'N/A'
                
            for i in range(5):
                col = get_column_letter(col_start + i)
                cell = safe_set_cell_value(ws, f'{col}{row}', value, f"{field_name} sample {i+1}")
                
                if cell:
                    try:
                        style = pass_style if value in ['OK', 'Yes'] else fail_style
                        cell.fill = style['fill']
                        cell.font = style['font']
                        cell.alignment = style['alignment']
                    except Exception as e:
                        logger.warning(f"Error styling {field_name}: {e}")
    
    # SIGNATURES
    signature_rows = {
        'Team Leader': {'row': 44, 'field': 'operator'},
        'Shift Supervisor': {'row': 45, 'field': 'shift_supervisor'},
        'Quality Supervisor': {'row': 46, 'field': 'quality_supervisor'},
    }
    
    for field_name, config in signature_rows.items():
        row = config['row']
        field = config['field']
        
        for sg_index, sg in enumerate(subgroups):
            col_start = 2 + sg_index * 6
            role_obj = getattr(checklist.shift, field, None)
            
            # Get name with fallbacks
            name = 'N/A'
            if role_obj:
                try:
                    name = role_obj.get_full_name()
                    if not name or name.strip() == '':
                        name = role_obj.username
                except:
                    name = getattr(role_obj, 'username', 'N/A')
            
            for i in range(5):
                col = get_column_letter(col_start + i)
                safe_set_cell_value(ws, f'{col}{row}', name, f"{field_name} signature")
    
    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=quality_control_{checklist.id}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    try:
        # Log all changes before saving
        logger.info("SAVING WORKBOOK TO RESPONSE")
        wb.save(response)
        logger.info("WORKBOOK SAVED SUCCESSFULLY")
    except Exception as e:
        logger.error(f"ERROR SAVING WORKBOOK: {e}")
        # Fallback to a simpler workbook if the template has issues
        from openpyxl import Workbook
        fallback_wb = Workbook()
        fallback_ws = fallback_wb.active
        fallback_ws['A1'] = "Error generating Excel from template"
        fallback_ws['A2'] = f"Please check logs: {str(e)}"
        fallback_wb.save(response)
    
    return response


from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from datetime import timedelta
from django.db.models import Avg, Count
from .models import ChecklistBase, SubgroupEntry, Verification, Concern

# views.py - Updated to handle 5 readings per parameter

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import timedelta
from django.contrib import messages
from .models import ChecklistBase, SubgroupEntry, SubgroupVerification

def checklist_detail(request, checklist_id):   
    checklist = get_object_or_404(
        ChecklistBase.objects.select_related(
            'verification_status__shift__operator',
            'verification_status__shift__shift_supervisor',
            'verification_status__shift__quality_supervisor',
            'verification_status__created_by',
            'frequency_config'
        ).prefetch_related(
            'subgroup_entries__verifications',
            'subgroup_entries__edit_history__edited_by',
            'verifications',
            'concern_set',
            'dynamic_values__parameter'  # NEW: Fetch dynamic values
        ), 
        id=checklist_id
    )
    
    # Get subgroups with validation status
    subgroups = process_subgroups(checklist.subgroup_entries.all().order_by('subgroup_number'))
    
    # Get frequency config
    frequency_hours = 2
    max_subgroups = 6
    
    if checklist.frequency_config and checklist.frequency_config.is_active:
        frequency_hours = checklist.frequency_config.frequency_hours
        max_subgroups = checklist.frequency_config.max_subgroups
    else:
        try:
            from .models import SubgroupFrequencyConfig
            config = SubgroupFrequencyConfig.objects.get(
                model_name=checklist.selected_model,
                is_active=True
            )
            frequency_hours = config.frequency_hours
            max_subgroups = config.max_subgroups
        except SubgroupFrequencyConfig.DoesNotExist:
            pass
    
    shift_max_subgroups = max_subgroups
    
    # Get maintenance and NOK subgroups
    maintenance_subgroups = [sg for sg in subgroups if getattr(sg, 'is_after_maintenance', False)]
    nok_approval_subgroups = [sg for sg in subgroups if not sg.all_checks_passed]
    
    # Get timing information
    timing_info = {}
    can_add_subgroup = False
    if checklist.verification_status and checklist.verification_status.shift:
        can_add_subgroup, next_allowed_time, available_slots = check_time_gap_shift_based(checklist)
        
        current_date = timezone.localtime(timezone.now()).date()
        expected_times = get_expected_subgroup_times(
            checklist.verification_status.shift.shift_type, 
            current_date,
            checklist
        )
        
        schedule_info = []
        current_time = timezone.localtime(timezone.now())
        
        for i, expected_time in enumerate(expected_times, 1):
            schedule_info.append({
                'subgroup_number': i,
                'expected_time': expected_time,
                'is_available': current_time >= expected_time,
                'is_completed': any(sg.subgroup_number == i for sg in subgroups),
                'time_display': timezone.localtime(expected_time).strftime('%H:%M')
            })
        
        timing_info = {
            'schedule': schedule_info,
            'shift_type': checklist.verification_status.shift.shift_type,
            'shift_display': checklist.verification_status.shift.get_shift_type_display(),
            'max_subgroups': shift_max_subgroups,
            'current_available_slots': available_slots,
            'frequency_hours': frequency_hours
        }
    
    # Calculate subgroup metrics
    subgroup_metrics = {
        'total_count': len(subgroups),
        'passed_count': sum(1 for s in subgroups if getattr(s, 'all_checks_passed', False)),
        'completion_percentage': (len(subgroups) / shift_max_subgroups) * 100 if subgroups and shift_max_subgroups else 0,
        'progress_width': (len(subgroups) / shift_max_subgroups) * 100 if shift_max_subgroups else 0,
        'total_readings_completed': len(subgroups) * 21,
        'total_readings_target': shift_max_subgroups * 21,
        'progress_text': f"{len(subgroups)}/{shift_max_subgroups} Complete ({len(subgroups) * 21}/{shift_max_subgroups * 21} readings)"
    }
    
    # Verification status
    verification_status = {
        'can_verify_supervisor': (
            request.user.user_type == 'shift_supervisor' and 
            checklist.status == 'pending' and 
            len(subgroups) == shift_max_subgroups
        ),
        'can_verify_quality': (
            request.user.user_type == 'quality_supervisor' and 
            checklist.status == 'supervisor_approved'
        ),
        'last_verification': checklist.verifications.last(),
        'verification_count': checklist.verifications.count()
    }
    
    # User permissions
    user_permissions = {
        'can_edit': request.user == checklist.verification_status.created_by and checklist.status == 'pending',
        'can_verify': request.user.user_type in ['shift_supervisor', 'quality_supervisor'],
        'can_add_concern': checklist.status != 'quality_approved',
        'can_view_all': request.user.user_type in ['shift_supervisor', 'quality_supervisor']
    }
    
    # Measurement validation
    measurement_validation = validate_measurements(checklist, subgroups)
    
    # Concerns
    concerns = process_concerns(checklist.concern_set.all().order_by('-created_at'))
    
    # NEW: Get dynamic values ordered by display_order
    dynamic_values = checklist.dynamic_values.select_related('parameter').order_by('parameter__order', 'parameter__id')
    
    context = {
        'checklist': checklist,
        'verification_status': checklist.verification_status,
        'subgroups': subgroups,
        'can_add_subgroup': can_add_subgroup,
        'total_subgroups': len(subgroups),
        'remaining_subgroups': shift_max_subgroups - len(subgroups),
        'subgroup_metrics': subgroup_metrics,
        'verification_status': verification_status,
        'user_permissions': user_permissions,
        'measurement_validation': measurement_validation,
        'concerns': concerns,
        'timing_info': timing_info,
        'shift_max_subgroups': shift_max_subgroups,
        'frequency_hours': frequency_hours,
        'maintenance_subgroups': maintenance_subgroups,
        'nok_approval_subgroups': nok_approval_subgroups,
        'dynamic_values': dynamic_values,  # NEW: Add to context
    }
    
    return render(request, 'main/checklist_detail.html', context)




def process_subgroups(subgroups):
    """Process and validate subgroups - handles None values and comments"""
    processed_subgroups = []
    for subgroup in subgroups:
        # Get readings, filtering out None values
        uv_vacuum_readings = [
            subgroup.uv_vacuum_test_1, subgroup.uv_vacuum_test_2, 
            subgroup.uv_vacuum_test_3, subgroup.uv_vacuum_test_4, 
            subgroup.uv_vacuum_test_5
        ]
        uv_flow_readings = [
            subgroup.uv_flow_value_1, subgroup.uv_flow_value_2, 
            subgroup.uv_flow_value_3, subgroup.uv_flow_value_4, 
            subgroup.uv_flow_value_5
        ]
        
        # Filter out None values
        valid_uv_vacuum = [r for r in uv_vacuum_readings if r is not None]
        valid_uv_flow = [r for r in uv_flow_readings if r is not None]
        
        # Calculate averages (0 if no valid readings)
        uv_vacuum_avg = sum(valid_uv_vacuum) / len(valid_uv_vacuum) if valid_uv_vacuum else 0
        uv_flow_avg = sum(valid_uv_flow) / len(valid_uv_flow) if valid_uv_flow else 0
        
        # Validation based on averages and individual readings
        subgroup.validation_status = {
            'uv_vacuum_test_ok': all(-43 <= reading <= -35 for reading in valid_uv_vacuum) if valid_uv_vacuum else False,
            'uv_flow_value_ok': all(30 <= reading <= 40 for reading in valid_uv_flow) if valid_uv_flow else False,
            'uv_vacuum_avg_ok': -43 <= uv_vacuum_avg <= -35 if uv_vacuum_avg > 0 else False,
            'uv_flow_avg_ok': 30 <= uv_flow_avg <= 40 if uv_flow_avg > 0 else False
        }
        
        # Get OK/NG and Yes/No readings, filtering None values
        umbrella_readings = [
            subgroup.umbrella_valve_assembly_1, subgroup.umbrella_valve_assembly_2,
            subgroup.umbrella_valve_assembly_3, subgroup.umbrella_valve_assembly_4,
            subgroup.umbrella_valve_assembly_5
        ]
        uv_clip_readings = [
            subgroup.uv_clip_pressing_1, subgroup.uv_clip_pressing_2,
            subgroup.uv_clip_pressing_3, subgroup.uv_clip_pressing_4,
            subgroup.uv_clip_pressing_5
        ]
        workstation_value = subgroup.workstation_clean
        
        bin_readings = [
            subgroup.bin_contamination_check_1, subgroup.bin_contamination_check_2,
            subgroup.bin_contamination_check_3, subgroup.bin_contamination_check_4,
            subgroup.bin_contamination_check_5
        ]
        
        # Filter out None values
        valid_umbrella = [r for r in umbrella_readings if r is not None]
        valid_uv_clip = [r for r in uv_clip_readings if r is not None]
        valid_bin = [r for r in bin_readings if r is not None]
        
        # Add averages as attributes for template access
        subgroup.uv_vacuum_avg_calculated = uv_vacuum_avg
        subgroup.uv_flow_avg_calculated = uv_flow_avg
        
        # Overall validation
        subgroup.all_checks_passed = (
            len(valid_uv_vacuum) == 5 and 
            len(valid_uv_flow) == 5 and
            len(valid_umbrella) == 5 and
            len(valid_uv_clip) == 5 and
            workstation_value is not None and
            len(valid_bin) == 5 and
            subgroup.validation_status['uv_vacuum_test_ok'] and
            subgroup.validation_status['uv_flow_value_ok'] and
            valid_umbrella.count('OK') == 5 and
            valid_uv_clip.count('OK') == 5 and
            workstation_value == 'Yes' and
            valid_bin.count('Yes') == 5
        )
        
        # Add verification information
        subgroup.supervisor_verification = SubgroupVerification.objects.filter(
            subgroup=subgroup,
            verifier_type='supervisor'
        ).first()
        
        subgroup.quality_verification = SubgroupVerification.objects.filter(
            subgroup=subgroup,
            verifier_type='quality'
        ).first()
        
        processed_subgroups.append(subgroup)
    return processed_subgroups
 
 
 
def validate_measurements(checklist, subgroups):
    """Validate all measurements including subgroups with single workstation - handles None values"""
    base_measurements = {
        'line_pressure_ok': 4.5 <= checklist.line_pressure <= 5.5,
        'uv_flow_input_ok': 11 <= checklist.uv_flow_input_pressure <= 15,
        'test_pressure_ok': 0.25 <= checklist.test_pressure_vacuum <= 0.3
    }
    
    # Collect all individual reading issues
    all_issues = []
    for subgroup in subgroups:
        # Check each of the 21 readings per subgroup (5+5+5+5+1+5), handling None values
        for i in range(1, 6):
            # Get readings, handling None values
            uv_vacuum = getattr(subgroup, f'uv_vacuum_test_{i}')
            uv_flow = getattr(subgroup, f'uv_flow_value_{i}')
            umbrella = getattr(subgroup, f'umbrella_valve_assembly_{i}')
            clip = getattr(subgroup, f'uv_clip_pressing_{i}')
            bin_check = getattr(subgroup, f'bin_contamination_check_{i}')
            
            # Only check non-None values
            if uv_vacuum is not None and not (-43 <= uv_vacuum <= -35):
                all_issues.append(f'Subgroup {subgroup.subgroup_number} UV Vacuum Reading {i}: {uv_vacuum} kPa out of range')
            if uv_flow is not None and not (30 <= uv_flow <= 40):
                all_issues.append(f'Subgroup {subgroup.subgroup_number} UV Flow Reading {i}: {uv_flow} LPM out of range')
            if umbrella is not None and umbrella != 'OK':
                all_issues.append(f'Subgroup {subgroup.subgroup_number} Umbrella Valve Reading {i}: {umbrella}')
            if clip is not None and clip != 'OK':
                all_issues.append(f'Subgroup {subgroup.subgroup_number} UV Clip Reading {i}: {clip}')
            if bin_check is not None and bin_check != 'Yes':
                all_issues.append(f'Subgroup {subgroup.subgroup_number} Bin Check Reading {i}: {bin_check}')
        
        # Check single workstation cleanliness field
        workstation = getattr(subgroup, 'workstation_clean')
        if workstation is not None and workstation != 'Yes':
            all_issues.append(f'Subgroup {subgroup.subgroup_number} Workstation: {workstation}')
    
    return {
        'base_measurements': base_measurements,
        'all_base_ok': all(base_measurements.values()),
        'subgroup_measurements_ok': all(s.all_checks_passed for s in subgroups),
        'total_issues': len(all_issues),
        'critical_issues': [issue for issue in all_issues if 'out of range' in issue],
        'quality_issues': [issue for issue in all_issues if 'out of range' not in issue]
    }
  






def calculate_subgroup_metrics(subgroups):
    """Calculate metrics for subgroups with 21 readings (single workstation)"""
    total_readings = len(subgroups) * 21  # 21 readings per subgroup (5+5+5+5+1+5)
    
    vacuum_issues = 0
    flow_issues = 0
    umbrella_issues = 0
    clip_issues = 0
    workstation_issues = 0
    bin_issues = 0
    
    for subgroup in subgroups:
        # Count individual reading issues for 5-reading parameters
        for i in range(1, 6):
            # UV Vacuum issues
            uv_vacuum = getattr(subgroup, f'uv_vacuum_test_{i}', None)
            if uv_vacuum is not None and not (-43 <= uv_vacuum <= -35):
                vacuum_issues += 1
                
            # UV Flow issues
            uv_flow = getattr(subgroup, f'uv_flow_value_{i}', None)
            if uv_flow is not None and not (30 <= uv_flow <= 40):
                flow_issues += 1
                
            # Umbrella valve issues
            umbrella = getattr(subgroup, f'umbrella_valve_assembly_{i}', None)
            if umbrella is not None and umbrella != 'OK':
                umbrella_issues += 1
                
            # UV Clip issues
            clip = getattr(subgroup, f'uv_clip_pressing_{i}', None)
            if clip is not None and clip != 'OK':
                clip_issues += 1
                
            # Bin contamination issues
            bin_check = getattr(subgroup, f'bin_contamination_check_{i}', None)
            if bin_check is not None and bin_check != 'Yes':
                bin_issues += 1
        
        # Check single workstation field
        workstation = getattr(subgroup, 'workstation_clean', None)
        if workstation is not None and workstation != 'Yes':
            workstation_issues += 1
    
    return {
        'total_count': len(subgroups),
        'total_readings': total_readings,
        'passed_count': sum(1 for s in subgroups if s.all_checks_passed),
        'vacuum_test_issues': vacuum_issues,
        'flow_value_issues': flow_issues,
        'umbrella_valve_issues': umbrella_issues,
        'uv_clip_issues': clip_issues,
        'workstation_issues': workstation_issues,
        'bin_contamination_issues': bin_issues,
        'completion_percentage': (len(subgroups) / 6) * 100 if subgroups else 0,
        'reading_completion_percentage': (total_readings / 126) * 100 if total_readings else 0  # 126 = 6 * 21
    }

def check_subgroup_addition(checklist, subgroups):
    """Check if new subgroup can be added"""
    if checklist.status != 'pending':
        return False
        
    current_count = len(subgroups)
    if current_count >= 6:
        return False
        
    if not subgroups:
        return True
        
    last_subgroup = subgroups[-1]
    time_since_last = timezone.now() - last_subgroup.timestamp
    return time_since_last >= timedelta(hours=2)

def get_verification_status(user, checklist, subgroups):
    """Get verification status and permissions"""
    return {
        'can_verify_supervisor': (
            user.user_type == 'shift_supervisor' and 
            checklist.status == 'pending' and 
            len(subgroups) == 6
        ),
        'can_verify_quality': (
            user.user_type == 'quality_supervisor' and 
            checklist.status == 'supervisor_approved'
        ),
        'last_verification': checklist.verifications.last(),
        'verification_count': checklist.verifications.count()
    }

def process_concerns(concerns):
    """Process concerns with additional metadata"""
    return [{
        'concern': concern,
        'is_resolved': bool(concern.action_taken),
        'requires_attention': not concern.action_taken
    } for concern in concerns]





def calculate_timing_metrics(checklist, subgroups):
    """Calculate timing related metrics"""
    if not subgroups:
        return None
        
    return {
        'total_duration': (subgroups[-1].timestamp - checklist.created_at).total_seconds() / 3600,
        'average_interval': calculate_average_interval(subgroups),
        'completion_rate': len(subgroups) / ((timezone.now() - checklist.created_at).total_seconds() / 3600)
    }

def get_user_permissions(user, checklist):
    """Get user-specific permissions"""
    return {
        'can_edit': user.user_type == 'operator' and checklist.status == 'pending',
        'can_verify': user.user_type in ['shift_supervisor', 'quality_supervisor'],
        'can_add_concern': checklist.status != 'quality_approved',
        'can_view_all': user.user_type in ['shift_supervisor', 'quality_supervisor']
    }

def calculate_average_interval(subgroups):
    """Calculate average time interval between subgroups"""
    if len(subgroups) < 2:
        return None
        
    intervals = []
    for i in range(1, len(subgroups)):
        interval = (subgroups[i].timestamp - subgroups[i-1].timestamp).total_seconds() / 3600
        intervals.append(interval)
    
    return sum(intervals) / len(intervals)

def get_critical_issues(checklist, subgroups):
    """Identify critical issues in measurements"""
    issues = []
    
    # Check base measurements
    if not (4.5 <= checklist.line_pressure <= 5.5):
        issues.append({
            'type': 'base_measurement',
            'measurement': 'line_pressure',
            'value': checklist.line_pressure,
            'severity': 'critical'
        })
    
    # Check subgroup measurements
    for subgroup in subgroups:
        if not (-43 <= subgroup.uv_vacuum_test <= -35):
            issues.append({
                'type': 'subgroup',
                'subgroup': subgroup.subgroup_number,
                'measurement': 'uv_vacuum_test',
                'value': subgroup.uv_vacuum_test,
                'severity': 'critical'
            })
    
    return issues
    
    
    
    
    
from .forms import SubgroupVerificationForm
from django.db import IntegrityError
from .models import (
    ChecklistBase, 
    SubgroupEntry, 
    SubgroupVerification, 
    Shift, 
    User
)

@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'])
def verify_subgroup_measurement(request, subgroup_id):
    subgroup = get_object_or_404(SubgroupEntry.objects.select_related('checklist'), id=subgroup_id)
    verifier_type = 'supervisor' if request.user.user_type == 'shift_supervisor' else 'quality'
    
    # Get existing verification
    existing_verification = SubgroupVerification.objects.filter(
        subgroup=subgroup,
        verifier_type=verifier_type
    ).first()
    
    # Check verification rules
    if verifier_type == 'quality':
        supervisor_verification = SubgroupVerification.objects.filter(
            subgroup=subgroup,
            verifier_type='supervisor'
        ).first()
        
        if not supervisor_verification or supervisor_verification.status == 'rejected':
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Subgroup must be verified by supervisor first'
                })
            messages.error(request, 'Subgroup must be verified by supervisor first')
            return redirect('checklist_detail', checklist_id=subgroup.checklist.id)
    
    if request.method == 'POST':
        form = SubgroupVerificationForm(request.POST, instance=existing_verification)
        if form.is_valid():
            try:
                verification = form.save(commit=False)
                verification.subgroup = subgroup
                verification.verified_by = request.user
                verification.verifier_type = verifier_type
                
                # FIXED: Convert status BEFORE saving
                if verification.status == "approved":
                    verification.status = "supervisor_verified" if verifier_type == "supervisor" else "quality_verified"
                
                verification.save()  # Now save with correct status

                # Update subgroup status
                new_status = None
                if verification.status == 'rejected':
                    new_status = 'rejected'
                elif verifier_type == 'supervisor':
                    new_status = 'supervisor_verified'
                else:  # quality supervisor
                    new_status = 'quality_verified'
                    
                subgroup.verification_status = new_status
                subgroup.save()
                
                # Check if it's an AJAX request
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': 'Verification completed successfully'
                    })
                
                messages.success(request, 'Verification completed successfully')
                return redirect('dashboard')
                
            except IntegrityError:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'An error occurred during verification'
                    })
                messages.error(request, 'An error occurred during verification')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Please correct the errors below'
                })
            messages.error(request, 'Please correct the errors below')
    
    # Prepare measurements validation for template
    measurements = {
        'uv_vacuum_test_ok': -43 <= subgroup.uv_vacuum_test <= -35,
        'uv_flow_value_ok': 30 <= subgroup.uv_flow_value <= 40,
        'assembly_ok': subgroup.umbrella_valve_assembly == 'OK',
        'pressing_ok': subgroup.uv_clip_pressing == 'OK'
    }
    
    context = {
        'form': form,
        'subgroup': subgroup,
        'measurements': measurements,
        'verifier_type': verifier_type.title(),
        'existing_verification': existing_verification,
    }
    
    return render(request, 'main/verify_subgroup_measurement.html', context)


@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'])
def verify_subgroup_ajax(request):
    if request.method == 'POST' and request.is_ajax():
        subgroup_id = request.POST.get('subgroup_id')
        verifier_type = request.POST.get('verifier_type')
        status = request.POST.get('status')
        comments = request.POST.get('comments')
        
        # Get the subgroup object
        subgroup = get_object_or_404(SubgroupEntry, id=subgroup_id)
        
        # Check verification rules
        if verifier_type == 'quality':
            supervisor_verification = SubgroupVerification.objects.filter(
                subgroup=subgroup, verifier_type='supervisor'
            ).first()
            
            if not supervisor_verification or supervisor_verification.status == 'rejected':
                return JsonResponse({
                    'success': False,
                    'message': 'Subgroup must be verified by supervisor first'
                })
        
        # Get existing verification or create new one
        verification, created = SubgroupVerification.objects.get_or_create(
            subgroup=subgroup,
            verifier_type=verifier_type,
            defaults={
                'verified_by': request.user,
                'status': status,
                'comments': comments
            }
        )
        
        # If updating an existing verification
        if not created:
            verification.status = status
            verification.comments = comments
            verification.save()
        
        # Update subgroup status
        new_status = None
        if status == 'rejected':
            new_status = 'rejected'
        elif verifier_type == 'supervisor':
            new_status = 'supervisor_verified'
        else:  # quality supervisor
            new_status = 'quality_verified'
            
        subgroup.verification_status = new_status
        subgroup.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Verification completed successfully'
        })
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request'
    })
    
        
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'])
def edit_verification(request, verification_id):
    verification = get_object_or_404(SubgroupVerification, id=verification_id)
    
    # Check if user has permission to edit this verification
    if (request.user.user_type == 'shift_supervisor' and verification.verifier_type != 'supervisor') or \
       (request.user.user_type == 'quality_supervisor' and verification.verifier_type != 'quality'):
        messages.error(request, 'You do not have permission to edit this verification')
        return redirect('checklist_detail', checklist_id=verification.subgroup.checklist.id)
    
    if request.method == 'POST':
        form = SubgroupVerificationForm(request.POST, instance=verification)
        if form.is_valid():
            verification = form.save(commit=False)
            verification.verified_at = timezone.now()  # Update verification time
            verification.save()
            
            # Update subgroup status
            subgroup = verification.subgroup
            if verification.status == 'rejected':
                subgroup.verification_status = 'rejected'
            elif verification.verifier_type == 'supervisor':
                subgroup.verification_status = 'supervisor_verified'
            else:  # quality supervisor
                subgroup.verification_status = 'quality_verified'
            subgroup.save()
            
            messages.success(request, 'Verification updated successfully')
            return redirect('dashboard')
    else:
        form = SubgroupVerificationForm(instance=verification)
    
    context = {
        'form': form,
        'verification': verification,
        'subgroup': verification.subgroup,
        'measurements': get_subgroup_measurements(verification.subgroup),
    }
    
    return render(request, 'main/edit_verification.html', context)

def get_subgroup_measurements(subgroup):
    """Helper function to get formatted measurements for a subgroup"""
    return {
        'uv_vacuum_test': {
            'value': subgroup.uv_vacuum_test,
            'is_valid': -43 <= subgroup.uv_vacuum_test <= -35,
            'range': '-43 to -35 kPa'
        },
        'uv_flow_value': {
            'value': subgroup.uv_flow_value,
            'is_valid': 30 <= subgroup.uv_flow_value <= 40,
            'range': '30-40 LPM'
        },
        'assembly_ok': subgroup.umbrella_valve_assembly == 'OK',
        'pressing_ok': subgroup.uv_clip_pressing == 'OK',
        'cleanliness': {
            'workstation': subgroup.workstation_clean == 'Yes',
            'contamination': subgroup.bin_contamination_check == 'Yes'
        }
    }    
# Verification Views
@login_required
@user_passes_test(lambda u: u.user_type == 'shift_supervisor')
def supervisor_verify(request, checklist_id):
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    
    # Verify if checklist can be verified
    if checklist.status != 'pending':
        messages.error(request, 'This checklist has already been verified')
        return redirect('supervisor_dashboard')
        
    if checklist.subgroup_entries.count() < 6:
        messages.error(request, 'Cannot verify incomplete checklist. All 6 subgroups must be completed.')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    if request.method == 'POST':
        form = VerificationForm(request.POST)
        if form.is_valid():
            action = request.POST.get('action', 'approve')
            
            # Perform validation checks
            measurements_ok = all([
                4.5 <= checklist.line_pressure <= 5.5,
                11 <= checklist.uv_flow_input_pressure <= 15,
                0.25 <= checklist.test_pressure_vacuum <= 0.3,
                checklist.oring_condition == 'OK',
                checklist.master_verification_lvdt == 'OK',
                checklist.good_bad_master_verification == 'OK',
                checklist.tool_alignment == 'OK'
            ])
            
            # Check subgroup measurements
            subgroups_ok = True
            measurement_issues = []
            for subgroup in checklist.subgroup_entries.all():
                if not (-43 <= subgroup.uv_vacuum_test <= -35):
                    subgroups_ok = False
                    measurement_issues.append(f"Subgroup {subgroup.subgroup_number}: UV vacuum test out of range")
                if not (30 <= subgroup.uv_flow_value <= 40):
                    subgroups_ok = False
                    measurement_issues.append(f"Subgroup {subgroup.subgroup_number}: UV flow value out of range")
            
            if action == 'approve':
                if not measurements_ok or not subgroups_ok:
                    message = "Warning: Some measurements are out of range:\n"
                    if not measurements_ok:
                        message += "- Initial measurements are out of range\n"
                    if measurement_issues:
                        message += "\n".join(measurement_issues)
                    messages.warning(request, message)
                    return render(request, 'main/verify_checklist.html', {
                        'form': form,
                        'checklist': checklist,
                        'verification_type': 'Supervisor',
                        'measurements_ok': measurements_ok,
                        'subgroups_ok': subgroups_ok,
                        'measurement_issues': measurement_issues
                    })
                
                checklist.status = 'supervisor_approved'
                success_message = 'Checklist approved successfully'
            else:
                checklist.status = 'rejected'
                success_message = 'Checklist rejected'
            
            checklist.supervisor_verified_at = timezone.now()
            checklist.supervisor_comments = form.cleaned_data['comments']
            checklist.save()
            
            # Create verification record
            Verification.objects.create(
                checklist=checklist,
                verified_by=request.user,
                verifier_type='supervisor',
                status=checklist.status,
                comments=form.cleaned_data['comments']
            )
            
            messages.success(request, success_message)
            return redirect('supervisor_dashboard')
    else:
        form = VerificationForm()
    
    # Pre-check measurements for the template
    measurements_ok = all([
        4.5 <= checklist.line_pressure <= 5.5,
        11 <= checklist.uv_flow_input_pressure <= 15,
        0.25 <= checklist.test_pressure_vacuum <= 0.3,
        checklist.oring_condition == 'OK',
        checklist.master_verification_lvdt == 'OK',
        checklist.good_bad_master_verification == 'OK',
        checklist.tool_alignment == 'OK'
    ])
    
    measurement_issues = []
    for subgroup in checklist.subgroup_entries.all():
        if not (-43 <= subgroup.uv_vacuum_test <= -35):
            measurement_issues.append(f"Subgroup {subgroup.subgroup_number}: UV vacuum test out of range")
        if not (30 <= subgroup.uv_flow_value <= 40):
            measurement_issues.append(f"Subgroup {subgroup.subgroup_number}: UV flow value out of range")
    
    return render(request, 'main/verify_checklist.html', {
        'form': form,
        'checklist': checklist,
        'verification_type': 'Supervisor',
        'measurements_ok': measurements_ok,
        'measurement_issues': measurement_issues,
        'subgroups': checklist.subgroup_entries.all().order_by('subgroup_number')
    })
    
@login_required
@user_passes_test(lambda u: u.user_type == 'quality_supervisor')
def quality_verify(request, checklist_id):
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    
    if request.method == 'POST':
        form = VerificationForm(request.POST)
        if form.is_valid():
            checklist.status = 'quality_approved'
            checklist.quality_verified_at = timezone.now()
            checklist.quality_comments = form.cleaned_data['comments']
            checklist.save()
            
            messages.success(request, 'Quality verification completed')
            return redirect('quality_dashboard')
    else:
        form = VerificationForm()
    
    return render(request, 'main/verify_checklist.html', {
        'form': form,
        'checklist': checklist,
        'verification_type': 'Quality'
    })

# Report Views
@login_required
def daily_report(request):
    date = request.GET.get('date', timezone.now().date())
    if isinstance(date, str):
        date = datetime.strptime(date, '%Y-%m-%d').date()
    
    checklists = ChecklistBase.objects.filter(shift__date=date)
    
    summary = {
        'total': checklists.count(),
        'pending': checklists.filter(status='pending').count(),
        'supervisor_approved': checklists.filter(status='supervisor_approved').count(),
        'quality_approved': checklists.filter(status='quality_approved').count(),
        'rejected': checklists.filter(status='rejected').count(),
    }
    
    return render(request, 'main/reports/daily_report.html', {
        'date': date,
        'checklists': checklists,
        'summary': summary
    })

@login_required
def weekly_report(request):
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=7)
    
    checklists = ChecklistBase.objects.filter(
        shift__date__range=[start_date, end_date]
    ).order_by('shift__date')
    
    daily_stats = {}
    for i in range(7):
        current_date = start_date + timedelta(days=i)
        daily_checklists = checklists.filter(shift__date=current_date)
        
        daily_stats[current_date] = {
            'total': daily_checklists.count(),
            'approved': daily_checklists.filter(status='quality_approved').count(),
            'rejected': daily_checklists.filter(status='rejected').count(),
        }
    
    return render(request, 'main/reports/weekly_report.html', {
        'start_date': start_date,
        'end_date': end_date,
        'daily_stats': daily_stats
    })

@login_required
def monthly_report(request):
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    checklists = ChecklistBase.objects.filter(
        shift__date__range=[start_date, end_date]
    )
    
    # Calculate statistics by model
    model_stats = {}
    for model in ChecklistBase.MODEL_CHOICES:
        model_checklists = checklists.filter(selected_model=model[0])
        model_stats[model[0]] = {
            'total': model_checklists.count(),
            'approved': model_checklists.filter(status='quality_approved').count(),
            'rejected': model_checklists.filter(status='rejected').count(),
        }
    
    return render(request, 'main/reports/monthly_report.html', {
        'start_date': start_date,
        'end_date': end_date,
        'model_stats': model_stats
    })

# Profile Views
@login_required
def user_profile(request):
    return render(request, 'main/profile/user_profile.html', {
        'user': request.user
    })

@login_required
def edit_profile(request):
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully')
            return redirect('user_profile')
    else:
        form = UserProfileForm(instance=request.user)
    
    return render(request, 'main/profile/edit_profile.html', {
        'form': form
    })

# API Views
@login_required
def validate_checklist(request):
    """API endpoint for validating checklist data with warnings instead of errors"""
    if request.method == 'POST':
        data = request.POST
        warnings = []
        
        try:
            # Validate line pressure
            line_pressure = float(data.get('line_pressure', 0))
            if not (4.5 <= line_pressure <= 5.5):
                warnings.append({
                    'field': 'line_pressure',
                    'message': f'Line pressure value {line_pressure} bar is outside recommended range (4.5 - 5.5 bar)',
                    'value': line_pressure,
                    'recommended_range': {
                        'min': 4.5,
                        'max': 5.5,
                        'unit': 'bar'
                    }
                })
            
            # Validate UV flow test pressure
            uv_pressure = float(data.get('uv_flow_test_pressure', 0))
            if not (11 <= uv_pressure <= 15):
                warnings.append({
                    'field': 'uv_flow_test_pressure',
                    'message': f'UV flow test pressure value {uv_pressure} kPa is outside recommended range (11 - 15 kPa)',
                    'value': uv_pressure,
                    'recommended_range': {
                        'min': 11,
                        'max': 15,
                        'unit': 'kPa'
                    }
                })
            
            # Validate UV vacuum test
            uv_vacuum = float(data.get('uv_vacuum_test', 0))
            if not (-43 <= uv_vacuum <= -35):
                warnings.append({
                    'field': 'uv_vacuum_test',
                    'message': f'UV vacuum test value {uv_vacuum} kPa is outside recommended range (-43 to -35 kPa)',
                    'value': uv_vacuum,
                    'recommended_range': {
                        'min': -43,
                        'max': -35,
                        'unit': 'kPa'
                    }
                })
            
            return JsonResponse({
                'is_valid': True,  # Always valid as we're only showing warnings
                'has_warnings': len(warnings) > 0,
                'warnings': warnings,
                'values': {  # Return validated values for reference
                    'line_pressure': line_pressure,
                    'uv_flow_test_pressure': uv_pressure,
                    'uv_vacuum_test': uv_vacuum
                }
            })
            
        except (TypeError, ValueError) as e:
            return JsonResponse({
                'is_valid': False,
                'errors': ['Please enter valid numerical values'],
                'details': str(e)
            })
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)    
@login_required
def operator_history(request):
    # Get all checklists grouped by date
    checklists = ChecklistBase.objects.filter(
        shift__operator=request.user
    ).annotate(
        subgroup_count=Count('subgroup_entries')
    ).order_by('-created_at')
    
    # Group by status
    pending_checklists = checklists.filter(status='pending')
    verified_checklists = checklists.filter(status__in=['supervisor_approved', 'quality_approved'])
    rejected_checklists = checklists.filter(status='rejected')
    
    context = {
        'title': 'Operator History',
        'pending_checklists': pending_checklists,
        'verified_checklists': verified_checklists,
        'rejected_checklists': rejected_checklists,
        'total_checklists': checklists.count()
    }
    
    return render(request, 'main/history/operator_history.html', context)
@login_required
@user_passes_test(lambda u: u.user_type == 'shift_supervisor')
def supervisor_history(request):
    # Get all verifications by this supervisor
    verifications = SubgroupVerification.objects.filter(
        verified_by=request.user,
        verifier_type='supervisor'
    ).select_related(
        'subgroup__checklist',
        'subgroup__checklist__shift__operator'
    ).order_by('-verified_at')

    # Get subgroups pending verification
    pending_subgroups = SubgroupEntry.objects.filter(
        checklist__shift__shift_supervisor=request.user,
        verification_status='pending'
    ).select_related(
        'checklist',
        'checklist__shift__operator'
    ).order_by('-timestamp')

    return render(request, 'main/history/supervisor_history.html', {
        'pending_verifications': pending_subgroups,
        'verified_entries': verifications,
        'title': 'Supervisor History'
    })
@login_required
@user_passes_test(lambda u: u.user_type == 'quality_supervisor')
def quality_history(request):
    # Get all quality verifications by this supervisor
    verified_entries = SubgroupVerification.objects.filter(
        verified_by=request.user,
        verifier_type='quality'
    ).select_related(
        'subgroup__checklist',
        'subgroup__checklist__shift__operator',
        'subgroup__checklist__shift__shift_supervisor'
    ).order_by('-verified_at')

    # Get subgroups that have supervisor verification but no quality verification
    pending_verifications = DailyVerificationStatus.objects.filter(
        quality_notified=True,
        status='in_progress',
        shift__quality_supervisor=request.user
    ).select_related(
        'shift__operator',
        'shift__shift_supervisor',
        'created_by'
    ).order_by('-date', '-created_at')

    context = {
        'pending_verifications': pending_subgroups,
        'verified_entries': verified_entries,
        'title': 'Quality History'
    }

    return render(request, 'main/history/quality_history.html', context)    

@login_required
def user_settings(request):
    if request.method == 'POST':
        # Handle settings update
        notification_settings = request.POST.get('notification_settings', False)
        theme_preference = request.POST.get('theme_preference', 'light')
        
        # Save settings to user profile or settings model
        profile = request.user
        profile.email_notifications = notification_settings == 'on'
        profile.theme_preference = theme_preference
        profile.save()
        
        messages.success(request, 'Settings updated successfully')
        return redirect('user_settings')
    
    return render(request, 'main/profile/user_settings.html', {
        'user': request.user
    })
    
@login_required
def notification_settings(request):
    if request.method == 'POST':
        # Handle notification settings update
        request.user.email_notifications = request.POST.get('email_notifications') == 'on'
        request.user.save()
        messages.success(request, 'Notification settings updated successfully.')
        return redirect('user_settings')
        
    return render(request, 'main/settings/notifications.html', {
        'user': request.user,
        'active_tab': 'notifications'
    })

@login_required
def user_preferences(request):
    if request.method == 'POST':
        # Handle user preferences update
        messages.success(request, 'Preferences updated successfully.')
        return redirect('user_settings')
        
    return render(request, 'main/settings/preferences.html', {
        'user': request.user,
        'active_tab': 'preferences'
    })    
    
@login_required
def start_verification(request):
    """Start a new daily verification workflow"""
    if request.method == 'POST':
        form = DailyVerificationWorkflowForm(request.POST, user=request.user)
        if form.is_valid():
            # Create workflow in a transaction
            with transaction.atomic():
                # 1. Create a new shift
                shift = Shift.objects.create(
                    date=form.cleaned_data['date'],
                    shift_type=form.cleaned_data['shift_type'],
                    operator=form.cleaned_data['operator'],
                    shift_supervisor=form.cleaned_data['shift_supervisor'],
                    quality_supervisor=form.cleaned_data['quality_supervisor']
                )
                
                # 2. Create the verification status
                verification_status = DailyVerificationStatus.objects.create(
                    date=form.cleaned_data['date'],
                    shift=shift,
                    status='pending',
                    created_by=request.user
                )
                
                # 3. Create the initial checklist
                checklist = ChecklistBase.objects.create(
                    verification_status=verification_status,
                    selected_model=form.cleaned_data['model'],
                    # Set default values for required fields
                    line_pressure=0,
                    oring_condition='OK',
                    uv_flow_input_pressure=0,
                    master_verification_lvdt='OK',
                    good_bad_master_verification='OK',
                    test_pressure_vacuum=0,
                    tool_alignment='OK',
                    top_tool_id=ChecklistBase.TOP_TOOL_CHOICES[0][0],
                    bottom_tool_id=ChecklistBase.BOTTOM_TOOL_CHOICES[0][0],
                    uv_assy_stage_id=ChecklistBase.UV_ASSY_STAGE_CHOICES[0][0],
                    retainer_part_no=ChecklistBase.RETAINER_PART_CHOICES[0][0],
                    uv_clip_part_no=ChecklistBase.UV_CLIP_PART_CHOICES[0][0],
                    umbrella_part_no=ChecklistBase.UMBRELLA_PART_CHOICES[0][0],
                    retainer_id_lubrication='OK',
                    error_proofing_verification='Yes'
                )
                
                messages.success(request, 'Daily verification workflow started successfully')
                return redirect('edit_checklist', checklist_id=checklist.id)
    else:
        form = DailyVerificationWorkflowForm(user=request.user)
    
    return render(request, 'main/start_verification.html', {
        'form': form,
        'title': 'Start New Verification'
    })    
    
@login_required
def edit_checklist(request, checklist_id):
    """Edit the initial checklist information"""
    checklist = get_object_or_404(ChecklistBase, id=checklist_id)
    
    # Verify ownership and status
    if checklist.verification_status.created_by != request.user:
        messages.error(request, 'You do not have permission to edit this checklist')
        return redirect('dashboard')
    
    if checklist.status != 'pending':
        messages.error(request, 'Cannot edit a verified checklist')
        return redirect('checklist_detail', checklist_id=checklist.id)
    
    if request.method == 'POST':
        form = ChecklistBaseForm(request.POST, instance=checklist)
        if form.is_valid():
            # Save the form and update status
            checklist = form.save()
            
            # Update verification status
            verification_status = checklist.verification_status
            verification_status.status = 'in_progress'
            verification_status.save()
            
            messages.success(request, 'Checklist updated successfully')
            return redirect('checklist_detail', checklist_id=checklist.id)
    else:
        form = ChecklistBaseForm(instance=checklist)
    
    return render(request, 'main/edit_checklist.html', {
        'form': form,
        'checklist': checklist,
        'title': 'Edit Checklist'
    })
        
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'])
def verify_subgroup(request, subgroup_id):
    subgroup = get_object_or_404(SubgroupEntry, id=subgroup_id)
    verifier_type = 'supervisor' if request.user.user_type == 'shift_supervisor' else 'quality'
    
    # Check if subgroup can be verified
    if verifier_type == 'quality' and subgroup.verification_status != 'supervisor_verified':
        messages.error(request, 'Subgroup must be verified by supervisor first')
        return redirect('checklist_detail', checklist_id=subgroup.checklist.id)
    
    if request.method == 'POST':
        form = SubgroupVerificationForm(request.POST)
        if form.is_valid():
            verification = form.save(commit=False)
            verification.subgroup = subgroup
            verification.verified_by = request.user
            verification.verifier_type = verifier_type
            
            # Update subgroup status based on verification
            if verification.status == 'rejected':
                subgroup.verification_status = 'rejected'
            elif verifier_type == 'supervisor':
                subgroup.verification_status = 'supervisor_verified'
            else:  # quality supervisor
                subgroup.verification_status = 'quality_verified'
            
            verification.save()
            subgroup.save()
            
            # Check if all subgroups are verified to update checklist status
            update_checklist_status(subgroup.checklist)
            
            messages.success(request, 'Subgroup verified successfully')
            return redirect('checklist_detail', checklist_id=subgroup.checklist.id)
    else:
        form = SubgroupVerificationForm()
    
    context = {
        'form': form,
        'subgroup': subgroup,
        'verifier_type': verifier_type.title(),
        'measurements': {
            'uv_vacuum_test_ok': -43 <= subgroup.uv_vacuum_test <= -35,
            'uv_flow_value_ok': 30 <= subgroup.uv_flow_value <= 40,
            'assembly_ok': subgroup.umbrella_valve_assembly == 'OK',
            'pressing_ok': subgroup.uv_clip_pressing == 'OK'
        }
    }
    
    return render(request, 'main/verify_subgroup.html', context)

def update_checklist_status(checklist):
    """Update checklist status based on subgroup verifications"""
    subgroups = checklist.subgroup_entries.all()
    total_subgroups = subgroups.count()
    
    if total_subgroups == 0:
        return
    
    # Count verifications
    supervisor_verified = subgroups.filter(verification_status='supervisor_verified').count()
    quality_verified = subgroups.filter(verification_status='quality_verified').count()
    rejected = subgroups.filter(verification_status='rejected').count()
    
    # Update checklist status
    if rejected > 0:
        checklist.status = 'rejected'
    elif quality_verified == total_subgroups:
        checklist.status = 'quality_approved'
    elif supervisor_verified == total_subgroups:
        checklist.status = 'supervisor_approved'
    else:
        checklist.status = 'pending'
    
    checklist.save()    
    
    
    
    
    
    
    
    
    
# New code 


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Sum, Avg, Q, F
from django.core.paginator import Paginator
from datetime import datetime, timedelta
import json

from .models import (
    OperationNumber, DefectCategory, DefectType, CustomDefectType, 
    FTQRecord, DefectRecord, Shift, User, ChecklistBase
)
from .forms import (
    OperationNumberForm, DefectCategoryForm, DefectTypeForm, 
    FTQRecordForm, DefectRecordForm, CustomDefectTypeForm
)

# Operation Number Views
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def operation_number_list(request):
    """List all operation numbers"""
    operations = OperationNumber.objects.all().order_by('number')
    
    return render(request, 'main/operations/operation_list.html', {
        'operations': operations,
        'title': 'Operation Numbers'
    })

@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def operation_number_create(request):
    """Create a new operation number"""
    if request.method == 'POST':
        form = OperationNumberForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Operation number created successfully')
            return redirect('operation_number_list')
    else:
        form = OperationNumberForm()
    
    return render(request, 'main/operations/operation_form.html', {
        'form': form,
        'title': 'Create Operation Number'
    })

@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def operation_number_edit(request, pk):
    """Edit an operation number"""
    operation = get_object_or_404(OperationNumber, pk=pk)
    
    if request.method == 'POST':
        form = OperationNumberForm(request.POST, instance=operation)
        if form.is_valid():
            form.save()
            messages.success(request, 'Operation number updated successfully')
            return redirect('operation_number_list')
    else:
        form = OperationNumberForm(instance=operation)
    
    return render(request, 'main/operations/operation_form.html', {
        'form': form,
        'title': 'Edit Operation Number',
        'operation': operation
    })

@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def operation_number_delete(request, pk):
    """Delete an operation number"""
    operation = get_object_or_404(OperationNumber, pk=pk)
    
    if request.method == 'POST':
        # Check if this operation is used by any defect types
        if DefectType.objects.filter(operation_number=operation).exists():
            messages.error(request, 'Cannot delete operation number because it is associated with defect types')
            return redirect('operation_number_list')
        
        operation.delete()
        messages.success(request, 'Operation number deleted successfully')
        return redirect('operation_number_list')
    
    return render(request, 'main/operations/operation_confirm_delete.html', {
        'operation': operation
    })

# Defect Category Views
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def defect_category_list(request):
    """List all defect categories"""
    categories = DefectCategory.objects.all().order_by('name')
    
    return render(request, 'main/operations/category_list.html', {
        'categories': categories,
        'title': 'Defect Categories'
    })

@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def defect_category_create(request):
    """Create a new defect category"""
    if request.method == 'POST':
        form = DefectCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Defect category created successfully')
            return redirect('defect_category_list')
    else:
        form = DefectCategoryForm()
    
    return render(request, 'main/operations/category_form.html', {
        'form': form,
        'title': 'Create Defect Category'
    })

@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def defect_category_edit(request, pk):
    """Edit a defect category"""
    category = get_object_or_404(DefectCategory, pk=pk)
    
    if request.method == 'POST':
        form = DefectCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Defect category updated successfully')
            return redirect('defect_category_list')
    else:
        form = DefectCategoryForm(instance=category)
    
    return render(request, 'main/operations/category_form.html', {
        'form': form,
        'title': 'Edit Defect Category',
        'category': category
    })

@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def defect_category_delete(request, pk):
    """Delete a defect category"""
    category = get_object_or_404(DefectCategory, pk=pk)
    
    if request.method == 'POST':
        # Check if this category is used by any defect types
        if DefectType.objects.filter(category=category).exists():
            messages.error(request, 'Cannot delete category because it is associated with defect types')
            return redirect('defect_category_list')
        
        category.delete()
        messages.success(request, 'Defect category deleted successfully')
        return redirect('defect_category_list')
    
    return render(request, 'main/operations/category_confirm_delete.html', {
        'category': category
    })

# Defect Type Views
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def defect_type_list(request):
    """List all defect types"""
    operation_id = request.GET.get('operation', None)
    category_id = request.GET.get('category', None)
    
    defect_types = DefectType.objects.select_related('operation_number', 'category')
    
    # Apply filters if provided
    if operation_id:
        defect_types = defect_types.filter(operation_number_id=operation_id)
    if category_id:
        defect_types = defect_types.filter(category_id=category_id)
    
    defect_types = defect_types.order_by('operation_number__number', 'order', 'name')
    
    # Get list of operations and categories for filter dropdowns
    operations = OperationNumber.objects.all().order_by('number')
    categories = DefectCategory.objects.all().order_by('name')
    
    return render(request, 'main/operations/defect_type_list.html', {
        'defect_types': defect_types,
        'operations': operations,
        'categories': categories,
        'selected_operation': operation_id,
        'selected_category': category_id,
        'title': 'Defect Types'
    })

@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def defect_type_create(request):
    """Create a new defect type"""
    if request.method == 'POST':
        form = DefectTypeForm(request.POST)
        if form.is_valid():
            defect_type = form.save()
            messages.success(request, 'Defect type created successfully')
            
            # If this is set as default, check the order
            if defect_type.is_default:
                # Count existing default defects for this operation
                count = DefectType.objects.filter(
                    operation_number=defect_type.operation_number,
                    is_default=True
                ).count()
                
                if count > 10:
                    messages.warning(request, f'Note: There are now {count} default defect types for Operation {defect_type.operation_number}. Only 10 will be shown by default.')
            
            return redirect('defect_type_list')
    else:
        form = DefectTypeForm()
    
    return render(request, 'main/operations/defect_type_form.html', {
        'form': form,
        'title': 'Create Defect Type'
    })

@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def defect_type_edit(request, pk):
    """Edit a defect type"""
    defect_type = get_object_or_404(DefectType, pk=pk)
    
    if request.method == 'POST':
        form = DefectTypeForm(request.POST, instance=defect_type)
        if form.is_valid():
            form.save()
            messages.success(request, 'Defect type updated successfully')
            return redirect('defect_type_list')
    else:
        form = DefectTypeForm(instance=defect_type)
    
    return render(request, 'main/operations/defect_type_form.html', {
        'form': form,
        'title': 'Edit Defect Type',
        'defect_type': defect_type
    })

@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def defect_type_delete(request, pk):
    """Delete a defect type"""
    defect_type = get_object_or_404(DefectType, pk=pk)
    
    if request.method == 'POST':
        # Check if this defect type is used in any records
        if DefectRecord.objects.filter(defect_type=defect_type).exists():
            messages.error(request, 'Cannot delete defect type because it is associated with defect records')
            return redirect('defect_type_list')
        
        defect_type.delete()
        messages.success(request, 'Defect type deleted successfully')
        return redirect('defect_type_list')
    
    return render(request, 'main/operations/defect_type_confirm_delete.html', {
        'defect_type': defect_type
    })

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum
from .models import (
    FTQRecord, TimeBasedDefectEntry, DefectType, CustomDefectType, 
    OperationNumber, DailyVerificationStatus, ChecklistBase
)
from .forms import FTQRecordForm

# FTQ Record Views
@login_required
def ftq_list(request):
    """List FTQ records with filtering options"""
    # Get filter parameters
    start_date_str = request.GET.get('start_date', (timezone.now().date() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('end_date', timezone.now().date().strftime('%Y-%m-%d'))
    model_name = request.GET.get('model_name', '')
    shift_type = request.GET.get('shift_type', '')
    
    # Parse dates with default fallback
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = timezone.now().date() - timedelta(days=30)
    
    try:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        end_date = timezone.now().date()
    
    # Query FTQ records with prefetch for time-based defects
    ftq_records = FTQRecord.objects.select_related(
        'created_by', 
        'verified_by'
    ).prefetch_related(
        'time_based_defects'
    )
    
    # Apply filters
    ftq_records = ftq_records.filter(date__range=[start_date, end_date])
    
    if model_name:
        ftq_records = ftq_records.filter(model_name=model_name)
    
    if shift_type:
        ftq_records = ftq_records.filter(shift_type=shift_type)
    
    # Order by date (newest first)
    ftq_records = ftq_records.order_by('-date', '-shift_type')
    
    # Paginate results
    paginator = Paginator(ftq_records, 20)  # Show 20 records per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate summary statistics manually (since total_defects is a property)
    total_inspected = 0
    total_defects = 0
    
    for record in ftq_records:
        total_inspected += record.total_inspected
        total_defects += record.total_defects  # This uses the property
    
    stats = {
        'total_inspected': total_inspected,
        'total_defects': total_defects,
        'record_count': ftq_records.count()
    }
    
    # Calculate FTQ percentage if there are records
    if stats['total_inspected'] > 0:
        stats['ftq_percentage'] = ((stats['total_inspected'] - stats['total_defects']) / stats['total_inspected']) * 100
    else:
        stats['ftq_percentage'] = 0
    
    # Get model choices for filter dropdown
    model_choices = FTQRecord.MODEL_CHOICES
    
    return render(request, 'main/operations/ftq_list.html', {
        'page_obj': page_obj,
        'stats': stats,
        'model_choices': model_choices,
        'shift_choices': FTQRecord.SHIFTS,  # Fixed: use SHIFTS not SHIFT_CHOICES
        'filters': {
            'start_date': start_date,
            'end_date': end_date,
            'model_name': model_name,
            'shift_type': shift_type
        },
        'title': 'FTQ Records'
    })



@login_required
def ftq_record_create(request):
    """Create a new FTQ record with time-based defect tracking"""
    active_verification = DailyVerificationStatus.objects.filter(
        created_by=request.user,
        date=timezone.now().date(),
        status__in=['pending', 'in_progress']
    ).first()
    
    active_checklist = None
    if active_verification:
        active_checklist = ChecklistBase.objects.filter(
            verification_status=active_verification
        ).first()
    
    initial_data = {}
    if active_checklist:
        initial_data = {
            'date': active_verification.date,
            'shift_type': active_verification.shift.shift_type,
            'model_name': active_checklist.selected_model,
            'julian_date': active_verification.date
        }
    
    if request.method == 'POST':
        form = FTQRecordForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    ftq_record = form.save(commit=False)
                    ftq_record.created_by = request.user
                    
                    if active_verification:
                        ftq_record.verification_status = active_verification
                    
                    ftq_record.save()
                    
                    # Get operation number
                    operation = OperationNumber.objects.get(number='OP#35')
                    
                    # Process time-based defect entries
                    defect_data = {}
                    
                    # Parse all defect entries from POST data
                    for key in request.POST:
                        if key.startswith('defect_time_'):
                            parts = key.split('_')
                            defect_id = parts[2]
                            time_index = parts[3]
                            
                            time_value = request.POST.get(key)
                            count_key = f'defect_count_{defect_id}_{time_index}'
                            count_value = request.POST.get(count_key, 0)
                            
                            if time_value and int(count_value) > 0:
                                if defect_id not in defect_data:
                                    defect_data[defect_id] = []
                                
                                defect_data[defect_id].append({
                                    'time': time_value,
                                    'count': int(count_value)
                                })
                    
                    # Create TimeBasedDefectEntry records
                    for defect_id, entries in defect_data.items():
                        if defect_id.startswith('custom_'):
                            custom_name = request.POST.get(f'custom_defect_name_{defect_id}')
                            if custom_name:
                                custom_defect = CustomDefectType.objects.create(
                                    ftq_record=ftq_record,
                                    name=custom_name,
                                    operation_number=operation,
                                    added_by=request.user
                                )
                                
                                for entry in entries:
                                    TimeBasedDefectEntry.objects.create(
                                        ftq_record=ftq_record,
                                        defect_type_custom=custom_defect,
                                        recorded_at=entry['time'],
                                        count=entry['count']
                                    )
                        else:
                            try:
                                defect_type = DefectType.objects.get(id=int(defect_id))
                                for entry in entries:
                                    TimeBasedDefectEntry.objects.create(
                                        ftq_record=ftq_record,
                                        defect_type=defect_type,
                                        recorded_at=entry['time'],
                                        count=entry['count']
                                    )
                            except DefectType.DoesNotExist:
                                continue
                    
                    messages.success(request, 'FTQ record created successfully with time-based defect tracking')
                    return redirect('operator_dashboard')
                    
            except Exception as e:
                messages.error(request, f'Error creating FTQ record: {str(e)}')
                
    else:
        form = FTQRecordForm(initial=initial_data)
    
    # Get default defect types
    try:
        operation = OperationNumber.objects.get(number='OP#35')
        default_defects = DefectType.objects.filter(
            operation_number=operation,
            is_default=True
        ).order_by('order')[:10]
    except OperationNumber.DoesNotExist:
        default_defects = []
    
    context = {
        'form': form,
        'default_defects': default_defects,
        'title': 'Create FTQ Record',
        'active_verification': active_verification,
        'active_checklist': active_checklist
    }
    
    return render(request, 'main/operations/ftq_record_form.html', context)


@login_required
def ftq_record_edit(request, pk):
    """Edit an FTQ record with time-based defects"""
    ftq_record = get_object_or_404(FTQRecord, pk=pk)
    
    if not (request.user.user_type in ['shift_supervisor', 'quality_supervisor'] or 
            request.user.is_superuser or request.user == ftq_record.created_by):
        messages.error(request, 'You do not have permission to edit this record')
        return redirect('ftq_list')
    
    # Get existing time-based entries organized by defect type
    time_entries = TimeBasedDefectEntry.objects.filter(
        ftq_record=ftq_record
    ).select_related('defect_type', 'defect_type_custom').order_by('recorded_at')
    
    # Organize existing entries by defect type
    standard_defect_entries = {}
    custom_defect_entries = {}
    
    for entry in time_entries:
        if entry.defect_type:
            key = entry.defect_type.id
            if key not in standard_defect_entries:
                standard_defect_entries[key] = []
            standard_defect_entries[key].append({
                'time': entry.recorded_at.strftime('%H:%M'),
                'count': entry.count
            })
        else:
            key = entry.defect_type_custom.id
            if key not in custom_defect_entries:
                custom_defect_entries[key] = {
                    'name': entry.defect_type_custom.name,
                    'entries': []
                }
            custom_defect_entries[key]['entries'].append({
                'time': entry.recorded_at.strftime('%H:%M'),
                'count': entry.count
            })
    
    if request.method == 'POST':
        form = FTQRecordForm(request.POST, instance=ftq_record)
        if form.is_valid():
            try:
                with transaction.atomic():
                    ftq_record = form.save()
                    
                    # Delete all existing time-based entries
                    TimeBasedDefectEntry.objects.filter(ftq_record=ftq_record).delete()
                    
                    # Delete existing custom defect types for this record
                    CustomDefectType.objects.filter(ftq_record=ftq_record).delete()
                    
                    # Get operation number
                    operation = OperationNumber.objects.get(number='OP#35')
                    
                    # Process time-based defect entries
                    defect_data = {}
                    
                    for key in request.POST:
                        if key.startswith('defect_time_'):
                            parts = key.split('_')
                            defect_id = parts[2]
                            time_index = parts[3]
                            
                            time_value = request.POST.get(key)
                            count_key = f'defect_count_{defect_id}_{time_index}'
                            count_value = request.POST.get(count_key, 0)
                            
                            if time_value and int(count_value) > 0:
                                if defect_id not in defect_data:
                                    defect_data[defect_id] = []
                                
                                defect_data[defect_id].append({
                                    'time': time_value,
                                    'count': int(count_value)
                                })
                    
                    # Create new TimeBasedDefectEntry records
                    for defect_id, entries in defect_data.items():
                        if defect_id.startswith('custom_'):
                            custom_name = request.POST.get(f'custom_defect_name_{defect_id}')
                            if custom_name:
                                custom_defect = CustomDefectType.objects.create(
                                    ftq_record=ftq_record,
                                    name=custom_name,
                                    operation_number=operation,
                                    added_by=request.user
                                )
                                
                                for entry in entries:
                                    TimeBasedDefectEntry.objects.create(
                                        ftq_record=ftq_record,
                                        defect_type_custom=custom_defect,
                                        recorded_at=entry['time'],
                                        count=entry['count']
                                    )
                        else:
                            try:
                                defect_type = DefectType.objects.get(id=int(defect_id))
                                for entry in entries:
                                    TimeBasedDefectEntry.objects.create(
                                        ftq_record=ftq_record,
                                        defect_type=defect_type,
                                        recorded_at=entry['time'],
                                        count=entry['count']
                                    )
                            except DefectType.DoesNotExist:
                                continue
                    
                    messages.success(request, 'FTQ record updated successfully with time-based defect tracking')
                    return redirect('ftq_record_detail', pk=ftq_record.pk)
                    
            except Exception as e:
                messages.error(request, f'Error updating FTQ record: {str(e)}')
    else:
        form = FTQRecordForm(instance=ftq_record)
    
    # Get default defect types with their entries
    try:
        operation = OperationNumber.objects.get(number='OP#35')
        default_defects_qs = DefectType.objects.filter(
            operation_number=operation,
            is_default=True
        ).order_by('order')[:10]
        
        # Add entries JSON to each defect
        default_defects = []
        for defect in default_defects_qs:
            entries = standard_defect_entries.get(defect.id, [])
            import json
            defect.entries_json = json.dumps(entries)
            default_defects.append(defect)
    except OperationNumber.DoesNotExist:
        default_defects = []
    
    # Get existing custom defects for this record with their entries
    existing_custom_defects = []
    for custom in CustomDefectType.objects.filter(ftq_record=ftq_record):
        entries = custom_defect_entries.get(custom.id, {}).get('entries', [])
        import json
        existing_custom_defects.append({
            'id': custom.id,
            'name': custom.name,
            'entries_json': json.dumps(entries)
        })
    
    return render(request, 'main/operations/ftq_record_edit.html', {
        'form': form,
        'default_defects': default_defects,
        'standard_defect_entries': standard_defect_entries,
        'custom_defect_entries': custom_defect_entries,
        'existing_custom_defects': existing_custom_defects,
        'ftq_record': ftq_record,
        'title': 'Edit FTQ Record'
    })    
    
@login_required
def ftq_record_detail(request, pk):
    """View FTQ record with time-based defect details"""
    ftq_record = get_object_or_404(FTQRecord, pk=pk)
    
    # Organize defects by type with their time entries
    defect_summary = {}
    
    for entry in ftq_record.time_based_defects.all():
        if entry.defect_type:
            key = f"standard_{entry.defect_type.id}"
            name = entry.defect_type.name
            is_critical = entry.defect_type.is_critical
        else:
            key = f"custom_{entry.defect_type_custom.id}"
            name = entry.defect_type_custom.name
            is_critical = False
        
        if key not in defect_summary:
            defect_summary[key] = {
                'name': name,
                'is_critical': is_critical,
                'entries': [],
                'total_count': 0
            }
        
        defect_summary[key]['entries'].append({
            'time': entry.recorded_at,
            'count': entry.count
        })
        defect_summary[key]['total_count'] += entry.count
    
    # Check if user can verify
    can_verify = (
        request.user.user_type in ['shift_supervisor', 'quality_supervisor'] or 
        request.user.is_superuser
    ) and not ftq_record.verified_by
    
    context = {
        'ftq_record': ftq_record,
        'defect_summary': defect_summary,
        'can_verify': can_verify,
        'title': 'FTQ Record Details'
    }
    
    return render(request, 'main/operations/ftq_record_detail.html', context)



@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def ftq_record_verify(request, pk):
    """Verify an FTQ record (supervisor approval)"""
    ftq_record = get_object_or_404(FTQRecord, pk=pk)
    
    if ftq_record.verified_by:
        messages.warning(request, 'This FTQ record has already been verified')
        return redirect('ftq_record_detail', pk=ftq_record.pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        comments = request.POST.get('comments', '')
        
        if action == 'approve':
            ftq_record.verified_by = request.user
            ftq_record.save()
            messages.success(request, f'FTQ record approved successfully. Comments: {comments}' if comments else 'FTQ record approved successfully')
            return redirect('ftq_record_detail', pk=ftq_record.pk)
        elif action == 'reject':
            messages.warning(request, f'FTQ record rejected. Reason: {comments}' if comments else 'FTQ record rejected')
            return redirect('ftq_list')
    
    # Get time-based defect entries grouped by defect type
    time_entries = TimeBasedDefectEntry.objects.filter(
        ftq_record=ftq_record
    ).select_related('defect_type', 'defect_type_custom').order_by('defect_type', 'defect_type_custom', 'recorded_at')
    
    # Organize defects by type with aggregated counts
    standard_defects = {}
    custom_defects = {}
    
    for entry in time_entries:
        if entry.defect_type:
            if entry.defect_type.id not in standard_defects:
                standard_defects[entry.defect_type.id] = {
                    'defect_type': entry.defect_type,
                    'count': 0,
                    'entries': []
                }
            standard_defects[entry.defect_type.id]['count'] += entry.count
            standard_defects[entry.defect_type.id]['entries'].append({
                'time': entry.recorded_at.strftime('%H:%M'),
                'count': entry.count
            })
        else:
            if entry.defect_type_custom.id not in custom_defects:
                custom_defects[entry.defect_type_custom.id] = {
                    'defect_type_custom': entry.defect_type_custom,
                    'count': 0,
                    'entries': []
                }
            custom_defects[entry.defect_type_custom.id]['count'] += entry.count
            custom_defects[entry.defect_type_custom.id]['entries'].append({
                'time': entry.recorded_at.strftime('%H:%M'),
                'count': entry.count
            })
    
    # Convert to lists for template
    default_defects = list(standard_defects.values())
    custom_defects_list = list(custom_defects.values())
    
    # Calculate FTQ percentage
    ftq_percentage = ftq_record.ftq_percentage
    
    return render(request, 'main/operations/ftq_record_verify.html', {
        'ftq_record': ftq_record,
        'default_defects': default_defects,
        'custom_defects': custom_defects_list,
        'ftq_percentage': ftq_percentage,
        'title': 'Verify FTQ Record'
    })



@login_required
def ftq_record_delete(request, pk):
    """Delete an FTQ record"""
    ftq_record = get_object_or_404(FTQRecord, pk=pk)
    
    # Check if user has permission to delete
    if not (request.user.is_superuser or request.user == ftq_record.created_by):
        messages.error(request, 'You do not have permission to delete this record')
        return redirect('ftq_list')
    
    if request.method == 'POST':
        ftq_record.delete()
        messages.success(request, 'FTQ record deleted successfully')
        return redirect('ftq_list')
    
    return render(request, 'main/operations/ftq_record_confirm_delete.html', {
        'ftq_record': ftq_record
    })

# API Views for AJAX
@login_required
def get_defect_types_by_operation(request):
    """API to get defect types for a specific operation"""
    operation_id = request.GET.get('operation_id')
    
    if not operation_id:
        return JsonResponse({'error': 'Operation ID is required'}, status=400)
    
    try:
        defect_types = DefectType.objects.filter(
            operation_number_id=operation_id,
            is_default=True
        ).order_by('order')
        
        data = [{
            'id': defect.id,
            'name': defect.name,
            'is_critical': defect.is_critical,
            'order': defect.order
        } for defect in defect_types]
        
        return JsonResponse({'defect_types': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def add_custom_defect(request):
    """API to add a custom defect type to an FTQ record"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method is allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        ftq_record_id = data.get('ftq_record_id')
        defect_name = data.get('defect_name')
        count = data.get('count', 0)
        
        if not all([ftq_record_id, defect_name, count]):
            return JsonResponse({'error': 'Missing required fields'}, status=400)
        
        ftq_record = get_object_or_404(FTQRecord, id=ftq_record_id)
        
        # Get the OP#35 operation (default)
        operation = OperationNumber.objects.get(number='OP#35')
        
        # Create custom defect type
        custom_defect = CustomDefectType.objects.create(
            ftq_record=ftq_record,
            name=defect_name,
            operation_number=operation,
            added_by=request.user
        )
        
        # Create defect record
        defect_record = DefectRecord.objects.create(
            ftq_record=ftq_record,
            defect_type_custom=custom_defect,
            count=count
        )
        
        # Update total defects
        ftq_record.total_defects = ftq_record.calculate_total_defects
        ftq_record.save()
        
        return JsonResponse({
            'success': True,
            'custom_defect_id': custom_defect.id,
            'defect_record_id': defect_record.id,
            'total_defects': ftq_record.total_defects
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def update_defect_count(request):
    """API to update a defect count"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method is allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        record_id = data.get('record_id')
        count = data.get('count', 0)
        
        if not all([record_id, count is not None]):
            return JsonResponse({'error': 'Missing required fields'}, status=400)
        
        defect_record = get_object_or_404(DefectRecord, id=record_id)
        defect_record.count = count
        defect_record.save()
        
        # Update total defects
        ftq_record = defect_record.ftq_record
        ftq_record.total_defects = ftq_record.calculate_total_defects
        ftq_record.save()
        
        return JsonResponse({
            'success': True,
            'total_defects': ftq_record.total_defects
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Dashboard and Reports
@login_required
def ftq_dashboard(request):
    """FTQ Dashboard with charts and statistics using time-based defects"""
    # Get date range filters
    today = timezone.now().date()
    
    # Check for custom date range
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            days = (end_date - start_date).days + 1
        except ValueError:
            days = int(request.GET.get('days', 30))
            end_date = today
            start_date = end_date - timedelta(days=days-1)
    else:
        days = int(request.GET.get('days', 30))
        end_date = today
        start_date = end_date - timedelta(days=days-1)
    
    # Get FTQ records in range
    ftq_records = FTQRecord.objects.filter(
        date__range=[start_date, end_date]
    ).prefetch_related('time_based_defects').order_by('date')
    
    # Calculate summary statistics manually (since total_defects is a property)
    total_inspected = 0
    total_defects = 0
    
    for record in ftq_records:
        total_inspected += record.total_inspected
        total_defects += record.total_defects
    
    stats = {
        'total_records': ftq_records.count(),
        'total_inspected': total_inspected,
        'total_defects': total_defects,
    }
    
    if stats['total_inspected'] > 0:
        stats['ftq_percentage'] = round(((stats['total_inspected'] - stats['total_defects']) / stats['total_inspected']) * 100, 2)
    else:
        stats['ftq_percentage'] = 0
    
    # Get model-wise statistics
    model_data = {}
    for record in ftq_records:
        model = record.model_name
        if model not in model_data:
            model_data[model] = {'total_inspected': 0, 'total_defects': 0}
        model_data[model]['total_inspected'] += record.total_inspected
        model_data[model]['total_defects'] += record.total_defects
    
    model_stats = []
    for model, data in model_data.items():
        if data['total_inspected'] > 0:
            ftq_pct = round(((data['total_inspected'] - data['total_defects']) / data['total_inspected']) * 100, 2)
        else:
            ftq_pct = 0
        
        model_stats.append({
            'model_name': model,
            'total_inspected': data['total_inspected'],
            'total_defects': data['total_defects'],
            'ftq_percentage': ftq_pct
        })
    
    # Get shift-wise statistics
    shift_data = {}
    for record in ftq_records:
        shift = record.get_shift_type_display() if record.shift_type else 'Unknown'
        if shift not in shift_data:
            shift_data[shift] = {'total_inspected': 0, 'total_defects': 0}
        shift_data[shift]['total_inspected'] += record.total_inspected
        shift_data[shift]['total_defects'] += record.total_defects
    
    shift_stats = []
    for shift, data in shift_data.items():
        if data['total_inspected'] > 0:
            ftq_pct = round(((data['total_inspected'] - data['total_defects']) / data['total_inspected']) * 100, 2)
        else:
            ftq_pct = 0
        
        shift_stats.append({
            'shift_type': shift,
            'total_inspected': data['total_inspected'],
            'total_defects': data['total_defects'],
            'ftq_percentage': ftq_pct
        })
    
    # Get top defects from time-based entries
    defect_counts = {}
    
    for record in ftq_records:
        for entry in record.time_based_defects.all():
            if entry.defect_type:
                defect_name = entry.defect_type.name
            else:
                defect_name = entry.defect_type_custom.name
            
            if defect_name not in defect_counts:
                defect_counts[defect_name] = 0
            defect_counts[defect_name] += entry.count
    
    # Sort and get top 10
    top_defects = [
        {'defect_type__name': name, 'total_count': count}
        for name, count in sorted(defect_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    ]
    
    # Prepare chart data for daily FTQ trend
    daily_data = {}
    for record in ftq_records:
        date_str = record.date.strftime('%Y-%m-%d')
        if date_str not in daily_data:
            daily_data[date_str] = {'total_inspected': 0, 'total_defects': 0}
        daily_data[date_str]['total_inspected'] += record.total_inspected
        daily_data[date_str]['total_defects'] += record.total_defects
    
    # Calculate daily FTQ percentages
    chart_dates = []
    chart_ftq = []
    total_inspected_daily = []
    total_defects_daily = []
    
    for date_str, data in sorted(daily_data.items()):
        chart_dates.append(date_str)
        if data['total_inspected'] > 0:
            ftq = round(((data['total_inspected'] - data['total_defects']) / data['total_inspected']) * 100, 2)
        else:
            ftq = 0
        chart_ftq.append(ftq)
        total_inspected_daily.append(data['total_inspected'])
        total_defects_daily.append(data['total_defects'])
    
    # Prepare daily defect data for top defects
    daily_defect_data = {}
    for defect in top_defects:
        defect_name = defect['defect_type__name']
        daily_defect_data[defect_name] = {}
        
        # Initialize with zeros
        for date_str in chart_dates:
            daily_defect_data[defect_name][date_str] = 0
    
    # Populate daily defect data
    for record in ftq_records:
        date_str = record.date.strftime('%Y-%m-%d')
        for entry in record.time_based_defects.all():
            if entry.defect_type:
                defect_name = entry.defect_type.name
            else:
                defect_name = entry.defect_type_custom.name
            
            if defect_name in daily_defect_data and date_str in daily_defect_data[defect_name]:
                daily_defect_data[defect_name][date_str] += entry.count
    
    # Calculate defect trends
    defect_trends = {}
    for defect_name, date_data in daily_defect_data.items():
        dates = sorted(date_data.keys())
        if len(dates) >= 7:
            last_7_days = dates[-7:]
            first_3_days_avg = sum(date_data[date] for date in last_7_days[:3]) / 3
            last_3_days_avg = sum(date_data[date] for date in last_7_days[-3:]) / 3
            
            if first_3_days_avg > 0:
                percent_change = ((last_3_days_avg - first_3_days_avg) / first_3_days_avg) * 100
            else:
                percent_change = 0 if last_3_days_avg == 0 else 100
                
            defect_trends[defect_name] = {
                'percent_change': round(percent_change, 1),
                'direction': 'up' if percent_change > 5 else ('down' if percent_change < -5 else 'stable')
            }
        else:
            defect_trends[defect_name] = {'percent_change': 0, 'direction': 'stable'}
    
    # Find most common model and shift for each defect
    defect_common_models = {}
    defect_common_shifts = {}
    
    for defect in top_defects:
        defect_name = defect['defect_type__name']
        model_defect_counts = {}
        shift_defect_counts = {}
        
        for record in ftq_records:
            for entry in record.time_based_defects.all():
                entry_name = entry.defect_type.name if entry.defect_type else entry.defect_type_custom.name
                
                if entry_name == defect_name:
                    # Count by model
                    if record.model_name not in model_defect_counts:
                        model_defect_counts[record.model_name] = 0
                    model_defect_counts[record.model_name] += entry.count
                    
                    # Count by shift
                    shift = record.get_shift_type_display() if record.shift_type else 'Unknown'
                    if shift not in shift_defect_counts:
                        shift_defect_counts[shift] = 0
                    shift_defect_counts[shift] += entry.count
        
        # Find most common
        if model_defect_counts:
            defect_common_models[defect_name] = max(model_defect_counts, key=model_defect_counts.get)
        else:
            defect_common_models[defect_name] = 'N/A'
            
        if shift_defect_counts:
            defect_common_shifts[defect_name] = max(shift_defect_counts, key=shift_defect_counts.get)
        else:
            defect_common_shifts[defect_name] = 'N/A'
    
    # Prepare chart data
    chart_data = {
        'dates': chart_dates,
        'ftq': chart_ftq,
        'total_inspected_daily': total_inspected_daily,
        'total_defects_daily': total_defects_daily,
        'models': [model['model_name'] for model in model_stats],
        'model_ftq': [model['ftq_percentage'] for model in model_stats],
        'model_inspected': [model['total_inspected'] for model in model_stats],
        'model_defects': [model['total_defects'] for model in model_stats],
        'shifts': [shift['shift_type'] for shift in shift_stats],
        'shift_ftq': [shift['ftq_percentage'] for shift in shift_stats],
        'shift_inspected': [shift['total_inspected'] for shift in shift_stats],
        'shift_defects': [shift['total_defects'] for shift in shift_stats],
        'top_defect_names': [defect['defect_type__name'] for defect in top_defects],
        'top_defect_counts': [defect['total_count'] for defect in top_defects],
        'daily_defect_data': daily_defect_data,
        'defect_trends': defect_trends,
        'defect_common_models': defect_common_models,
        'defect_common_shifts': defect_common_shifts
    }
    
    return render(request, 'main/operations/ftq_dashboard.html', {
        'stats': stats,
        'model_stats': model_stats,
        'shift_stats': shift_stats,
        'top_defects': top_defects,
        'chart_data': json.dumps(chart_data),
        'chart_ftq': chart_ftq,
        'days': days,
        'start_date': start_date,
        'end_date': end_date,
        'title': 'FTQ Dashboard'
    })


@login_required
def ftq_report(request, report_type='daily'):
    """Generate FTQ reports (daily, weekly, monthly)"""
    today = timezone.now().date()
    
    if report_type == 'daily':
        # Get date from request or use today
        report_date = request.GET.get('date', today.strftime('%Y-%m-%d'))
        if isinstance(report_date, str):
            try:
                report_date = datetime.strptime(report_date, '%Y-%m-%d').date()
            except ValueError:
                report_date = today
        
        # Get records for the date
        ftq_records = FTQRecord.objects.filter(date=report_date).order_by('shift_type')
        
        # Group by shift and model
        shifts = {}
        for record in ftq_records:
            shift_key = record.shift_type
            if shift_key not in shifts:
                shifts[shift_key] = {
                    'shift_name': record.get_shift_type_display(),
                    'models': {},
                    'total_inspected': 0,
                    'total_defects': 0
                }
            
            model_key = record.model_name
            if model_key not in shifts[shift_key]['models']:
                shifts[shift_key]['models'][model_key] = {
                    'model_name': record.get_model_name_display(),
                    'records': [],
                    'total_inspected': 0,
                    'total_defects': 0
                }
            
            shifts[shift_key]['models'][model_key]['records'].append(record)
            shifts[shift_key]['models'][model_key]['total_inspected'] += record.total_inspected
            shifts[shift_key]['models'][model_key]['total_defects'] += record.total_defects
            shifts[shift_key]['total_inspected'] += record.total_inspected
            shifts[shift_key]['total_defects'] += record.total_defects
        
        # Calculate FTQ percentages
        daily_total_inspected = 0
        daily_total_defects = 0
        
        for shift_key, shift in shifts.items():
            if shift['total_inspected'] > 0:
                shift['ftq_percentage'] = round(((shift['total_inspected'] - shift['total_defects']) / shift['total_inspected']) * 100, 2)
            else:
                shift['ftq_percentage'] = 0
            
            for model_key, model in shift['models'].items():
                if model['total_inspected'] > 0:
                    model['ftq_percentage'] = round(((model['total_inspected'] - model['total_defects']) / model['total_inspected']) * 100, 2)
                else:
                    model['ftq_percentage'] = 0
            
            daily_total_inspected += shift['total_inspected']
            daily_total_defects += shift['total_defects']
        
        # Overall daily FTQ
        if daily_total_inspected > 0:
            daily_ftq = round(((daily_total_inspected - daily_total_defects) / daily_total_inspected) * 100, 2)
        else:
            daily_ftq = 0
        
        # Get all defects for this day
        defect_records = DefectRecord.objects.filter(
            ftq_record__date=report_date
        ).select_related('defect_type', 'defect_type_custom', 'ftq_record')
        
        # Group defects by type
        defects_by_type = {}
        for record in defect_records:
            defect_name = record.defect_type.name if record.defect_type else record.defect_type_custom.name
            if defect_name not in defects_by_type:
                defects_by_type[defect_name] = {
                    'name': defect_name,
                    'count': 0
                }
            defects_by_type[defect_name]['count'] += record.count
        
        # Sort defects by count (descending)
        defects_sorted = sorted(defects_by_type.values(), key=lambda x: x['count'], reverse=True)
        
        return render(request, 'main/operations/ftq_report_daily.html', {
            'report_date': report_date,
            'shifts': shifts,
            'daily_total_inspected': daily_total_inspected,
            'daily_total_defects': daily_total_defects,
            'daily_ftq': daily_ftq,
            'defects': defects_sorted,
            'title': f'Daily FTQ Report - {report_date}'
        })
    
    elif report_type == 'weekly':
        # Get week end date (default to Sunday of current week)
        end_date_str = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = today
        
        # Calculate week start (Monday)
        weekday = end_date.weekday()
        start_date = end_date - timedelta(days=weekday + 6)  # Go to Monday of the week
        
        # Get records for the week
        ftq_records = FTQRecord.objects.filter(
            date__range=[start_date, end_date]
        ).order_by('date', 'shift_type')
        
        # Group by day
        days = {}
        for i in range(7):
            current_date = start_date + timedelta(days=i)
            days[current_date] = {
                'date': current_date,
                'records': [],
                'total_inspected': 0,
                'total_defects': 0,
                'ftq_percentage': 0
            }
        
        # Add records to respective days
        for record in ftq_records:
            if record.date in days:
                days[record.date]['records'].append(record)
                days[record.date]['total_inspected'] += record.total_inspected
                days[record.date]['total_defects'] += record.total_defects
        
        # Calculate FTQ percentage for each day
        weekly_total_inspected = 0
        weekly_total_defects = 0
        
        for day in days.values():
            if day['total_inspected'] > 0:
                day['ftq_percentage'] = round(((day['total_inspected'] - day['total_defects']) / day['total_inspected']) * 100, 2)
            
            weekly_total_inspected += day['total_inspected']
            weekly_total_defects += day['total_defects']
        
        # Calculate weekly FTQ
        if weekly_total_inspected > 0:
            weekly_ftq = round(((weekly_total_inspected - weekly_total_defects) / weekly_total_inspected) * 100, 2)
        else:
            weekly_ftq = 0
        
        # Get top defects for the week
        top_defects = DefectRecord.objects.filter(
            ftq_record__date__range=[start_date, end_date]
        ).values(
            'defect_type__name'
        ).annotate(
            total_count=Sum('count')
        ).order_by('-total_count')[:10]
        
        # Chart data
        chart_dates = []
        chart_ftq = []
        
        for date, day in sorted(days.items()):
            chart_dates.append(date.strftime('%Y-%m-%d'))
            chart_ftq.append(day['ftq_percentage'])
        
        chart_data = {
            'dates': chart_dates,
            'ftq': chart_ftq
        }
        
        return render(request, 'main/operations/ftq_report_weekly.html', {
            'start_date': start_date,
            'end_date': end_date,
            'days': days,
            'weekly_total_inspected': weekly_total_inspected,
            'weekly_total_defects': weekly_total_defects,
            'weekly_ftq': weekly_ftq,
            'top_defects': top_defects,
            'chart_data': json.dumps(chart_data),
            'title': f'Weekly FTQ Report ({start_date} - {end_date})'
        })
    
    elif report_type == 'monthly':
        # Get month and year from request or use current
        month = int(request.GET.get('month', today.month))
        year = int(request.GET.get('year', today.year))
        
        # Validate month
        if month < 1 or month > 12:
            month = today.month
        
        # Get first and last day of the month
        first_day = datetime(year, month, 1).date()
        if month == 12:
            last_day = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        # Get records for the month
        ftq_records = FTQRecord.objects.filter(
            date__range=[first_day, last_day]
        ).order_by('date')
        
        # Group by model
        models = {}
        for record in ftq_records:
            model_key = record.model_name
            if model_key not in models:
                models[model_key] = {
                    'model_name': record.get_model_name_display(),
                    'total_inspected': 0,
                    'total_defects': 0
                }
            
            models[model_key]['total_inspected'] += record.total_inspected
            models[model_key]['total_defects'] += record.total_defects
        
        # Calculate FTQ percentage for each model
        monthly_total_inspected = 0
        monthly_total_defects = 0
        
        for model in models.values():
            if model['total_inspected'] > 0:
                model['ftq_percentage'] = round(((model['total_inspected'] - model['total_defects']) / model['total_inspected']) * 100, 2)
            else:
                model['ftq_percentage'] = 0
            
            monthly_total_inspected += model['total_inspected']
            monthly_total_defects += model['total_defects']
        
        # Calculate monthly FTQ
        if monthly_total_inspected > 0:
            monthly_ftq = round(((monthly_total_inspected - monthly_total_defects) / monthly_total_inspected) * 100, 2)
        else:
            monthly_ftq = 0
        
        # Get top defects for the month
        top_defects = DefectRecord.objects.filter(
            ftq_record__date__range=[first_day, last_day]
        ).values(
            'defect_type__name'
        ).annotate(
            total_count=Sum('count')
        ).order_by('-total_count')[:15]
        
        # Weekly trend within the month
        weekly_trend = []
        current_date = first_day
        while current_date <= last_day:
            # Find Sunday of the week
            days_to_sunday = 6 - current_date.weekday()
            if days_to_sunday < 0:
                days_to_sunday += 7
            
            end_of_week = current_date + timedelta(days=days_to_sunday)
            if end_of_week > last_day:
                end_of_week = last_day
            
            # Get records for this week
            week_records = ftq_records.filter(
                date__range=[current_date, end_of_week]
            )
            
            # Calculate weekly totals
            week_inspected = week_records.aggregate(total=Sum('total_inspected'))['total'] or 0
            week_defects = week_records.aggregate(total=Sum('total_defects'))['total'] or 0
            
            # Calculate FTQ
            if week_inspected > 0:
                week_ftq = round(((week_inspected - week_defects) / week_inspected) * 100, 2)
            else:
                week_ftq = 0
            
            weekly_trend.append({
                'start_date': current_date,
                'end_date': end_of_week,
                'total_inspected': week_inspected,
                'total_defects': week_defects,
                'ftq_percentage': week_ftq
            })
            
            # Move to next week
            current_date = end_of_week + timedelta(days=1)
        
        # Chart data for models
        chart_models = []
        chart_model_ftq = []
        
        for model_key, model in sorted(models.items()):
            chart_models.append(model['model_name'])
            chart_model_ftq.append(model['ftq_percentage'])
        
        chart_data = {
            'models': chart_models,
            'model_ftq': chart_model_ftq,
            'defect_names': [defect['defect_type__name'] for defect in top_defects],
            'defect_counts': [defect['total_count'] for defect in top_defects],
            'week_labels': [f'Week {i+1}' for i in range(len(weekly_trend))],
            'week_ftq': [week['ftq_percentage'] for week in weekly_trend]
        }
        
        return render(request, 'main/operations/ftq_report_monthly.html', {
            'year': year,
            'month': month,
            'month_name': datetime(year, month, 1).strftime('%B'),
            'models': models,
            'monthly_total_inspected': monthly_total_inspected,
            'monthly_total_defects': monthly_total_defects,
            'monthly_ftq': monthly_ftq,
            'top_defects': top_defects,
            'weekly_trend': weekly_trend,
            'chart_data': json.dumps(chart_data),
            'title': f'Monthly FTQ Report - {datetime(year, month, 1).strftime("%B %Y")}'
        })
    
    # Default case or unsupported report type
    messages.warning(request, f'Unsupported report type: {report_type}')
    return redirect('ftq_dashboard')

@login_required
def export_ftq_excel(request):
    """Export FTQ data to Excel"""
    # Get filter parameters
    start_date = request.GET.get('start_date', (timezone.now().date() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', timezone.now().date().strftime('%Y-%m-%d'))
    model_name = request.GET.get('model_name', '')
    shift_type = request.GET.get('shift_type', '')
    
    # Parse dates
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = timezone.now().date() - timedelta(days=30)
        end_date = timezone.now().date()
    
    # Query FTQ records
    ftq_records = FTQRecord.objects.select_related('shift', 'created_by', 'verified_by')
    
    # Apply filters
    ftq_records = ftq_records.filter(date__range=[start_date, end_date])
    
    if model_name:
        ftq_records = ftq_records.filter(model_name=model_name)
    
    if shift_type:
        ftq_records = ftq_records.filter(shift_type=shift_type)
    
    # Order by date
    ftq_records = ftq_records.order_by('date', 'shift_type')
    
    # Create Excel workbook
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "FTQ Report"
    
    # Define styles
    header_style = {
        'fill': PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid'),
        'font': Font(color='FFFFFF', bold=True),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    
    date_style = {
        'alignment': Alignment(horizontal='center')
    }
    
    number_style = {
        'alignment': Alignment(horizontal='right')
    }
    
    # Create header row
    headers = [
        'Date', 'Shift', 'Model', 'Julian Date', 'Production/Day', 
        'Total Reject', 'FTQ %', 'Created By', 'Verified By'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_style['fill']
        cell.font = header_style['font']
        cell.alignment = header_style['alignment']
    
    # Add data rows
    for row_num, record in enumerate(ftq_records, 2):
        # Calculate FTQ percentage
        if record.total_inspected > 0:
            ftq_percentage = ((record.total_inspected - record.total_defects) / record.total_inspected) * 100
        else:
            ftq_percentage = 0
        
        # Format verified_by
        verified_by = record.verified_by.get_full_name() if record.verified_by else 'Not Verified'
        
        # Add row data
        row_data = [
            record.date,
            record.get_shift_type_display(),
            record.get_model_name_display(),
            record.julian_date,
            record.total_inspected,
            record.total_defects,
            f"{ftq_percentage:.2f}%",
            record.created_by.get_full_name(),
            verified_by
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            
            # Apply styles
            if col_num == 1:  # Date column
                cell.alignment = date_style['alignment']
            elif col_num in [5, 6, 7]:  # Numeric columns
                cell.alignment = number_style['alignment']
    
    # Auto adjust column widths
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15
    
    # Add summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    
    # Add summary headers
    summary_headers = ['Metric', 'Value']
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_num, value=header)
        cell.fill = header_style['fill']
        cell.font = header_style['font']
        cell.alignment = header_style['alignment']
    
    # Calculate summary statistics
    total_inspected = ftq_records.aggregate(total=Sum('total_inspected'))['total'] or 0
    total_defects = ftq_records.aggregate(total=Sum('total_defects'))['total'] or 0
    
    if total_inspected > 0:
        overall_ftq = ((total_inspected - total_defects) / total_inspected) * 100
    else:
        overall_ftq = 0
    
    # Add summary data
    summary_data = [
        ['Date Range', f"{start_date} to {end_date}"],
        ['Total FTQ Records', ftq_records.count()],
        ['Total Production', total_inspected],
        ['Total Rejects', total_defects],
        ['Overall FTQ', f"{overall_ftq:.2f}%"]
    ]
    
    for row_num, (metric, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row_num, column=1, value=metric)
        ws_summary.cell(row=row_num, column=2, value=value)
    
    # Auto adjust column widths
    for col_num, _ in enumerate(summary_headers, 1):
        col_letter = get_column_letter(col_num)
        ws_summary.column_dimensions[col_letter].width = 20
    
    # Add defects sheet
    ws_defects = wb.create_sheet(title="Defects")
    
    # Get defect records
    defect_records = DefectRecord.objects.filter(
        ftq_record__in=ftq_records
    ).select_related('defect_type', 'defect_type_custom', 'ftq_record')
    
    # Group defects by type
    defects_by_type = {}
    for record in defect_records:
        defect_name = record.defect_type.name if record.defect_type else record.defect_type_custom.name
        if defect_name not in defects_by_type:
            defects_by_type[defect_name] = 0
        defects_by_type[defect_name] += record.count
    
    # Sort defects by count
    sorted_defects = sorted(defects_by_type.items(), key=lambda x: x[1], reverse=True)
    
    # Add defect headers
    defect_headers = ['Defect Type', 'Count', '% of Total']
    for col_num, header in enumerate(defect_headers, 1):
        cell = ws_defects.cell(row=1, column=col_num, value=header)
        cell.fill = header_style['fill']
        cell.font = header_style['font']
        cell.alignment = header_style['alignment']
    
    # Add defect data
    for row_num, (defect_name, count) in enumerate(sorted_defects, 2):
        percentage = (count / total_defects * 100) if total_defects > 0 else 0
        
        ws_defects.cell(row=row_num, column=1, value=defect_name)
        ws_defects.cell(row=row_num, column=2, value=count)
        ws_defects.cell(row=row_num, column=3, value=f"{percentage:.2f}%")
    
    # Auto adjust column widths
    for col_num, _ in enumerate(defect_headers, 1):
        col_letter = get_column_letter(col_num)
        ws_defects.column_dimensions[col_letter].width = 20
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ftq_report_{start_date}_to_{end_date}.xlsx'
    
    wb.save(response)
    return response 



# new code 


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Sum
from datetime import datetime, timedelta, date

from .models import DTPMChecklistFMA03, DTPMCheckResult, DTPMIssue, Shift, User
from .forms import DTPMChecklistForm, DTPMCheckResultForm, DTPMCheckResultInlineFormSet, DTPMIssueForm, DTPMIssueResolveForm, DTPMChecklistFilterForm

# DTPM List View
@login_required
def dtpm_list(request):
    """List all DTPM checklists with filtering options"""
    form = DTPMChecklistFilterForm(request.GET)
    checklists = DTPMChecklistFMA03.objects.select_related('shift', 'operator', 'supervisor')
    
    # Apply filters if form is valid
    if form.is_valid():
        data = form.cleaned_data
        if data.get('date_from'):
            checklists = checklists.filter(date__gte=data['date_from'])
        if data.get('date_to'):
            checklists = checklists.filter(date__lte=data['date_to'])
        if data.get('shift'):
            checklists = checklists.filter(shift__shift_type=data['shift'])
        if data.get('status'):
            checklists = checklists.filter(status=data['status'])
    
    # Order by date, newest first
    checklists = checklists.order_by('-date')
    
    # Calculate completion statistics
    stats = {
        'total': checklists.count(),
        'completed': checklists.filter(status__in=['verified', 'rejected']).count(),
        'pending': checklists.filter(status='pending').count(),
        'issues_count': DTPMIssue.objects.filter(check_result__checklist__in=checklists).count()
    }
    
    # Calculate completion percentage
    if stats['total'] > 0:
        stats['completion_percentage'] = (stats['completed'] / stats['total']) * 100
    else:
        stats['completion_percentage'] = 0
    
    context = {
        'checklists': checklists,
        'filter_form': form,
        'stats': stats,
        'title': 'DTPM FMA03 Checklists'
    }
    
    return render(request, 'main/operations/dtpm_list.html', context)

# DTPM Detail View
@login_required
def dtpm_detail(request, pk):
    """View details of a specific DTPM checklist"""
    checklist = get_object_or_404(
        DTPMChecklistFMA03.objects.select_related('shift', 'operator', 'supervisor'),
        pk=pk
    )
    
    # Get all check results for this checklist
    check_results = DTPMCheckResult.objects.filter(checklist=checklist).order_by('item_number')
    
    # Get issues for this checklist
    issues = DTPMIssue.objects.filter(
        check_result__checklist=checklist
    ).select_related('check_result', 'reported_by', 'assigned_to').order_by('-created_at')
    
    # Check if user can edit the checklist
    can_edit = (request.user == checklist.operator and checklist.status == 'pending')
    
    # Check if user can verify the checklist
    can_verify = (request.user == checklist.supervisor and checklist.status == 'pending')
    
    context = {
        'checklist': checklist,
        'check_results': check_results,
        'issues': issues,
        'can_edit': can_edit,
        'can_verify': can_verify,
        'title': f'DTPM Checklist - {checklist.date}'
    }
    
    return render(request, 'main/operations/dtpm_detail.html', context)

@login_required
@user_passes_test(lambda u: u.user_type == 'operator')
def dtpm_create(request):
    """Create a new DTPM checklist"""
    # Check if user is a supervisor
    is_supervisor = request.user.user_type in ['shift_supervisor', 'quality_supervisor'] or request.user.is_superuser
    
    if request.method == 'POST':
        form = DTPMChecklistForm(request.POST, request.FILES, is_supervisor=is_supervisor)
        if form.is_valid():
            try:
                # Get the current shift
                current_shift = get_current_shift(request.user)
                
                checklist = form.save(commit=False)
                checklist.shift = current_shift
                checklist.operator = request.user
                checklist.supervisor = current_shift.shift_supervisor
                checklist.status = 'pending'
                checklist.save()
                
                # Redirect directly to the edit_checks view
                messages.success(request, 'DTPM checklist created successfully. Please fill in the check results.')
                return redirect('operator_dashboard', pk=checklist.pk)
            except Exception as e:
                messages.error(request, f"Error creating checklist: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # For GET requests, just render the form without checking for existing checklists
        form = DTPMChecklistForm(is_supervisor=is_supervisor)
    
    context = {
        'form': form,
        'title': 'Create DTPM Checklist',
        'current_shift': f"{'Day' if 8 <= timezone.now().hour < 20 else 'Night'} Shift",
        'is_supervisor': is_supervisor
    }
    
    return render(request, 'main/operations/dtpm_form.html', context)
# DTPM Edit View
@login_required
def dtpm_edit(request, pk):
    """Edit a DTPM checklist's basic information"""
    checklist = get_object_or_404(DTPMChecklistFMA03, pk=pk)
    
    # Check if user is a supervisor
    is_supervisor = request.user.user_type in ['shift_supervisor', 'quality_supervisor'] or request.user.is_superuser
    
    # Verify permission to edit
    if not is_supervisor and checklist.operator != request.user:
        messages.error(request, 'You do not have permission to edit this checklist.')
        return redirect('dtpm_detail', pk=checklist.pk)
        
    if checklist.status != 'pending':
        messages.error(request, 'Cannot edit a verified or rejected checklist.')
        return redirect('dtpm_detail', pk=checklist.pk)
    
    if request.method == 'POST':
        form = DTPMChecklistForm(request.POST, request.FILES, instance=checklist, is_supervisor=is_supervisor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Checklist information updated successfully.')
            return redirect('dtpm_detail', pk=checklist.pk)
    else:
        form = DTPMChecklistForm(instance=checklist, is_supervisor=is_supervisor)
    
    context = {
        'form': form,
        'checklist': checklist,
        'title': 'Edit DTPM Checklist',
        'is_supervisor': is_supervisor
    }
    
    return render(request, 'main/operations/dtpm_form.html', context)

# DTPM Delete View
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def dtpm_delete(request, pk):
    """Delete a DTPM checklist"""
    checklist = get_object_or_404(DTPMChecklistFMA03, pk=pk)
    
    if request.method == 'POST':
        checklist.delete()
        messages.success(request, 'DTPM checklist deleted successfully.')
        return redirect('dtpm_list')
    
    context = {
        'checklist': checklist,
        'title': 'Delete DTPM Checklist'
    }
    
    return render(request, 'main/operations/dtpm_confirm_delete.html', context)

# DTPM Edit Check Results View
@login_required
def dtpm_edit_checks(request, pk):
    """Edit all check results for a DTPM checklist"""
    checklist = get_object_or_404(DTPMChecklistFMA03, pk=pk)
    
    # Verify permission to edit
    if checklist.status != 'pending':
        messages.error(request, 'Cannot edit a verified or rejected checklist.')
        return redirect('dtpm_detail', pk=checklist.pk)
        
    if request.user != checklist.operator and not request.user.is_superuser:
        if request.user != checklist.supervisor:
            messages.error(request, 'You do not have permission to edit this checklist.')
            return redirect('dtpm_detail', pk=checklist.pk)
    
    if request.method == 'POST':
        formset = DTPMCheckResultInlineFormSet(request.POST, request.FILES, instance=checklist)
        if formset.is_valid():
            # Save the formset instances
            instances = formset.save(commit=False)
            
            for instance in instances:
                # Set the user who performed the check
                instance.checked_by = request.user
                instance.checked_at = timezone.now()
                
                # Preserve existing image if no new image is uploaded
                if not instance.image:
                    try:
                        existing_instance = DTPMCheckResult.objects.get(pk=instance.pk)
                        if existing_instance.image:
                            instance.image = existing_instance.image
                    except DTPMCheckResult.DoesNotExist:
                        pass
                
                # Manual handling of result field from the custom form
                result_key = f"{formset.prefix}-{instance.id}-result"
                if result_key in request.POST:
                    instance.result = request.POST[result_key]
                
                instance.save()
            
            messages.success(request, 'Check results updated successfully.')
            return redirect('dtpm_detail', pk=checklist.pk)
        else:
            for form in formset:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"Error in {form.instance.get_item_number_display}: {field} - {error}")
    else:
        formset = DTPMCheckResultInlineFormSet(instance=checklist)
    
    context = {
        'checklist': checklist,
        'formset': formset,
        'title': 'Edit Check Results'
    }
    
    return render(request, 'main/operations/dtpm_check_results_form.html', context)

# DTPM Verify View
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'])
def dtpm_verify(request, pk):
    """Verify or reject a DTPM checklist"""
    checklist = get_object_or_404(DTPMChecklistFMA03, pk=pk)
    
    # Verify permission to verify
    if checklist.status != 'pending':
        messages.error(request, 'This checklist has already been verified or rejected.')
        return redirect('dtpm_detail', pk=checklist.pk)
        
    if request.user != checklist.supervisor:
        messages.error(request, 'Only the assigned supervisor can verify this checklist.')
        return redirect('dtpm_detail', pk=checklist.pk)
    
    if request.method == 'POST':
        action = request.POST.get('action', 'verify')
        
        # Check if all required check results have been filled
        incomplete_checks = DTPMCheckResult.objects.filter(checklist=checklist, result='').exists()
        
        if incomplete_checks and action == 'verify':
            messages.error(request, 'Cannot verify checklist with incomplete check results.')
            return redirect('dtpm_edit_checks', pk=checklist.pk)
        
        # Update checklist status
        if action == 'verify':
            checklist.status = 'verified'
            success_message = 'Checklist verified successfully.'
        else:
            checklist.status = 'rejected'
            success_message = 'Checklist rejected.'
            
        checklist.save()
        messages.success(request, success_message)
        return redirect('dtpm_detail', pk=checklist.pk)
    
    # Get all check results for verification review
    check_results = DTPMCheckResult.objects.filter(checklist=checklist).order_by('item_number')
    
    context = {
        'checklist': checklist,
        'check_results': check_results,
        'title': 'Verify DTPM Checklist'
    }
    
    return render(request, 'main/operations/dtpm_verify.html', context)

# DTPM Report Issue View
@login_required
def dtpm_report_issue(request, check_id):
    """Report an issue for a specific check result"""
    check_result = get_object_or_404(DTPMCheckResult.objects.select_related('checklist'), pk=check_id)
    
    if request.method == 'POST':
        form = DTPMIssueForm(request.POST, request.FILES)
        if form.is_valid():
            issue = form.save(commit=False)
            issue.check_result = check_result
            issue.reported_by = request.user
            issue.save()
            
            messages.success(request, 'Issue reported successfully.')
            return redirect('dtpm_detail', pk=check_result.checklist.pk)
    else:
        form = DTPMIssueForm()
    
    context = {
        'form': form,
        'check_result': check_result,
        'title': 'Report Issue'
    }
    
    return render(request, 'main/operations/dtpm_issue_form.html', context)

# DTPM Resolve Issue View
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'])
def dtpm_resolve_issue(request, issue_id):
    """Resolve a reported issue"""
    issue = get_object_or_404(DTPMIssue.objects.select_related('check_result__checklist'), pk=issue_id)
    
    if request.method == 'POST':
        form = DTPMIssueResolveForm(request.POST, request.FILES, instance=issue)
        if form.is_valid():
            resolved_issue = form.save(commit=False)
            resolved_issue.assigned_to = request.user
            resolved_issue.save()
            
            messages.success(request, 'Issue updated successfully.')
            return redirect('dtpm_detail', pk=issue.check_result.checklist.pk)
    else:
        form = DTPMIssueResolveForm(instance=issue)
    
    context = {
        'form': form,
        'issue': issue,
        'title': 'Resolve Issue'
    }
    
    return render(request, 'main/operations/dtpm_issue_resolve_form.html', context)

# DTPM Dashboard View
@login_required
def dtpm_dashboard(request):
    """Dashboard with DTPM statistics and charts"""
    # Get date range for filtering
    days = int(request.GET.get('days', 30))
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=days)
    
    # Get checklists in date range
    checklists = DTPMChecklistFMA03.objects.filter(
        date__range=[start_date, end_date]
    ).select_related('shift', 'operator')
    
    # Calculate summary statistics
    stats = {
        'total_checklists': checklists.count(),
        'verified_count': checklists.filter(status='verified').count(),
        'rejected_count': checklists.filter(status='rejected').count(),
        'pending_count': checklists.filter(status='pending').count(),
    }
    
    if stats['total_checklists'] > 0:
        stats['completion_rate'] = ((stats['verified_count'] + stats['rejected_count']) / stats['total_checklists']) * 100
        stats['success_rate'] = (stats['verified_count'] / (stats['verified_count'] + stats['rejected_count'])) * 100 if (stats['verified_count'] + stats['rejected_count']) > 0 else 0
    else:
        stats['completion_rate'] = 0
        stats['success_rate'] = 0
    
    # Calculate statistics by check item
    check_items = {}
    for i in range(1, 8):  # For all 7 check items
        # Get all results for this item
        results = DTPMCheckResult.objects.filter(
            checklist__in=checklists,
            item_number=i
        )
        
        total = results.count()
        ok_count = results.filter(result='OK').count()
        ng_count = results.filter(result='NG').count()
        
        if total > 0:
            ok_rate = (ok_count / total) * 100
        else:
            ok_rate = 0
            
        description = DTPMCheckResult.CHECK_ITEMS[i-1][1]
        
        check_items[i] = {
            'description': description,
            'total': total,
            'ok_count': ok_count,
            'ng_count': ng_count,
            'ok_rate': ok_rate
        }
    
    # Get open issues
    open_issues = DTPMIssue.objects.filter(
        check_result__checklist__in=checklists,
        status__in=['open', 'in_progress']
    ).select_related('check_result', 'reported_by').order_by('-created_at')
    
    # Prepare chart data
    chart_data = {
        'labels': [str(i) for i in range(1, 8)],  # Check item numbers
        'ok_rates': [check_items[i]['ok_rate'] for i in range(1, 8)],
        'item_descriptions': [check_items[i]['description'][:30] + '...' if len(check_items[i]['description']) > 30 else check_items[i]['description'] for i in range(1, 8)]
    }
    
    context = {
        'stats': stats,
        'check_items': check_items,
        'open_issues': open_issues,
        'chart_data': chart_data,
        'days': days,
        'start_date': start_date,
        'end_date': end_date,
        'title': 'DTPM Dashboard'
    }
    
    return render(request, 'main/operations/dtpm_dashboard.html', context)

# Export DTPM to Excel
@login_required
def export_dtpm_excel(request, pk):
    """Export a DTPM checklist to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    checklist = get_object_or_404(
        DTPMChecklistFMA03.objects.select_related('shift', 'operator', 'supervisor'),
        pk=pk
    )
    
    # Get all check results for this checklist
    check_results = DTPMCheckResult.objects.filter(checklist=checklist).order_by('item_number')
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DTPM Checklist"
    
    # Define styles
    header_style = {
        'fill': PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid'),
        'font': Font(color='FFFFFF', bold=True),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    
    ok_style = {
        'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'font': Font(color='006100'),
        'alignment': Alignment(horizontal='center')
    }
    
    ng_style = {
        'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'font': Font(color='9C0006'),
        'alignment': Alignment(horizontal='center')
    }
    
    # Add title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "DTPM FMA03 Operation 35 Checklist"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Add basic info
    ws['A3'] = "Date:"
    ws['B3'] = checklist.date
    
    ws['D3'] = "Shift:"
    ws['E3'] = checklist.shift.get_shift_type_display()
    
    ws['A4'] = "Operator:"
    ws['B4'] = checklist.operator.username
    
    ws['D4'] = "Status:"
    ws['E4'] = checklist.get_status_display()
    
    # Add headers for check items
    headers = ['Item #', 'Description', 'Result', 'Comments', 'Checked By', 'Checked At']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.fill = header_style['fill']
        cell.font = header_style['font']
        cell.alignment = header_style['alignment']
    
    # Add check results
    for row_idx, check in enumerate(check_results, 7):
        ws.cell(row=row_idx, column=1, value=check.item_number)
        
        # Get description from the CHECK_ITEMS choices
        description = next((desc for code, desc in DTPMCheckResult.CHECK_ITEMS if code == check.item_number), '')
        ws.cell(row=row_idx, column=2, value=description)
        
        # Add result with color-coding
        result_cell = ws.cell(row=row_idx, column=3, value=check.result)
        if check.result == 'OK':
            result_cell.fill = ok_style['fill']
            result_cell.font = ok_style['font']
            result_cell.alignment = ok_style['alignment']
        elif check.result == 'NG':
            result_cell.fill = ng_style['fill']
            result_cell.font = ng_style['font']
            result_cell.alignment = ng_style['alignment']
        
        ws.cell(row=row_idx, column=4, value=check.comments)
        ws.cell(row=row_idx, column=5, value=check.checked_by.username if check.checked_by else '')
        ws.cell(row=row_idx, column=6, value=check.checked_at)
    
    # Add issues section if there are any
    issues = DTPMIssue.objects.filter(check_result__checklist=checklist).select_related('check_result', 'reported_by')
    
    if issues.exists():
        # Add header for issues section
        row_idx = len(check_results) + 9  # Leave a gap after check results
        
        ws.merge_cells(f'A{row_idx}:G{row_idx}')
        issues_header = ws[f'A{row_idx}']
        issues_header.value = "Issues"
        issues_header.font = Font(size=12, bold=True)
        
        # Add issue headers
        issue_headers = ['Check Item', 'Description', 'Priority', 'Status', 'Reported By', 'Created At']
        row_idx += 1
        
        for col_idx, header in enumerate(issue_headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.fill = header_style['fill']
            cell.font = header_style['font']
            cell.alignment = header_style['alignment']
        
        # Add issues
        for issue_idx, issue in enumerate(issues, 1):
            row_idx += 1
            
            ws.cell(row=row_idx, column=1, value=issue.check_result.item_number)
            ws.cell(row=row_idx, column=2, value=issue.description)
            ws.cell(row=row_idx, column=3, value=issue.get_priority_display())
            ws.cell(row=row_idx, column=4, value=issue.get_status_display())
            ws.cell(row=row_idx, column=5, value=issue.reported_by.username if issue.reported_by else '')
            ws.cell(row=row_idx, column=6, value=issue.created_at)
    
    # Adjust column widths
    col_widths = [10, 50, 10, 30, 15, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=dtpm_checklist_{checklist.date}.xlsx'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def dtpm_manage_images(request, pk):
    """Special view for supervisors to manage check result images"""
    checklist = get_object_or_404(DTPMChecklistFMA03, pk=pk)
    
    if request.method == 'POST':
        # Handle image uploads
        try:
            for key, file in request.FILES.items():
                if key.startswith('check_image_'):
                    check_id = key.split('_')[-1]
                    check = get_object_or_404(DTPMCheckResult, id=check_id, checklist=checklist)
                    check.image = file
                    check.save()
            
            # Handle header image upload
            if 'header_image' in request.FILES:
                checklist.header_image = request.FILES['header_image']
                checklist.save()
                
            messages.success(request, 'Images updated successfully.')
            return redirect('dtpm_detail', pk=checklist.pk)
        except Exception as e:
            messages.error(request, f'Error saving images: {str(e)}')
    
    # Get all check results for this checklist
    check_results = DTPMCheckResult.objects.filter(checklist=checklist).order_by('item_number')
    
    context = {
        'checklist': checklist,
        'check_results': check_results,
        'title': 'Manage Check Images'
    }
    
    return render(request, 'main/operations/dtpm_manage_images.html', context)






#  new code 
# Views for Error Prevention functionality

@login_required
def ep_check_list(request):
    """List all Error Prevention checks with filtering options"""
    form = ErrorPreventionFilterForm(request.GET)
    
    # Start with all EP checks
    ep_checks = ErrorPreventionCheck.objects.select_related(
        'verification_status__shift', 
        'operator', 
        'supervisor', 
        'quality_supervisor'
    )
    
    # Apply filters if form is valid
    if form.is_valid():
        data = form.cleaned_data
        if data.get('date_from'):
            ep_checks = ep_checks.filter(date__gte=data['date_from'])
        if data.get('date_to'):
            ep_checks = ep_checks.filter(date__lte=data['date_to'])
        if data.get('model'):
            ep_checks = ep_checks.filter(current_model=data['model'])
        if data.get('status'):
            ep_checks = ep_checks.filter(status=data['status'])
    
    # Order by date, newest first
    ep_checks = ep_checks.order_by('-date')
    
    # Calculate stats
    stats = {
        'total': ep_checks.count(),
        'pending': ep_checks.filter(status='pending').count(),
        'approved': ep_checks.filter(status__in=['supervisor_approved', 'quality_approved']).count(),
        'rejected': ep_checks.filter(status='rejected').count(),
    }
    
    context = {
        'ep_checks': ep_checks,
        'filter_form': form,
        'stats': stats,
        'title': 'Error Prevention Checks'
    }
    
    return render(request, 'main/operations/ep_check_list.html', context)


from .models import ErrorPreventionCheck, ErrorPreventionMechanismStatus
from .history_utils import create_initial_history, get_ep_check_timeline


@login_required
def ep_check_detail(request, pk):
    """View a specific Error Prevention check with change history - updated permissions"""
    ep_check = get_object_or_404(
        ErrorPreventionCheck.objects.select_related(
            'verification_status__shift', 
            'operator', 
            'supervisor', 
            'quality_supervisor'
        ),
        pk=pk
    )
    
    # Get all mechanism statuses for this check, ordered by mechanism display order
    mechanism_statuses = ErrorPreventionMechanismStatus.objects.filter(
        ep_check=ep_check
    ).select_related('mechanism').order_by(
        'mechanism__display_order',
        'mechanism__mechanism_id'
    )
    
    # Get change history timeline
    timeline = get_ep_check_timeline(ep_check)
    
    # Get recent changes (last 24 hours)
    from datetime import timedelta
    recent_cutoff = timezone.now() - timedelta(hours=24)
    recent_changes = [change for change in timeline if change['timestamp'] >= recent_cutoff]
    
    # Updated permission logic
    # Operators can edit when status is 'pending' OR 'rejected'
    can_edit = (
        request.user == ep_check.operator and 
        ep_check.status in ['pending', 'rejected']
    )
    
    # Supervisors can verify at any time
    can_verify_supervisor = (request.user.user_type == 'shift_supervisor')
    
    # Quality supervisors can verify at any time
    can_verify_quality = (request.user.user_type == 'quality_supervisor')
    
    context = {
        'ep_check': ep_check,
        'mechanism_statuses': mechanism_statuses,
        'can_edit': can_edit,
        'can_verify_supervisor': can_verify_supervisor,
        'can_verify_quality': can_verify_quality,
        'timeline': timeline[:10],  # Show last 10 changes
        'recent_changes': recent_changes,
        'has_recent_changes': len(recent_changes) > 0,
        'title': f'EP Check - {ep_check.date}'
    }
    
    return render(request, 'main/operations/ep_check_detail.html', context)




@login_required
@user_passes_test(lambda u: u.user_type == 'operator')
def ep_check_edit(request, pk):
    """Edit an existing Error Prevention check - operators can edit when pending or rejected"""
    ep_check = get_object_or_404(ErrorPreventionCheck, pk=pk)
    
    # Updated permission check: allow editing when pending OR rejected
    if request.user != ep_check.operator:
        messages.error(request, 'You can only edit your own EP checks.')
        return redirect('ep_check_detail', pk=pk)
    
    if ep_check.status not in ['pending', 'rejected']:
        messages.error(request, 'You can only edit EP checks that are pending or rejected.')
        return redirect('ep_check_detail', pk=pk)
    
    # Get all mechanism statuses for this check, ordered by mechanism display order
    mechanism_statuses = ep_check.mechanism_statuses.select_related('mechanism').order_by(
        'mechanism__display_order', 
        'mechanism__mechanism_id'
    )
    
    if request.method == 'POST':
        form = ErrorPreventionCheckForm(
            request.POST,
            instance=ep_check,
            verification_status=ep_check.verification_status,
            user=request.user
        )
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Save the main EP check
                    ep_check = form.save()
                    
                    # Track changes for history
                    changes_made = []
                    
                    # Update mechanism statuses from POST data
                    for status in mechanism_statuses:
                        if not status.mechanism:
                            continue
                            
                        # Use mechanism ID as the key (consistent with create view)
                        mech_key = f"mech_{status.mechanism.id}"
                        
                        # Get form data
                        new_status = request.POST.get(f'status_{mech_key}', '')
                        new_is_na = request.POST.get(f'is_na_{mech_key}') == 'on'
                        new_comments = request.POST.get(f'comments_{mech_key}', '')
                        new_alternative = request.POST.get(f'alternative_method_{mech_key}', '')
                        
                        # Track changes
                        if status.status != new_status and not new_is_na:
                            ErrorPreventionMechanismHistory.objects.create(
                                mechanism_status=status,
                                changed_by=request.user,
                                field_name='status',
                                old_value=status.status,
                                new_value=new_status
                            )
                            changes_made.append(f"{status.mechanism.mechanism_id}: {status.status} → {new_status}")
                        
                        if status.is_not_applicable != new_is_na:
                            ErrorPreventionMechanismHistory.objects.create(
                                mechanism_status=status,
                                changed_by=request.user,
                                field_name='is_not_applicable',
                                old_value=str(status.is_not_applicable),
                                new_value=str(new_is_na)
                            )
                        
                        if status.comments != new_comments:
                            ErrorPreventionMechanismHistory.objects.create(
                                mechanism_status=status,
                                changed_by=request.user,
                                field_name='comments',
                                old_value=status.comments,
                                new_value=new_comments
                            )
                        
                        # Update the status
                        status.status = '' if new_is_na else new_status
                        status.is_not_applicable = new_is_na
                        status.comments = new_comments
                        # Alternative method can be edited by operator
                        status.alternative_method = new_alternative
                        # is_working remains synced from master (not editable)
                        status.last_edited_by = request.user
                        status.last_edited_at = timezone.now()
                        status.save()
                    
                    # Create EP check history entry
                    ErrorPreventionCheckHistory.objects.create(
                        ep_check=ep_check,
                        changed_by=request.user,
                        action='updated',
                        description=f'EP check updated. Changes: {", ".join(changes_made) if changes_made else "No mechanism changes"}',
                        additional_data={'changes': changes_made}
                    )
                    
                    messages.success(request, 'EP check updated successfully!')
                    return redirect('ep_check_detail', pk=ep_check.pk)
                    
            except Exception as e:
                messages.error(request, f"Error updating EP check: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # Initialize form with existing data
        form = ErrorPreventionCheckForm(
            instance=ep_check,
            verification_status=ep_check.verification_status,
            user=request.user
        )
    
    context = {
        'form': form,
        'ep_check': ep_check,
        'mechanism_statuses': mechanism_statuses,
        'title': 'Edit Error Prevention Check',
    }
    
    return render(request, 'main/operations/ep_check_edit.html', context)

@login_required
@user_passes_test(lambda u: u.user_type == 'operator')
def ep_check_create(request):
    """Create a new Error Prevention check with dynamic mechanisms from master list"""
    
    # Get the active verification status
    active_verification = DailyVerificationStatus.objects.filter(
        created_by=request.user,
        date=timezone.now().date(),
        status__in=['pending', 'in_progress']
    ).first()
    
    if not active_verification:
        messages.warning(request, 'You need to create a Daily Verification Sheet first.')
        return redirect('create_checklist')
    
    # Check if EP check already exists
    existing_check = ErrorPreventionCheck.objects.filter(
        verification_status=active_verification
    ).first()
    
    if existing_check:
        messages.warning(request, 'An EP check already exists for this verification status.')
        return redirect('ep_check_detail', pk=existing_check.pk)
    
    # Get active checklist
    active_checklist = ChecklistBase.objects.filter(
        verification_status=active_verification
    ).first()
    
    if not active_checklist:
        messages.error(request, 'No checklist found. Please create a checklist first.')
        return redirect('create_checklist')
    
    # Get current model to filter mechanisms
    current_model = active_checklist.selected_model
    
    # Get applicable mechanisms for this model
    mechanisms = ErrorPreventionMechanism.objects.filter(
        is_active=True,
        applicable_models__icontains=current_model
    ).order_by('display_order')
    
    if request.method == 'POST':
        form = ErrorPreventionCheckForm(
            request.POST,
            verification_status=active_verification,
            user=request.user
        )
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Create the main EP check
                    ep_check = form.save()
                    
                    # Create mechanism statuses from POST data
                    for mechanism in mechanisms:
                        # Use mechanism ID as the key
                        mech_key = f"mech_{mechanism.id}"
                        
                        # Get form data
                        status = request.POST.get(f'status_{mech_key}', '')
                        is_na = request.POST.get(f'is_na_{mech_key}') == 'on'
                        comments = request.POST.get(f'comments_{mech_key}', '')
                        
                        # Create mechanism status
                        ErrorPreventionMechanismStatus.objects.create(
                            ep_check=ep_check,
                            mechanism=mechanism,
                            status='' if is_na else status,
                            is_not_applicable=is_na,
                            is_working=mechanism.is_currently_working,
                            alternative_method=mechanism.default_alternative_method,
                            comments=comments,
                            last_edited_by=request.user
                        )
                    
                    messages.success(
                        request, 
                        f'EP check created successfully with {mechanisms.count()} mechanisms!'
                    )
                    return redirect('operator_dashboard')
                    
            except Exception as e:
                messages.error(request, f"Error creating EP check: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # Initialize form
        initial_data = {'date': active_verification.date}
        form = ErrorPreventionCheckForm(
            initial=initial_data,
            verification_status=active_verification,
            user=request.user
        )
    
    context = {
        'form': form,
        'title': 'Create Error Prevention Check',
        'active_verification': active_verification,
        'active_checklist': active_checklist,
        'current_model': current_model,
        'current_shift': active_verification.shift.get_shift_type_display() if active_verification.shift else 'N/A',
        'mechanisms': mechanisms,  # Pass mechanisms to template
        'mechanisms_count': mechanisms.count(),
    }
    
    return render(request, 'main/operations/ep_check_create.html', context)


@login_required
def ep_check_edit_statuses(request, pk):
    """Edit all mechanism statuses for an EP check"""
    ep_check = get_object_or_404(ErrorPreventionCheck, pk=pk)
    
    # Verify permission to edit
    if ep_check.status != 'pending':
        messages.error(request, 'Cannot edit a verified or rejected EP check.')
        return redirect('ep_check_detail', pk=ep_check.pk)
        
    if ep_check.operator != request.user and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to edit this EP check.')
        return redirect('ep_check_detail', pk=ep_check.pk)
    
    if request.method == 'POST':
        formset = ErrorPreventionStatusFormSet(request.POST, instance=ep_check)
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Mechanism statuses updated successfully.')
            return redirect('ep_check_detail', pk=ep_check.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        formset = ErrorPreventionStatusFormSet(instance=ep_check)
    
    # Get mechanism details for displaying names
    mechanism_choices = dict(ErrorPreventionCheck.EP_MECHANISM_CHOICES)
    
    context = {
        'ep_check': ep_check,
        'formset': formset,
        'mechanism_choices': mechanism_choices,
        'title': 'Edit Mechanism Statuses'
    }
    
    return render(request, 'main/operations/ep_check_statuses_form.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'shift_supervisor')
def ep_check_verify_supervisor(request, pk):
    """Verify an EP check as a supervisor - can verify/re-verify at any time"""
    ep_check = get_object_or_404(ErrorPreventionCheck, pk=pk)
    
    # Supervisors can verify at any time - no status restrictions
    if request.method == 'POST':
        action = request.POST.get('action', 'approve')
        comments = request.POST.get('comments', '')
        
        # Check if all mechanisms have a status
        incomplete_statuses = ErrorPreventionMechanismStatus.objects.filter(
            ep_check=ep_check, 
            status='',
            is_not_applicable=False
        ).exists()
        
        if incomplete_statuses and action == 'approve':
            messages.error(request, 'Cannot approve with incomplete status entries. Please mark non-applicable items as N/A or provide a status.')
            return redirect('ep_check_detail', pk=ep_check.pk)
        
        # Store previous status for messaging
        previous_status = ep_check.status
        
        # Update check status
        if action == 'approve':
            ep_check.status = 'supervisor_approved'
            ep_check.supervisor = request.user
            ep_check.comments = comments
            
            # Clear any previous rejection info
            if hasattr(ep_check, 'rejected_at'):
                ep_check.rejected_at = None
            if hasattr(ep_check, 'rejected_by'):
                ep_check.rejected_by = None
            
            # Clear quality verification if re-approving (needs fresh quality review)
            if previous_status in ['quality_approved']:
                ep_check.quality_supervisor = None
                if hasattr(ep_check, 'quality_approved_at'):
                    ep_check.quality_approved_at = None
                success_message = 'EP check re-approved by supervisor. Previous quality verification cleared - requires fresh quality review.'
            else:
                success_message = 'EP check approved by supervisor.'
            
            # Update verification status to notify quality supervisor
            if ep_check.verification_status:
                verification_status = ep_check.verification_status
                verification_status.quality_notified = True
                verification_status.save()
            
        else:  # reject action
            ep_check.status = 'rejected'
            ep_check.supervisor = request.user
            ep_check.comments = comments
            
            # Clear any approvals when rejecting
            ep_check.quality_supervisor = None
            if hasattr(ep_check, 'quality_approved_at'):
                ep_check.quality_approved_at = None
            
            # Update verification status
            if ep_check.verification_status:
                verification_status = ep_check.verification_status
                verification_status.status = 'rejected'
                verification_status.quality_notified = False
                verification_status.save()
            
            success_message = 'EP check rejected by supervisor. Operator can edit and resubmit.'
            
        ep_check.save()
        
        # Create history entry
        ErrorPreventionCheckHistory.objects.create(
            ep_check=ep_check,
            changed_by=request.user,
            action='supervisor_verified' if action == 'approve' else 'rejected',
            description=f'Supervisor {"approved" if action == "approve" else "rejected"} EP check',
            additional_data={'comments': comments}
        )
        
        messages.success(request, success_message)
        return redirect('dashboard')
    
    # Get all mechanism statuses with related mechanism data, ordered properly
    mechanism_statuses = ErrorPreventionMechanismStatus.objects.filter(
        ep_check=ep_check
    ).select_related('mechanism').order_by(
        'mechanism__display_order',
        'mechanism__mechanism_id'
    )
    
    # Check for incomplete statuses
    incomplete_statuses = ErrorPreventionMechanismStatus.objects.filter(
        ep_check=ep_check, 
        status='',
        is_not_applicable=False
    ).exists()
    
    context = {
        'ep_check': ep_check,
        'mechanism_statuses': mechanism_statuses,
        'incomplete_statuses': incomplete_statuses,
        'is_re_verification': ep_check.status not in ['pending', 'rejected'],
        'title': 'Supervisor Verification - EP Check'
    }
    
    return render(request, 'main/operations/ep_check_verify.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'quality_supervisor')
def ep_check_verify_quality(request, pk):
    """Verify an EP check as a quality supervisor - can verify at any time"""
    ep_check = get_object_or_404(ErrorPreventionCheck, pk=pk)
    
    # Quality supervisors can verify at any time, but warn if not supervisor approved
    show_warning = ep_check.status != 'supervisor_approved'
    
    if request.method == 'POST':
        action = request.POST.get('action', 'approve')
        comments = request.POST.get('comments', '')
        
        # Store previous status for messaging
        previous_status = ep_check.status
        
        # Update check status
        if action == 'approve':
            ep_check.status = 'quality_approved'
            ep_check.quality_supervisor = request.user
            ep_check.comments = comments
            
            # Update verification status
            if ep_check.verification_status:
                verification_status = ep_check.verification_status
                verification_status.status = 'completed'
                verification_status.save()
            
            success_message = 'EP check approved by quality. Process complete.'
            
        else:  # reject action
            ep_check.status = 'rejected'
            # DO NOT set quality_supervisor when rejecting - clear it instead
            ep_check.quality_supervisor = None
            ep_check.comments = comments
            
            # Also clear supervisor approval to force re-verification
            ep_check.supervisor = None
            
            # Update verification status
            if ep_check.verification_status:
                verification_status = ep_check.verification_status
                verification_status.status = 'rejected'
                verification_status.quality_notified = False
                verification_status.save()
            
            success_message = 'EP check rejected by quality. Sent back for review.'
            
        ep_check.save()
        
        # Create history entry
        ErrorPreventionCheckHistory.objects.create(
            ep_check=ep_check,
            changed_by=request.user,
            action='quality_verified' if action == 'approve' else 'rejected',
            description=f'Quality {"approved" if action == "approve" else "rejected"} EP check',
            additional_data={'comments': comments}
        )
        
        messages.success(request, success_message)
        return redirect('dashboard')
    
    # Get all mechanism statuses with related mechanism data, ordered properly
    mechanism_statuses = ErrorPreventionMechanismStatus.objects.filter(
        ep_check=ep_check
    ).select_related('mechanism').order_by(
        'mechanism__display_order',
        'mechanism__mechanism_id'
    )
    
    # Check for any non-working mechanisms
    non_working_mechanisms = ErrorPreventionMechanismStatus.objects.filter(
        ep_check=ep_check,
        is_working=False,
        is_not_applicable=False
    )
    
    context = {
        'ep_check': ep_check,
        'mechanism_statuses': mechanism_statuses,
        'non_working_mechanisms': non_working_mechanisms,
        'has_non_working': non_working_mechanisms.exists(),
        'show_warning': show_warning,
        'is_re_verification': ep_check.status in ['quality_approved'],
        'title': 'Quality Verification - EP Check'
    }
    
    return render(request, 'main/operations/ep_check_verify_quality.html', context)


@login_required
def start_ep_workflow(request):
    """Start a new EP workflow with verification status"""
    if request.method == 'POST':
        form = ErrorPreventionWorkflowForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Create a new shift
                    shift = Shift.objects.create(
                        date=form.cleaned_data['date'],
                        shift_type=form.cleaned_data['shift_type'],
                        operator=form.cleaned_data['operator'],
                        shift_supervisor=form.cleaned_data['shift_supervisor'],
                        quality_supervisor=form.cleaned_data['quality_supervisor']
                    )
                    
                    # Create a verification status
                    verification_status = DailyVerificationStatus.objects.create(
                        date=form.cleaned_data['date'],
                        shift=shift,
                        status='pending',
                        created_by=request.user
                    )
                    
                    # Create the EP check
                    ep_check = ErrorPreventionCheck.objects.create(
                        verification_status=verification_status,
                        date=form.cleaned_data['date'],
                        current_model=form.cleaned_data['current_model'],
                        operator=form.cleaned_data['operator'],
                        supervisor=None,  # Will be set when verified
                        quality_supervisor=None,  # Will be set when verified
                        status='pending',
                        comments=form.cleaned_data['comments']
                    )
                    
                    # Create mechanism statuses
                    for i, mech_id in enumerate(dict(ErrorPreventionCheck.EP_MECHANISM_CHOICES).keys(), 1):
                        ErrorPreventionMechanismStatus.objects.create(
                            ep_check=ep_check,
                            ep_mechanism_id=mech_id,
                            is_working=True,
                            is_not_applicable=False,
                            status='OK',  # Default to OK
                            alternative_method='100% Inspection By Operator',
                            comments=''
                        )
                    
                    messages.success(request, 'Error Prevention workflow created successfully')
                    return redirect('ep_check_edit_statuses', pk=ep_check.pk)
            except Exception as e:
                messages.error(request, f'Error creating EP workflow: {str(e)}')
    else:
        form = ErrorPreventionWorkflowForm(user=request.user)
    
    context = {
        'form': form,
        'title': 'Start Error Prevention Workflow'
    }
    
    return render(request, 'main/operations/ep_workflow_form.html', context)

@login_required
def ep_check_dashboard(request):
    """Dashboard with EP check statistics and charts"""
    # Set today's date for max date values
    today = timezone.now().date()
    
    # Handle custom date range if provided
    custom_start_date = request.GET.get('custom_start_date')
    custom_end_date = request.GET.get('custom_end_date')
    
    if custom_start_date and custom_end_date:
        try:
            start_date = datetime.strptime(custom_start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(custom_end_date, '%Y-%m-%d').date()
            # Calculate days for active button state
            days = (end_date - start_date).days + 1
        except ValueError:
            # Fall back to default if dates are invalid
            days = int(request.GET.get('days', 30))
            end_date = today
            start_date = end_date - timedelta(days=days-1)
    else:
        # Get date range for filtering from days parameter
        days = int(request.GET.get('days', 30))
        end_date = today
        start_date = end_date - timedelta(days=days-1)
    
    # Get EP checks in date range
    ep_checks = ErrorPreventionCheck.objects.filter(
        date__range=[start_date, end_date]
    ).select_related('verification_status__shift', 'operator')
    
    # Calculate summary statistics
    stats = {
        'total_checks': ep_checks.count(),
        'approved_count': ep_checks.filter(status='quality_approved').count(),
        'rejected_count': ep_checks.filter(status='rejected').count(),
        'pending_count': ep_checks.filter(status__in=['pending', 'supervisor_approved']).count(),
    }
    
    # Add derived statistics
    if stats['total_checks'] > 0:
        stats['completion_rate'] = ((stats['approved_count'] + stats['rejected_count']) / stats['total_checks']) * 100
        stats['success_rate'] = (stats['approved_count'] / max(stats['approved_count'] + stats['rejected_count'], 1)) * 100
        stats['rejection_rate'] = (stats['rejected_count'] / max(stats['approved_count'] + stats['rejected_count'], 1)) * 100
        stats['pending_percentage'] = (stats['pending_count'] / stats['total_checks']) * 100
    else:
        stats['completion_rate'] = 0
        stats['success_rate'] = 0
        stats['rejection_rate'] = 0
        stats['pending_percentage'] = 0
    
    # Get mechanism statuses for checks in range
    mechanism_statuses = ErrorPreventionMechanismStatus.objects.filter(
        ep_check__in=ep_checks
    ).select_related('mechanism')
    
    # Get all active mechanisms from master list
    active_mechanisms = ErrorPreventionMechanism.objects.filter(
        is_active=True
    ).order_by('display_order', 'mechanism_id')
    
    # Calculate statistics by mechanism using the master mechanism list
    mechanism_stats = {}
    for mechanism in active_mechanisms:
        # Get all statuses for this mechanism
        statuses = mechanism_statuses.filter(mechanism=mechanism)
        
        total = statuses.count()
        ok_count = statuses.filter(status='OK').count()
        ng_count = statuses.filter(status='NG').count()
        na_count = statuses.filter(is_not_applicable=True).count()
        
        applicable_count = total - na_count
        
        if applicable_count > 0:
            ok_rate = (ok_count / applicable_count) * 100
            ng_rate = (ng_count / applicable_count) * 100
        else:
            ok_rate = 0
            ng_rate = 0
            
        mechanism_stats[mechanism.mechanism_id] = {
            'name': mechanism.description[:50],  # Truncate long descriptions
            'mechanism_id': mechanism.mechanism_id,
            'total': total,
            'ok_count': ok_count,
            'ng_count': ng_count,
            'na_count': na_count,
            'ok_rate': ok_rate,
            'ng_rate': ng_rate
        }
    
    # Prepare chart data for mechanisms
    chart_data = {
        'labels': [m_stats['mechanism_id'] for m_id, m_stats in mechanism_stats.items()],
        'ok_rates': [m_stats['ok_rate'] for m_id, m_stats in mechanism_stats.items()],
        'ok_counts': [m_stats['ok_count'] for m_id, m_stats in mechanism_stats.items()],
        'ng_counts': [m_stats['ng_count'] for m_id, m_stats in mechanism_stats.items()]
    }
    
    # Generate time series data for trend chart
    date_range = []
    current_date = start_date
    while current_date <= end_date:
        date_range.append(current_date)
        current_date += timedelta(days=1)
    
    # Format date range based on total days
    if days <= 14:
        formatted_dates = [d.strftime('%b %d') for d in date_range]
    elif days <= 90:
        # Group by week for medium ranges
        formatted_dates = []
        week_starts = []
        for i, d in enumerate(date_range):
            if i % 7 == 0 or i == len(date_range) - 1:
                formatted_dates.append(d.strftime('%b %d'))
                week_starts.append(d)
            else:
                formatted_dates.append('')
        date_range = week_starts
    else:
        # Group by month for long ranges
        month_grouped_dates = []
        month_labels = []
        current_month = None
        
        for d in date_range:
            if current_month != d.month:
                current_month = d.month
                month_grouped_dates.append(d)
                month_labels.append(d.strftime('%b %Y'))
        
        date_range = month_grouped_dates
        formatted_dates = month_labels
    
    # Get daily counts
    total_counts = []
    approved_counts = []
    rejected_counts = []
    
    for date in date_range:
        if days <= 14:
            day_checks = ep_checks.filter(date=date)
        elif days <= 90:
            week_end = date + timedelta(days=6)
            day_checks = ep_checks.filter(date__range=[date, min(week_end, end_date)])
        else:
            month_end = date.replace(day=28) + timedelta(days=4)
            month_end = month_end.replace(day=1) - timedelta(days=1)
            day_checks = ep_checks.filter(date__range=[date, min(month_end, end_date)])
        
        total_counts.append(day_checks.count())
        approved_counts.append(day_checks.filter(status='quality_approved').count())
        rejected_counts.append(day_checks.filter(status='rejected').count())
    
    # Package trend data for the template
    trends_data = {
        'dates': formatted_dates,
        'total_counts': total_counts,
        'approved_counts': approved_counts,
        'rejected_counts': rejected_counts
    }
    
    context = {
        'stats': stats,
        'mechanism_stats': mechanism_stats,
        'chart_data': json.dumps(chart_data),
        'trends_data': json.dumps(trends_data),
        'days': days,
        'start_date': start_date,
        'end_date': end_date,
        'today': today,
        'custom_start_date': custom_start_date or start_date.strftime('%Y-%m-%d'),
        'custom_end_date': custom_end_date or end_date.strftime('%Y-%m-%d'),
        'title': 'EP Check Dashboard'
    }
    
    return render(request, 'main/operations/ep_check_dashboard.html', context)

@login_required
def export_ep_excel(request, pk):
    """Export an EP check to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    ep_check = get_object_or_404(
        ErrorPreventionCheck.objects.select_related(
            'verification_status__shift', 'operator', 'supervisor', 'quality_supervisor'
        ),
        pk=pk
    )
    
    # Get all mechanism statuses
    mechanism_statuses = ErrorPreventionMechanismStatus.objects.filter(
        ep_check=ep_check
    ).order_by('ep_mechanism_id')
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "EP Check"
    
    # Define styles
    header_style = {
        'fill': PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid'),
        'font': Font(color='FFFFFF', bold=True),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    
    ok_style = {
        'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'font': Font(color='006100'),
        'alignment': Alignment(horizontal='center')
    }
    
    ng_style = {
        'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'font': Font(color='9C0006'),
        'alignment': Alignment(horizontal='center')
    }
    
    na_style = {
        'fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'font': Font(color='9C6500'),
        'alignment': Alignment(horizontal='center')
    }
    
    # Add title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "Error Prevention Check Master List"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Add basic info
    ws['A3'] = "Date:"
    ws['B3'] = ep_check.date
    
    ws['D3'] = "Shift:"
    ws['E3'] = ep_check.verification_status.shift.get_shift_type_display()
    
    ws['A4'] = "Operator:"
    ws['B4'] = ep_check.operator.username
    
    ws['D4'] = "Model:"
    ws['E4'] = ep_check.current_model
    
    ws['A5'] = "Status:"
    ws['B5'] = ep_check.get_status_display()
    
    # Add headers for mechanisms
    headers = ['Mechanism ID', 'Description', 'Status', 'N/A', 'Working', 'Alternative Method', 'Comments']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.fill = header_style['fill']
        cell.font = header_style['font']
        cell.alignment = header_style['alignment']
    
    # Add mechanism statuses
    mechanism_choices = dict(ErrorPreventionCheck.EP_MECHANISM_CHOICES)
    
    for row_idx, status in enumerate(mechanism_statuses, 8):
        # Mechanism ID
        ws.cell(row=row_idx, column=1, value=status.ep_mechanism_id)
        
        # Description
        description = mechanism_choices.get(status.ep_mechanism_id, 'Unknown')
        ws.cell(row=row_idx, column=2, value=description)
        
        # Status with color-coding
        result_cell = ws.cell(row=row_idx, column=3, value=status.status)
        if status.is_not_applicable:
            result_cell.value = 'N/A'
            result_cell.fill = na_style['fill']
            result_cell.font = na_style['font']
            result_cell.alignment = na_style['alignment']
        elif status.status == 'OK':
            result_cell.fill = ok_style['fill']
            result_cell.font = ok_style['font']
            result_cell.alignment = ok_style['alignment']
        elif status.status == 'NG':
            result_cell.fill = ng_style['fill']
            result_cell.font = ng_style['font']
            result_cell.alignment = ng_style['alignment']
        
        # N/A indicator
        ws.cell(row=row_idx, column=4, value='Yes' if status.is_not_applicable else 'No')
        
        # Working
        ws.cell(row=row_idx, column=5, value='Yes' if status.is_working else 'No')
        
        # Alternative Method
        ws.cell(row=row_idx, column=6, value=status.alternative_method)
        
        # Comments
        ws.cell(row=row_idx, column=7, value=status.comments)
    
    # Add verification section
    row_idx = len(mechanism_statuses) + 10  # Leave a gap after mechanisms
    
    ws.merge_cells(f'A{row_idx}:G{row_idx}')
    verification_header = ws[f'A{row_idx}']
    verification_header.value = "Verification Details"
    verification_header.font = Font(size=12, bold=True)
    
    row_idx += 2
    ws.cell(row=row_idx, column=1, value="Operator:")
    ws.cell(row=row_idx, column=2, value=ep_check.operator.username)
    
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Supervisor:")
    ws.cell(row=row_idx, column=2, value=ep_check.supervisor.username if ep_check.supervisor else 'Not verified')
    
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Quality Supervisor:")
    ws.cell(row=row_idx, column=2, value=ep_check.quality_supervisor.username if ep_check.quality_supervisor else 'Not verified')
    
    row_idx += 2
    ws.cell(row=row_idx, column=1, value="Comments:")
    ws.cell(row=row_idx, column=2, value=ep_check.comments)
    ws.merge_cells(f'B{row_idx}:G{row_idx}')
    
    # Add verification status info
    row_idx += 2
    ws.cell(row=row_idx, column=1, value="Verification Status:")
    ws.cell(row=row_idx, column=2, value=ep_check.verification_status.get_status_display())
    
    # Adjust column widths
    col_widths = [15, 40, 10, 10, 10, 25, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ep_check_{ep_check.date}.xlsx'
    
    wb.save(response)
    return response




@login_required
def ep_mechanism_status_update(request, status_id):
    """API endpoint to update a mechanism status via AJAX"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method is allowed'}, status=405)
    
    try:
        mechanism_status = get_object_or_404(ErrorPreventionMechanismStatus, pk=status_id)
        
        # Check if parent check is still editable
        if mechanism_status.ep_check.status != 'pending':
            return JsonResponse({'error': 'Cannot update status of a verified or rejected check'}, status=403)
        
        # Get data from request
        status = request.POST.get('status')
        is_not_applicable = request.POST.get('is_not_applicable') == 'true'
        comments = request.POST.get('comments', '')
        
        # Update the status
        mechanism_status.status = status
        mechanism_status.is_not_applicable = is_not_applicable
        mechanism_status.comments = comments
        mechanism_status.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Status updated successfully'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def ep_check_delete(request, pk):
    """Delete an Error Prevention check"""
    ep_check = get_object_or_404(ErrorPreventionCheck, pk=pk)
    
    # Only supervisors and admins can delete
    if not (request.user.user_type in ['shift_supervisor', 'quality_supervisor'] or request.user.is_superuser):
        messages.error(request, 'You do not have permission to delete this check.')
        return redirect('ep_check_detail', pk=ep_check.pk)
    
    if request.method == 'POST':
        try:
            # Get verification status to update
            verification_status = ep_check.verification_status
            
            # Delete all associated mechanism statuses first
            ErrorPreventionMechanismStatus.objects.filter(ep_check=ep_check).delete()
            
            # Then delete the checklist
            ep_check.delete()
            
            messages.success(request, 'Error Prevention check has been deleted successfully.')
            return redirect('ep_check_list')
        except Exception as e:
            messages.error(request, f'Error deleting check: {str(e)}')
            return redirect('ep_check_detail', pk=ep_check.pk)
    
    return render(request, 'main/operations/ep_check_delete.html', {
        'ep_check': ep_check,
        'title': 'Delete EP Check'
    })

@login_required
def operator_dashboard_ep(request):
    """Add EP check information to operator dashboard"""
    # Get the standard operator dashboard context
    context = operator_dashboard(request)
    
    # Get the active verification status
    active_verification = context.get('active_verification')
    
    # Check for EP checks if verification status exists
    if active_verification:
        # Look for an EP check for this verification
        ep_check = ErrorPreventionCheck.objects.filter(
            verification_status=active_verification
        ).first()
        
        # Add to context
        context['ep_check'] = ep_check
        context['can_create_ep'] = bool(active_verification) and not ep_check
    
    return render(request, 'main/operator_dashboard.html', context)

@login_required
def ep_check_from_verification(request, verification_id):
    """Create an EP check from an existing verification status"""
    verification_status = get_object_or_404(DailyVerificationStatus, id=verification_id)
    
    # Check if EP check already exists for this verification
    existing_check = ErrorPreventionCheck.objects.filter(verification_status=verification_status).first()
    if existing_check:
        messages.warning(request, 'An EP check already exists for this verification status.')
        return redirect('ep_check_detail', pk=existing_check.pk)
    
    # Check if user has permission
    if verification_status.created_by != request.user and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to create an EP check for this verification.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        # Get the model from form submission
        form = ErrorPreventionCheckForm(
            request.POST, 
            verification_status=verification_status,
            user=request.user
        )
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Create the EP check
                    ep_check = form.save()
                    
                    # Create default mechanism statuses
                    for mech_id, _ in ErrorPreventionCheck.EP_MECHANISM_CHOICES:
                        ErrorPreventionMechanismStatus.objects.create(
                            ep_check=ep_check,
                            ep_mechanism_id=mech_id,
                            is_working=True,
                            is_not_applicable=False,
                            status='OK',  # Default value
                            alternative_method='100% Inspection By Operator'
                        )
                    
                    messages.success(request, 'Error Prevention check created successfully.')
                    return redirect('ep_check_detail', pk=ep_check.pk)
            except Exception as e:
                messages.error(request, f'Error creating EP check: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # Pre-fill with verification status information
        form = ErrorPreventionCheckForm(
            initial={
                'date': verification_status.date,
                'current_model': 'T6'  # Default value
            },
            verification_status=verification_status,
            user=request.user
        )
    
    context = {
        'form': form,
        'verification_status': verification_status,
        'title': 'Create Error Prevention Check'
    }
    
    return render(request, 'main/operations/ep_check_from_verification.html', context)    
    
    
    
    
    
    
    
    #  new code
    
    
# New DTPM Views
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q, Sum, Exists, OuterRef
from django.db import transaction
from datetime import datetime, timedelta
import json

from .models import DTPMChecklistFMA03New, DTPMCheckResultNew, DTPMIssueNew, Shift, User, DailyVerificationStatus
from .forms import (
    DTPMChecklistFMA03NewForm, DTPMCheckResultNewForm, 
    DTPMCheckResultNewInlineFormSet, DTPMIssueNewForm, 
    DTPMIssueResolveNewForm, DTPMChecklistNewFilterForm
)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from .models import (
    DTPMCheckpoint, DTPMChecklistFMA03New, DTPMCheckResultNew, 
    DailyVerificationStatus, ChecklistBase
)
from .forms import DTPMChecklistFMA03NewForm



@login_required
def dtpm_new_list(request):
    """List all DTPM checklists with filtering options"""
    form = DTPMChecklistNewFilterForm(request.GET)
    
    # Annotate the queryset with issue count
    checklists = DTPMChecklistFMA03New.objects.select_related('shift', 'operator', 'supervisor', 'verification_status')\
        .annotate(issue_count=Count('check_results__issues', distinct=True))
    
    # Apply filters if form is valid
    if form.is_valid():
        data = form.cleaned_data
        if data.get('date_from'):
            checklists = checklists.filter(date__gte=data['date_from'])
        if data.get('date_to'):
            checklists = checklists.filter(date__lte=data['date_to'])
        if data.get('status'):
            checklists = checklists.filter(status=data['status'])
        if data.get('has_issues'):
            if data['has_issues'] == 'yes':
                checklists = checklists.filter(issue_count__gt=0)
            elif data['has_issues'] == 'no':
                checklists = checklists.filter(issue_count=0)
    
    # Order by date, newest first
    checklists = checklists.order_by('-date')
    
    # Calculate completion statistics
    stats = {
        'total': checklists.count(),
        'completed': checklists.filter(status__in=['verified', 'rejected']).count(),
        'pending': checklists.filter(status='pending').count(),
        'issues_count': DTPMIssueNew.objects.filter(
            check_result__checklist__in=checklists
        ).count()
    }
    
    # Calculate completion percentage
    if stats['total'] > 0:
        stats['completion_percentage'] = (stats['completed'] / stats['total']) * 100
    else:
        stats['completion_percentage'] = 0
    
    context = {
        'checklists': checklists,
        'filter_form': form,
        'stats': stats,
        'title': 'DTPM FMA03 Checklists'
    }
    
    return render(request, 'main/dtpm/dtpm_new_list.html', context)

@login_required
def dtpm_new_detail(request, pk):
    """View details of a specific DTPM checklist with dynamic checkpoints"""
    checklist = get_object_or_404(
        DTPMChecklistFMA03New.objects.select_related(
            'shift', 'operator', 'supervisor', 'verification_status'
        ),
        pk=pk
    )
    
    # Get all checkpoint results with their checkpoint details
    # Use the Meta ordering from DTPMCheckResultNew model
    checkpoint_results = DTPMCheckResultNew.objects.filter(
        checklist=checklist
    ).select_related('checkpoint')
    
    # Get issues for this checklist
    issues = DTPMIssueNew.objects.filter(
        check_result__checklist=checklist
    ).select_related('check_result', 'reported_by', 'assigned_to').order_by('-created_at')
    
    # Permissions
    can_edit = (
        request.user == checklist.operator and 
        checklist.status == 'pending'
    )
    
    can_verify_supervisor = (
        request.user.user_type == 'shift_supervisor' and
        checklist.status == 'pending'
    )
    
    can_verify_quality = (
        request.user.user_type == 'quality_supervisor' and
        checklist.status == 'supervisor_approved'
    )
    
    # Check for modifications (timestamp difference > 30 seconds)
    from datetime import timedelta
    has_modifications = False
    modifications = []
    
    for result in checkpoint_results:
        time_diff = abs((result.updated_at - result.checked_at).total_seconds())
        is_really_modified = time_diff > 30
        result.is_really_modified = is_really_modified
        
        if is_really_modified:
            has_modifications = True
            modifications.append({
                'checkpoint_number': result.checkpoint.checkpoint_number,
                'checkpoint_title': result.checkpoint.title_english,
                'current_status': result.status,
                'has_comments': bool(result.comments),
                'last_modified': result.updated_at,
                'originally_checked': result.checked_at,
                'time_diff_seconds': time_diff,
            })
    
    # Get machine overview image
    machine_overview = DTPMCheckpoint.objects.filter(
        checkpoint_number=8,
        is_active=True
    ).first()
    
    # Calculate statistics
    total_checkpoints = checkpoint_results.count()
    ok_count = checkpoint_results.filter(status='OK').count()
    ng_count = checkpoint_results.filter(status='NG').count()
    
    context = {
        'checklist': checklist,
        'checkpoint_results': checkpoint_results,
        'issues': issues,
        'can_edit': can_edit,
        'can_verify_supervisor': can_verify_supervisor,
        'can_verify_quality': can_verify_quality,
        'has_modifications': has_modifications,
        'modifications': modifications,
        'machine_overview_image': machine_overview,
        'total_checkpoints': total_checkpoints,
        'ok_count': ok_count,
        'ng_count': ng_count,
        'title': f'DTPM Checklist - {checklist.date}'
    }
    
    return render(request, 'main/dtpm/dtpm_new_detail.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'operator')
def dtpm_checklist_create(request):
    """Create a new DTPM checklist with dynamic checkpoints"""
    
    # Get the active verification status
    active_verification = DailyVerificationStatus.objects.filter(
        created_by=request.user,
        date=timezone.now().date(),
        status__in=['pending', 'in_progress']
    ).first()
    
    if not active_verification:
        messages.warning(request, 'You need to create a Daily Verification Sheet first.')
        return redirect('create_checklist')
    
    # Check if checklist exists (for auto-population)
    active_checklist = ChecklistBase.objects.filter(
        verification_status=active_verification
    ).first()
    
    if not active_checklist:
        messages.error(request, 'No checklist found. Please create a checklist first.')
        return redirect('create_checklist')
    
    # Check if DTPM checklist already exists
    existing_checklist = DTPMChecklistFMA03New.objects.filter(
        verification_status=active_verification
    ).first()
    
    if existing_checklist:
        messages.warning(request, 'A DTPM checklist already exists for this verification status.')
        return redirect('dtpm_new_detail', pk=existing_checklist.pk)
    
    # Get all active checkpoints ordered by display order
    active_checkpoints = DTPMCheckpoint.objects.filter(
        is_active=True
    ).exclude(
        checkpoint_number=8  # Exclude machine overview image from main checkpoints
    ).order_by('order', 'checkpoint_number')
    
    # Get machine overview image (checkpoint #8)
    machine_overview = DTPMCheckpoint.objects.filter(
        checkpoint_number=8,
        is_active=True
    ).first()
    
    if request.method == 'POST':
        form = DTPMChecklistFMA03NewForm(
            request.POST, 
            verification_status=active_verification, 
            user=request.user
        )
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Save the checklist
                    checklist = form.save(commit=False)
                    checklist.shift = active_verification.shift
                    checklist.operator = request.user
                    checklist.verification_status = active_verification
                    
                    # Set supervisor from verification shift
                    if active_verification.shift.shift_supervisor:
                        checklist.supervisor = active_verification.shift.shift_supervisor
                    
                    # Save checklist (this triggers signal to auto-create checkpoint results)
                    checklist.save()
                    
                    # UPDATE the auto-created checkpoint results with form data (don't create new ones)
                    for checkpoint in active_checkpoints:
                        status = request.POST.get(f'status_{checkpoint.id}', 'NG')
                        comments = request.POST.get(f'comments_{checkpoint.id}', '')
                        
                        # Update existing checkpoint result (created by signal)
                        DTPMCheckResultNew.objects.filter(
                            checklist=checklist,
                            checkpoint=checkpoint
                        ).update(
                            status=status,
                            comments=comments
                        )
                    
                    messages.success(
                        request, 
                        f'DTPM checklist created successfully for {checklist.current_model} model.'
                    )
                    return redirect('operator_dashboard')
                    
            except Exception as e:
                messages.error(request, f"Error creating DTPM checklist: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = DTPMChecklistFMA03NewForm(
            verification_status=active_verification, 
            user=request.user
        )
    
    context = {
        'form': form,
        'title': 'Create DTPM Checklist',
        'active_verification': active_verification,
        'active_checklist': active_checklist,
        'current_shift': active_verification.shift.get_shift_type_display(),
        'current_model': active_checklist.selected_model,
        'current_shift_from_checklist': dict(ChecklistBase.SHIFTS).get(
            active_checklist.shift, 
            active_checklist.shift
        ),
        'active_checkpoints': active_checkpoints,
        'machine_overview_image': machine_overview,
    }
    
    return render(request, 'main/dtpm/dtpm_new_form.html', context)
   

# Optional: Helper function for DTPM dashboard statistics
def get_dtpm_dashboard_stats(verification_status):
    """Get DTPM statistics for dashboard display"""
    if not verification_status:
        return None
    
    dtpm_checklist = DTPMChecklistFMA03New.objects.filter(
        verification_status=verification_status
    ).first()
    
    if not dtpm_checklist:
        return {
            'exists': False,
            'can_create': True
        }
    
    # Calculate checkpoint statistics
    total_checkpoints = dtpm_checklist.check_results.count()
    ok_count = dtpm_checklist.check_results.filter(status='OK').count()
    ng_count = dtpm_checklist.check_results.filter(status='NG').count()
    
    return {
        'exists': True,
        'checklist': dtpm_checklist,
        'total_checkpoints': total_checkpoints,
        'ok_count': ok_count,
        'ng_count': ng_count,
        'completion_percentage': (ok_count / total_checkpoints * 100) if total_checkpoints > 0 else 0,
        'current_model': dtpm_checklist.current_model,
        'shift': dtmp_checklist.checklist_shift,
        'can_create': False
    }


@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'] or u.is_superuser)
def dtpm_new_delete(request, pk):
    """Delete a DTPM checklist"""
    checklist = get_object_or_404(DTPMChecklistFMA03New, pk=pk)
    
    if request.method == 'POST':
        checklist.delete()
        messages.success(request, 'DTPM checklist deleted successfully.')
        return redirect('dtpm_new_list')
    
    context = {
        'checklist': checklist,
        'title': 'Delete DTPM Checklist'
    }
    
    return render(request, 'main/dtpm/dtpm_new_confirm_delete.html', context)



@login_required
def dtpm_new_edit_checks(request, pk):
    """Edit all checkpoint results for a DTPM checklist"""
    checklist = get_object_or_404(DTPMChecklistFMA03New, pk=pk)
    
    # Check permissions
    if request.user != checklist.operator and not request.user.is_superuser:
        if request.user != checklist.supervisor:
            messages.error(request, 'You do not have permission to edit this checklist.')
            return redirect('dtpm_new_detail', pk=checklist.pk)
    
    # Check if checklist can be edited
    if checklist.status not in ['pending', 'rejected']:
        messages.error(request, 'This checklist cannot be edited. Only pending or rejected checklists can be modified.')
        return redirect('dtpm_new_detail', pk=checklist.pk)
    
    # Get all active checkpoints (excluding machine overview)
    active_checkpoints = DTPMCheckpoint.objects.filter(
        is_active=True
    ).exclude(
        checkpoint_number=8
    ).order_by('order', 'checkpoint_number')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                updated_count = 0
                
                # Process each checkpoint result
                for checkpoint in active_checkpoints:
                    status = request.POST.get(f'status_{checkpoint.id}', '')
                    comments = request.POST.get(f'comments_{checkpoint.id}', '')
                    
                    # Get or create the checkpoint result
                    check_result, created = DTPMCheckResultNew.objects.get_or_create(
                        checklist=checklist,
                        checkpoint=checkpoint,
                        defaults={'status': 'NG'}
                    )
                    
                    # Check if there are actual changes
                    status_changed = check_result.status != status
                    comments_changed = check_result.comments != comments
                    
                    if status_changed or comments_changed:
                        check_result.status = status
                        check_result.comments = comments
                        
                        # Only update timestamp if this is a real edit (not initial creation)
                        from django.utils import timezone
                        time_since_creation = (timezone.now() - check_result.checked_at).total_seconds()
                        if time_since_creation > 30:  # More than 30 seconds since creation
                            check_result.updated_at = timezone.now()
                        
                        check_result.save()
                        updated_count += 1
                
                if updated_count > 0:
                    # If this was a rejected checklist, reset status to pending
                    if checklist.status == 'rejected':
                        checklist.status = 'pending'
                        checklist.rejected_at = None
                        checklist.rejected_by = None
                        checklist.save()
                        messages.success(request, f'Successfully updated {updated_count} checkpoint(s). Status reset to pending for re-verification.')
                    else:
                        messages.success(request, f'Successfully updated {updated_count} checkpoint(s).')
                else:
                    messages.info(request, 'No changes were made to the checklist.')
                    
                return redirect('dtpm_new_detail', pk=checklist.pk)
                
        except Exception as e:
            messages.error(request, f"Error updating checkpoint results: {str(e)}")
    
    # Get checkpoint results - use the Meta ordering from the model
    checkpoint_results = DTPMCheckResultNew.objects.filter(
        checklist=checklist
    ).select_related('checkpoint')
    
    # Get machine overview image
    machine_overview = DTPMCheckpoint.objects.filter(
        checkpoint_number=8,
        is_active=True
    ).first()
    
    context = {
        'checklist': checklist,
        'checkpoint_results': checkpoint_results,
        'active_checkpoints': active_checkpoints,
        'machine_overview_image': machine_overview,
        'title': 'Edit Checkpoint Results'
    }
    
    return render(request, 'main/dtpm/dtpm_new_check_results_form.html', context)
@login_required


@login_required
@user_passes_test(lambda u: u.user_type == 'shift_supervisor')
def dtpm_new_supervisor_verify(request, pk):
    """Supervisor verification of a DTPM checklist - can verify/re-verify at any time"""
    checklist = get_object_or_404(DTPMChecklistFMA03New, pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Check if all required checkpoint results have been filled
        incomplete_checks = DTPMCheckResultNew.objects.filter(
            checklist=checklist, 
            status=''
        ).exists()
        
        if incomplete_checks and action == 'supervisor_approve':
            messages.error(request, 'Cannot approve checklist with incomplete checkpoint results.')
            return redirect('dtpm_new_edit_checks', pk=checklist.pk)
        
        # Update checklist status based on supervisor action
        if action == 'supervisor_approve':
            previous_status = checklist.status
            
            checklist.status = 'supervisor_approved'
            checklist.supervisor_approved_at = timezone.now()
            checklist.supervisor_approved_by = request.user
            checklist.supervisor = request.user
            
            # Clear any previous rejection info
            checklist.rejected_at = None
            checklist.rejected_by = None
            
            # Clear quality verification if re-approving
            if previous_status in ['quality_certified', 'quality_rejected']:
                checklist.quality_certified_at = None
                checklist.quality_certified_by = None
                checklist.quality_rejected_at = None
                checklist.quality_rejected_by = None
                checklist.quality_comments = ''
                success_message = 'Checklist re-approved by supervisor. Previous quality verification cleared - requires fresh quality review.'
            else:
                success_message = 'Checklist approved by supervisor. Sent to Quality for verification.'
            
            # Create verification history record
            DTPMVerificationHistory.objects.create(
                checklist=checklist,
                verification_type='supervisor_approve',
                verified_by=request.user,
                comments=f"Re-approved (previous status: {previous_status})" if previous_status != 'pending' else ""
            )
            
        elif action == 'supervisor_reject':
            checklist.status = 'rejected'
            checklist.rejected_at = timezone.now()
            checklist.rejected_by = request.user
            
            # Clear approvals when rejecting
            checklist.supervisor_approved_at = None
            checklist.supervisor_approved_by = None
            checklist.quality_certified_at = None
            checklist.quality_certified_by = None
            checklist.quality_rejected_at = None
            checklist.quality_rejected_by = None
            checklist.quality_comments = ''
            
            success_message = 'Checklist rejected by supervisor. Operator can edit and resubmit.'
            
            # Create verification history record
            DTPMVerificationHistory.objects.create(
                checklist=checklist,
                verification_type='supervisor_reject',
                verified_by=request.user
            )
            
        else:
            messages.error(request, 'Invalid action specified.')
            return redirect('dtpm_new_supervisor_verify', pk=checklist.pk)
            
        checklist.save()
        messages.success(request, success_message)
        return redirect('dashboard')
    
    # Get all checkpoint results - use Meta ordering
    checkpoint_results = DTPMCheckResultNew.objects.filter(
        checklist=checklist
    ).select_related('checkpoint')
    
    # Check for incomplete checkpoints
    incomplete_checks = DTPMCheckResultNew.objects.filter(
        checklist=checklist, 
        status=''
    ).exists()
    
    # Get verification history
    verification_history = DTPMVerificationHistory.objects.filter(
        checklist=checklist
    ).order_by('-verified_at')[:5]
    
    context = {
        'checklist': checklist,
        'checkpoint_results': checkpoint_results,
        'incomplete_checks': incomplete_checks,
        'verification_history': verification_history,
        'is_re_verification': checklist.status not in ['pending', 'rejected'],
        'title': 'Supervisor Verify DTPM Checklist'
    }
    
    return render(request, 'main/dtpm/dtpm_new_supervisor_verify.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'quality_supervisor')
def dtpm_new_quality_verify(request, pk):
    """Quality verification of a DTPM checklist - can verify at any time"""
    checklist = get_object_or_404(DTPMChecklistFMA03New, pk=pk)
    
    show_warning = checklist.status != 'supervisor_approved'
    
    if request.method == 'POST':
        action = request.POST.get('action')
        quality_comments = request.POST.get('quality_comments', '')
        
        # Check for any critical issues (NG status checkpoints)
        critical_issues = DTPMCheckResultNew.objects.filter(
            checklist=checklist, 
            status='NG'
        ).exists()
        
        if action == 'quality_certify':
            checklist.status = 'quality_certified'
            checklist.quality_certified_at = timezone.now()
            checklist.quality_certified_by = request.user
            checklist.quality_comments = quality_comments
            
            # Clear any previous quality rejection
            checklist.quality_rejected_at = None
            checklist.quality_rejected_by = None
            
            success_message = 'Checklist certified by Quality. Process complete.'
            
            # Create verification history record
            DTPMVerificationHistory.objects.create(
                checklist=checklist,
                verification_type='quality_certify',
                verified_by=request.user,
                comments=quality_comments
            )
            
        elif action == 'quality_reject':
            checklist.status = 'quality_rejected'
            checklist.quality_rejected_at = timezone.now()
            checklist.quality_rejected_by = request.user
            checklist.quality_comments = quality_comments
            
            # Clear any previous quality certification
            checklist.quality_certified_at = None
            checklist.quality_certified_by = None
            
            success_message = 'Checklist rejected by Quality. Sent back for review.'
            
            # Create verification history record
            DTPMVerificationHistory.objects.create(
                checklist=checklist,
                verification_type='quality_reject',
                verified_by=request.user,
                comments=quality_comments
            )
            
        else:
            messages.error(request, 'Invalid action specified.')
            return redirect('dtpm_new_quality_verify', pk=checklist.pk)
            
        checklist.save()
        messages.success(request, success_message)
        return redirect('dashboard')
    
    # Get all checkpoint results - use Meta ordering
    checkpoint_results = DTPMCheckResultNew.objects.filter(
        checklist=checklist
    ).select_related('checkpoint')
    
    # Check for quality issues (NG status checkpoints)
    ng_checkpoints = DTPMCheckResultNew.objects.filter(
        checklist=checklist, 
        status='NG'
    )
    
    # Get verification history
    verification_history = DTPMVerificationHistory.objects.filter(
        checklist=checklist
    ).order_by('-verified_at')[:5]
    
    context = {
        'checklist': checklist,
        'checkpoint_results': checkpoint_results,
        'ng_checkpoints': ng_checkpoints,
        'has_ng_checkpoints': ng_checkpoints.exists(),
        'show_warning': show_warning,
        'verification_history': verification_history,
        'is_re_verification': checklist.status in ['quality_certified', 'quality_rejected'],
        'title': 'Quality Verify DTPM Checklist'
    }
    
    return render(request, 'main/dtpm/dtpm_new_quality_verify.html', context)

# Updated routing view with flexible permissions
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'])
def dtpm_new_verify(request, pk):
    """Route to appropriate verification based on user type - flexible permissions"""
    checklist = get_object_or_404(DTPMChecklistFMA03New, pk=pk)
    
    # Route based on user type - no status restrictions (supervisors can verify anytime)
    if request.user.user_type == 'shift_supervisor':
        return redirect('dtpm_new_supervisor_verify', pk=pk)
        
    elif request.user.user_type == 'quality_supervisor':
        return redirect('dtpm_new_quality_verify', pk=pk)
    
    # Fallback (shouldn't reach here due to decorator)
    messages.error(request, 'You do not have permission to verify this checklist.')
    return redirect('dtpm_new_detail', pk=pk)


# Optional: Add a view for supervisors to see all verification actions they can take
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'])
def dtpm_verification_dashboard(request):
    """Dashboard showing checklists that need verification or can be re-verified"""
    
    if request.user.user_type == 'shift_supervisor':
        # Show checklists that need supervisor verification or can be re-verified
        pending_checklists = DTPMChecklistFMA03New.objects.filter(
            status='pending'
        ).select_related('operator', 'shift').order_by('-created_at')
        
        rejected_checklists = DTPMChecklistFMA03New.objects.filter(
            status='rejected',
            rejected_by=request.user
        ).select_related('operator', 'shift').order_by('-rejected_at')
        
        approved_checklists = DTPMChecklistFMA03New.objects.filter(
            supervisor_approved_by=request.user
        ).select_related('operator', 'shift').order_by('-supervisor_approved_at')[:10]
        
        context = {
            'pending_checklists': pending_checklists,
            'rejected_checklists': rejected_checklists,
            'approved_checklists': approved_checklists,
            'user_type': 'supervisor',
            'title': 'Supervisor Verification Dashboard'
        }
        
    elif request.user.user_type == 'quality_supervisor':
        # Show checklists that need quality verification
        supervisor_approved = DTPMChecklistFMA03New.objects.filter(
            status='supervisor_approved'
        ).select_related('operator', 'supervisor', 'shift').order_by('-supervisor_approved_at')
        
        quality_certified = DTPMChecklistFMA03New.objects.filter(
            quality_certified_by=request.user
        ).select_related('operator', 'supervisor', 'shift').order_by('-quality_certified_at')[:10]
        
        quality_rejected = DTPMChecklistFMA03New.objects.filter(
            quality_rejected_by=request.user
        ).select_related('operator', 'supervisor', 'shift').order_by('-quality_rejected_at')[:10]
        
        context = {
            'supervisor_approved': supervisor_approved,
            'quality_certified': quality_certified,
            'quality_rejected': quality_rejected,
            'user_type': 'quality',
            'title': 'Quality Verification Dashboard'
        }
    
    return render(request, 'main/dtpm/dtpm_verification_dashboard.html', context)

@login_required
def dtpm_new_report_issue(request, check_id):
    """Report an issue for a specific checkpoint result"""
    check_result = get_object_or_404(
        DTPMCheckResultNew.objects.select_related('checklist'), 
        pk=check_id
    )
    
    if request.method == 'POST':
        form = DTPMIssueNewForm(request.POST)
        if form.is_valid():
            issue = form.save(commit=False)
            issue.check_result = check_result
            issue.reported_by = request.user
            issue.save()
            
            messages.success(request, 'Issue reported successfully.')
            return redirect('dtpm_new_detail', pk=check_result.checklist.pk)
    else:
        form = DTPMIssueNewForm()
    
    context = {
        'form': form,
        'check_result': check_result,
        'title': 'Report Issue'
    }
    
    return render(request, 'main/dtpm/dtpm_new_issue_form.html', context)
@login_required
@user_passes_test(lambda u: u.user_type in ['shift_supervisor', 'quality_supervisor'])
def dtpm_new_resolve_issue(request, issue_id):
    """Resolve a reported issue"""
    issue = get_object_or_404(
        DTPMIssueNew.objects.select_related('check_result__checklist'), 
        pk=issue_id
    )
    
    if request.method == 'POST':
        form = DTPMIssueResolveNewForm(request.POST, instance=issue)
        if form.is_valid():
            resolved_issue = form.save(commit=False)
            resolved_issue.resolved_by = request.user
            resolved_issue.save()
            
            messages.success(request, 'Issue updated successfully.')
            return redirect('dtpm_new_detail', pk=issue.check_result.checklist.pk)
    else:
        form = DTPMIssueResolveNewForm(instance=issue)
    
    context = {
        'form': form,
        'issue': issue,
        'title': 'Resolve Issue'
    }
    
    return render(request, 'main/dtpm/dtpm_new_issue_resolve_form.html', context)

@login_required
def dtpm_new_dashboard(request):
    """Dashboard with DTPM statistics and charts"""
    # Get date range for filtering
    custom_date = False
    
    # Handle custom date range
    if 'start_date' in request.GET and 'end_date' in request.GET:
        try:
            start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
            custom_date = True
        except ValueError:
            days = int(request.GET.get('date_filter', 30))
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=days)
    else:
        days = int(request.GET.get('date_filter', 30))
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days)
    
    # Get previous period for trend comparison
    prev_period_length = (end_date - start_date).days
    prev_end_date = start_date - timedelta(days=1)
    prev_start_date = prev_end_date - timedelta(days=prev_period_length)
    
    # Get checklists in date range
    checklists = DTPMChecklistFMA03New.objects.filter(
        date__range=[start_date, end_date]
    ).select_related('shift', 'operator', 'verification_status')
    
    # Get checklists from previous period for trend calculation
    prev_checklists = DTPMChecklistFMA03New.objects.filter(
        date__range=[prev_start_date, prev_end_date]
    )
    
    # Calculate summary statistics
    stats = {
        'total_checklists': checklists.count(),
        'verified_count': checklists.filter(status='quality_certified').count(),
        'rejected_count': checklists.filter(status__in=['rejected', 'quality_rejected']).count(),
        'pending_count': checklists.filter(status__in=['pending', 'supervisor_approved']).count(),
    }
    
    # Calculate previous period statistics
    prev_stats = {
        'total_checklists': prev_checklists.count(),
        'verified_count': prev_checklists.filter(status='quality_certified').count(),
        'rejected_count': prev_checklists.filter(status__in=['rejected', 'quality_rejected']).count(),
        'pending_count': prev_checklists.filter(status__in=['pending', 'supervisor_approved']).count(),
    }
    
    if stats['total_checklists'] > 0:
        stats['completion_rate'] = ((stats['verified_count'] + stats['rejected_count']) / stats['total_checklists']) * 100
        stats['success_rate'] = (stats['verified_count'] / (stats['verified_count'] + stats['rejected_count'])) * 100 if (stats['verified_count'] + stats['rejected_count']) > 0 else 0
    else:
        stats['completion_rate'] = 0
        stats['success_rate'] = 0
        
    if prev_stats['total_checklists'] > 0:
        prev_stats['completion_rate'] = ((prev_stats['verified_count'] + prev_stats['rejected_count']) / prev_stats['total_checklists']) * 100
        prev_stats['success_rate'] = (prev_stats['verified_count'] / (prev_stats['verified_count'] + prev_stats['rejected_count'])) * 100 if (prev_stats['verified_count'] + prev_stats['rejected_count']) > 0 else 0
    else:
        prev_stats['completion_rate'] = 0
        prev_stats['success_rate'] = 0
    
    # Calculate trend percentages
    if prev_stats['total_checklists'] > 0:
        stats['total_trend'] = ((stats['total_checklists'] - prev_stats['total_checklists']) / prev_stats['total_checklists']) * 100
    else:
        stats['total_trend'] = 0
        
    stats['success_rate_trend'] = stats['success_rate'] - prev_stats['success_rate'] if prev_stats['success_rate'] > 0 else 0
    stats['completion_rate_trend'] = stats['completion_rate'] - prev_stats['completion_rate'] if prev_stats['completion_rate'] > 0 else 0
    
    if prev_stats['rejected_count'] > 0:
        stats['rejected_trend'] = ((stats['rejected_count'] - prev_stats['rejected_count']) / prev_stats['rejected_count']) * 100
    else:
        stats['rejected_trend'] = 0
    
    # Get all active checkpoints (excluding machine overview)
    active_checkpoints = DTPMCheckpoint.objects.filter(
        is_active=True
    ).exclude(checkpoint_number=8).order_by('order', 'checkpoint_number')
    
    # Calculate statistics by checkpoint
    checkpoint_stats = {}
    prev_checkpoint_stats = {}
    
    # First collect previous period stats
    for checkpoint in active_checkpoints:
        prev_results = DTPMCheckResultNew.objects.filter(
            checklist__in=prev_checklists,
            checkpoint=checkpoint
        )
        
        prev_total = prev_results.count()
        prev_ok_count = prev_results.filter(status='OK').count()
        
        prev_ok_rate = (prev_ok_count / prev_total) * 100 if prev_total > 0 else 0
            
        prev_checkpoint_stats[checkpoint.checkpoint_number] = {
            'ok_rate': prev_ok_rate
        }
    
    # Now collect current period stats with comparison
    for checkpoint in active_checkpoints:
        results = DTPMCheckResultNew.objects.filter(
            checklist__in=checklists,
            checkpoint=checkpoint
        )
        
        total = results.count()
        ok_count = results.filter(status='OK').count()
        ng_count = results.filter(status='NG').count()
        
        ok_rate = (ok_count / total) * 100 if total > 0 else 0
        
        # Calculate change from previous period    
        prev_ok_rate = prev_checkpoint_stats.get(checkpoint.checkpoint_number, {}).get('ok_rate', 0)
        change = ok_rate - prev_ok_rate
            
        checkpoint_stats[checkpoint.checkpoint_number] = {
            'description': checkpoint.title_english,
            'total': total,
            'ok_count': ok_count,
            'ng_count': ng_count,
            'ok_rate': ok_rate,
            'change': change
        }
    
    # Get open issues
    open_issues = DTPMIssueNew.objects.filter(
        check_result__checklist__in=checklists,
        status__in=['open', 'in_progress']
    ).select_related('check_result__checkpoint', 'reported_by').order_by('-created_at')
    
    # Prepare bar chart data
    chart_data = {
        'labels': [f"CP{cp.checkpoint_number}" for cp in active_checkpoints],
        'ok_rates': [checkpoint_stats.get(cp.checkpoint_number, {'ok_rate': 0})['ok_rate'] for cp in active_checkpoints],
        'descriptions': [
            (checkpoint_stats.get(cp.checkpoint_number, {'description': ''})['description'][:30] + '...' 
            if len(checkpoint_stats.get(cp.checkpoint_number, {'description': ''})['description']) > 30 
            else checkpoint_stats.get(cp.checkpoint_number, {'description': ''})['description'])
            for cp in active_checkpoints
        ]
    }
    
    # Prepare trend line chart data (daily data within selected range)
    date_range = []
    current_date = start_date
    
    while current_date <= end_date:
        date_range.append(current_date)
        current_date += timedelta(days=1)
    
    success_rates = []
    completion_rates = []
    checklist_counts = []
    
    # Calculate daily statistics
    for day in date_range:
        day_checklists = checklists.filter(date=day)
        day_count = day_checklists.count()
        checklist_counts.append(day_count)
        
        day_verified = day_checklists.filter(status='quality_certified').count()
        day_rejected = day_checklists.filter(status__in=['rejected', 'quality_rejected']).count()
        
        if day_verified + day_rejected > 0:
            day_success_rate = (day_verified / (day_verified + day_rejected)) * 100
        else:
            day_success_rate = 0
        
        if day_count > 0:
            day_completion_rate = ((day_verified + day_rejected) / day_count) * 100
        else:
            day_completion_rate = 0
        
        success_rates.append(day_success_rate)
        completion_rates.append(day_completion_rate)
    
    formatted_dates = [d.strftime('%m/%d') for d in date_range]
    
    trend_data = {
        'dates': formatted_dates,
        'success_rates': success_rates,
        'completion_rates': completion_rates,
        'checklist_counts': checklist_counts
    }
    
    # Prepare weekly data for stacked bar chart
    weekly_data = {'weeks': [], 'verified': [], 'rejected': [], 'pending': []}
    
    start_week = start_date - timedelta(days=start_date.weekday())
    end_week = end_date + timedelta(days=(6 - end_date.weekday()))
    
    current_week_start = start_week
    while current_week_start <= end_week:
        current_week_end = current_week_start + timedelta(days=6)
        
        week_checklists = checklists.filter(
            date__range=[current_week_start, current_week_end]
        )
        
        week_label = f"{current_week_start.strftime('%m/%d')} - {current_week_end.strftime('%m/%d')}"
        weekly_data['weeks'].append(week_label)
        
        weekly_data['verified'].append(week_checklists.filter(status='quality_certified').count())
        weekly_data['rejected'].append(week_checklists.filter(status__in=['rejected', 'quality_rejected']).count())
        weekly_data['pending'].append(week_checklists.filter(status__in=['pending', 'supervisor_approved']).count())
        
        current_week_start += timedelta(days=7)
    
    context = {
        'stats': stats,
        'checkpoint_stats': checkpoint_stats,
        'open_issues': open_issues,
        'chart_data': json.dumps(chart_data),
        'trend_data': json.dumps(trend_data),
        'weekly_data': json.dumps(weekly_data),
        'days': days if not custom_date else 0,
        'custom_date': custom_date,
        'start_date': start_date,
        'end_date': end_date,
        'title': 'DTPM Dashboard'
    }
    
    return render(request, 'main/dtpm/dtpm_new_dashboard.html', context)




from django.http import FileResponse, Http404
import os

def serve_reference_image(request, checkpoint_id):
    """Serve a reference image for a checkpoint from a specific location"""
    try:
        # Path to the images
        image_path = f"C:\\Renata\\Phinia-Image\\Picture{checkpoint_id}.png"
        
        # Check if file exists
        if not os.path.exists(image_path):
            raise Http404(f"Image for checkpoint {checkpoint_id} not found")
        
        # Return the file
        return FileResponse(open(image_path, 'rb'), content_type='image/png')
    except:
        raise Http404("Image not found")
    
    
    
    
    
    
from django.http import FileResponse, Http404
import os

def serve_checkpoint_image(request, checkpoint_id):
    """Serve a checkpoint reference image from the specified local folder"""
    try:
        # Path to the image
        image_path = f"C:\\Renata\\Phinia-Image\\Picture{checkpoint_id}.png"
        
        # Check if file exists
        if not os.path.exists(image_path):
            raise Http404(f"Image for checkpoint {checkpoint_id} not found")
        
        # Return the file
        return FileResponse(open(image_path, 'rb'), content_type='image/png')
    except Exception as e:
        # Log the error for debugging
        print(f"Error serving image: {str(e)}")
        raise Http404("Image not found")
    
    
# Add this to your views.py file
from django.http import FileResponse, Http404
import os

def dtpm_image_view(request, image_id):
    """Serve a DTPM reference image from the specified location"""
    try:
        # Path to the image
        image_path = f"C:\\Renata\\Phinia-Image\\Picture{image_id}.png"
        
        # Check if file exists
        if not os.path.exists(image_path):
            raise Http404(f"Image for checkpoint {image_id} not found")
        
        # Return the file
        return FileResponse(open(image_path, 'rb'), content_type='image/png')
    except Exception as e:
        # Log the error for debugging
        print(f"Error serving image: {str(e)}")
        raise Http404("Image not found")        
    
    
    
    
    
    
    
# New CheckSheet  views 

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q, Count, Prefetch
from django.utils import timezone
from django.core.paginator import Paginator
from .models import (
    Checksheet, ChecksheetSection, ChecksheetField,
    ChecksheetResponse, ChecksheetFieldResponse,
    DailyVerificationStatus, Shift, User
)
import json


# ============ CHECKSHEET LIST VIEW ============

@login_required
def checksheet_list(request):
    """Display list of all active checksheets"""
    checksheets = Checksheet.objects.filter(is_active=True).annotate(
        total_sections=Count('sections', filter=Q(sections__is_active=True)),
        total_fields=Count('sections__fields', filter=Q(sections__is_active=True, sections__fields__is_active=True))
    ).order_by('name')
    
    context = {
        'checksheets': checksheets,
        'page_title': 'Available Checksheets'
    }
    return render(request, 'checksheets/checksheet_list.html', context)


# ============ CHECKSHEET RESPONSES LIST ============

@login_required
def checksheet_responses_list(request):
    """Display list of all checksheet responses"""
    responses = ChecksheetResponse.objects.select_related(
        'checksheet', 'filled_by', 'supervisor_approved_by', 'quality_approved_by'
    ).order_by('-created_at')
    
    # Filters
    status_filter = request.GET.get('status')
    checksheet_filter = request.GET.get('checksheet')
    
    if status_filter:
        responses = responses.filter(status=status_filter)
    
    if checksheet_filter:
        responses = responses.filter(checksheet_id=checksheet_filter)
    
    # Filter based on user role
    if request.user.user_type == 'operator':
        responses = responses.filter(filled_by=request.user)
    
    # Pagination
    paginator = Paginator(responses, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get all checksheets for filter
    checksheets = Checksheet.objects.filter(is_active=True).order_by('name')
    
    context = {
        'page_obj': page_obj,
        'responses': page_obj,
        'checksheets': checksheets,
        'selected_status': status_filter,
        'selected_checksheet': checksheet_filter,
        'page_title': 'Checksheet Responses'
    }
    return render(request, 'checksheets/responses_list.html', context)


# ============ CREATE CHECKSHEET RESPONSE ============

@login_required
def create_checksheet_response(request, checksheet_id):
    """Create a new checksheet response"""
    checksheet = get_object_or_404(
        Checksheet.objects.prefetch_related(
            Prefetch('sections', queryset=ChecksheetSection.objects.filter(is_active=True).order_by('order')),
            Prefetch('sections__fields', queryset=ChecksheetField.objects.filter(is_active=True).order_by('order'))
        ),
        pk=checksheet_id,
        is_active=True
    )
    
    if request.method == 'POST':
        # Create response
        response = ChecksheetResponse.objects.create(
            checksheet=checksheet,
            filled_by=request.user,
            status='draft'
        )
        
        # Process form data
        selected_model = None
        errors = []
        
        # First pass: collect selected model
        for section in checksheet.sections.filter(is_active=True):
            for field in section.fields.filter(is_active=True):
                if 'Program selection' in field.label or 'program' in field.label.lower():
                    field_key = f'field_{field.id}'
                    selected_model = request.POST.get(field_key, '')
                    break
        
        # Second pass: process all fields
        for section in checksheet.sections.filter(is_active=True).order_by('order'):
            for field in section.fields.filter(is_active=True).order_by('order'):
                field_key = f'field_{field.id}'
                status_key = f'status_{field.id}'
                comment_key = f'comment_{field.id}'
                
                # Initialize variables
                value = ''
                status = ''
                comment = request.POST.get(comment_key, '').strip()
                
                # CRITICAL FIX: Handle different field types correctly
                if field.field_type == 'ok_nok':
                    if field.has_status_field:
                        # For fields with status, the OK/NOK is in status_key and actual value is in field_key
                        status = request.POST.get(status_key, '').strip()
                        value = request.POST.get(field_key, '').strip()
                    else:
                        # For simple OK/NOK fields, the value IS the OK/NOK selection
                        value = request.POST.get(field_key, '').strip()
                        status = ''  # No separate status
                elif field.field_type == 'yes_no':
                    if field.has_status_field:
                        status = request.POST.get(status_key, '').strip()
                        value = request.POST.get(field_key, '').strip()
                    else:
                        value = request.POST.get(field_key, '').strip()
                        status = ''
                else:
                    # For all other field types
                    value = request.POST.get(field_key, '').strip()
                    if field.has_status_field:
                        status = request.POST.get(status_key, '').strip()
                
                # Auto-fill based on model (only if value is empty)
                if field.auto_fill_based_on_model and selected_model and field.model_value_mapping and not value:
                    auto_value = field.get_value_for_model(selected_model)
                    if auto_value:
                        value = auto_value
                
                # VALIDATION
                # Check if field is required
                if field.is_required:
                    # For OK/NOK fields without separate status, check the value
                    if field.field_type == 'ok_nok' and not field.has_status_field:
                        if not value:
                            errors.append(f"{field.label} is required")
                    # For fields with status field, check status
                    elif field.has_status_field:
                        if not status:
                            errors.append(f"{field.label} is required")
                    # For all other fields, check value
                    else:
                        if not value:
                            errors.append(f"{field.label} is required")
                
                # Check if comment is required when NOK/No
                if field.requires_comment_if_nok:
                    # Check status field if it exists
                    if field.has_status_field:
                        if status in ['NOK', 'No'] and not comment:
                            errors.append(f"Comment is required for '{field.label}' when {status}")
                    # Otherwise check the value itself for OK/NOK or Yes/No fields
                    elif field.field_type in ['ok_nok', 'yes_no']:
                        if value in ['NOK', 'No'] and not comment:
                            errors.append(f"Comment is required for '{field.label}' when {value}")
                
                # Validate numeric ranges
                if field.field_type in ['number', 'decimal'] and value:
                    try:
                        numeric_value = float(value)
                        if field.min_value is not None and numeric_value < field.min_value:
                            errors.append(f"{field.label} must be at least {field.min_value} {field.unit}")
                        if field.max_value is not None and numeric_value > field.max_value:
                            errors.append(f"{field.label} must be at most {field.max_value} {field.unit}")
                    except ValueError:
                        errors.append(f"{field.label} must be a valid number")
                
                # Create field response
                ChecksheetFieldResponse.objects.create(
                    response=response,
                    field=field,
                    value=value,
                    status=status,
                    comment=comment,
                    filled_by=request.user
                )
        
        if errors:
            # If there are errors, show them and delete the response
            for error in errors:
                messages.error(request, error)
            response.delete()
            
            # Properly prefetch related fields for re-rendering
            sections_with_fields = checksheet.sections.filter(is_active=True).prefetch_related(
                Prefetch('fields', queryset=ChecksheetField.objects.filter(is_active=True).order_by('order'))
            ).order_by('order')
            
            # Re-render form with data
            context = {
                'checksheet': checksheet,
                'sections': sections_with_fields,
                'page_title': f'Create {checksheet.name}',
                'is_edit': False,
                'form_data': request.POST
            }
            return render(request, 'checksheets/checksheet_form.html', context)
        
        # Check if user wants to submit or save as draft
        action = request.POST.get('action')
        if action == 'submit':
            response.status = 'submitted'
            response.submitted_at = timezone.now()
            response.save()
            messages.success(request, 'Checksheet submitted successfully! Waiting for supervisor approval.')
            return redirect('checksheet_response_detail', response_id=response.id)
        else:
            messages.success(request, 'Checksheet saved as draft!')
            return redirect('edit_checksheet_response', response_id=response.id)
    
    # GET request - show form
    sections = checksheet.sections.filter(is_active=True).prefetch_related(
        Prefetch('fields', queryset=ChecksheetField.objects.filter(is_active=True).order_by('order'))
    ).order_by('order')
    
    context = {
        'checksheet': checksheet,
        'sections': sections,
        'page_title': f'Create {checksheet.name}',
        'is_edit': False
    }
    return render(request, 'checksheets/checksheet_form.html', context)


@login_required
def edit_checksheet_response(request, response_id):
    """Edit an existing checksheet response"""
    response = get_object_or_404(
        ChecksheetResponse.objects.select_related('checksheet', 'filled_by'),
        pk=response_id
    )
    
    # Check if user is the one who filled it OR is a supervisor/admin
    is_owner = response.filled_by == request.user
    is_supervisor = request.user.user_type in ['shift_supervisor', 'quality_supervisor']
    is_admin = request.user.is_superuser
    
    if not (is_owner or is_supervisor or is_admin):
        messages.error(request, 'You do not have permission to edit this checksheet.')
        return redirect('checksheet_response_detail', response_id=response.id)
    
    # Check status - allow editing for drafts, rejected, and submitted (if you're a supervisor)
    editable_statuses = ['draft', 'rejected']
    if is_supervisor or is_admin:
        editable_statuses.extend(['submitted', 'supervisor_approved'])
    
    if response.status not in editable_statuses:
        messages.error(request, f'This checksheet cannot be edited because it is {response.get_status_display()}.')
        return redirect('checksheet_response_detail', response_id=response.id)
    
    checksheet = response.checksheet
    
    if request.method == 'POST':
        selected_model = None
        errors = []
        
        # First pass: collect selected model
        for section in checksheet.sections.filter(is_active=True):
            for field in section.fields.filter(is_active=True):
                if 'Program selection' in field.label or 'program' in field.label.lower():
                    field_key = f'field_{field.id}'
                    selected_model = request.POST.get(field_key, '')
                    break
        
        # Second pass: update all fields
        for section in checksheet.sections.filter(is_active=True).order_by('order'):
            for field in section.fields.filter(is_active=True).order_by('order'):
                field_key = f'field_{field.id}'
                status_key = f'status_{field.id}'
                comment_key = f'comment_{field.id}'
                
                # Initialize variables
                value = ''
                status = ''
                comment = request.POST.get(comment_key, '').strip()
                
                # Handle different field types correctly
                if field.field_type == 'ok_nok':
                    if field.has_status_field:
                        status = request.POST.get(status_key, '').strip()
                        value = request.POST.get(field_key, '').strip()
                    else:
                        value = request.POST.get(field_key, '').strip()
                        status = ''
                elif field.field_type == 'yes_no':
                    if field.has_status_field:
                        status = request.POST.get(status_key, '').strip()
                        value = request.POST.get(field_key, '').strip()
                    else:
                        value = request.POST.get(field_key, '').strip()
                        status = ''
                else:
                    value = request.POST.get(field_key, '').strip()
                    if field.has_status_field:
                        status = request.POST.get(status_key, '').strip()
                
                # Auto-fill based on model (only if value is empty)
                if field.auto_fill_based_on_model and selected_model and field.model_value_mapping and not value:
                    auto_value = field.get_value_for_model(selected_model)
                    if auto_value:
                        value = auto_value
                
                # VALIDATION
                if field.is_required:
                    if field.field_type == 'ok_nok' and not field.has_status_field:
                        if not value:
                            errors.append(f"{field.label} is required")
                    elif field.has_status_field:
                        if not status:
                            errors.append(f"{field.label} is required")
                    else:
                        if not value:
                            errors.append(f"{field.label} is required")
                
                if field.requires_comment_if_nok:
                    if field.has_status_field:
                        if status in ['NOK', 'No'] and not comment:
                            errors.append(f"Comment is required for '{field.label}' when {status}")
                    elif field.field_type in ['ok_nok', 'yes_no']:
                        if value in ['NOK', 'No'] and not comment:
                            errors.append(f"Comment is required for '{field.label}' when {value}")
                
                if field.field_type in ['number', 'decimal'] and value:
                    try:
                        numeric_value = float(value)
                        if field.min_value is not None and numeric_value < field.min_value:
                            errors.append(f"{field.label} must be at least {field.min_value} {field.unit}")
                        if field.max_value is not None and numeric_value > field.max_value:
                            errors.append(f"{field.label} must be at most {field.max_value} {field.unit}")
                    except ValueError:
                        errors.append(f"{field.label} must be a valid number")
                
                # Update or create field response
                field_response, created = ChecksheetFieldResponse.objects.update_or_create(
                    response=response,
                    field=field,
                    defaults={
                        'value': value,
                        'status': status,
                        'comment': comment,
                        'filled_by': request.user
                    }
                )
        
        if errors:
            for error in errors:
                messages.error(request, error)
            # Continue to re-render the form with errors
        else:
            # Check if user wants to submit or save as draft
            action = request.POST.get('action')
            if action == 'submit':
                response.status = 'submitted'
                response.submitted_at = timezone.now()
                response.save()
                messages.success(request, 'Checksheet submitted successfully! Waiting for supervisor approval.')
                return redirect('checksheet_response_detail', response_id=response.id)
            else:
                response.updated_at = timezone.now()
                response.save()
                messages.success(request, 'Checksheet updated successfully!')
                return redirect('edit_checksheet_response', response_id=response.id)
    
    # GET request or form has errors - show form with existing data
    sections = checksheet.sections.filter(is_active=True).prefetch_related(
        Prefetch('fields', queryset=ChecksheetField.objects.filter(is_active=True).order_by('order'))
    ).order_by('order')
    
    # Build response data for template
    existing_responses = {}
    for field_response in response.field_responses.select_related('field'):
        existing_responses[field_response.field.id] = {
            'field_id': field_response.field.id,
            'value': field_response.value,
            'status': field_response.status,
            'comment': field_response.comment,
        }
    
    context = {
        'checksheet': checksheet,
        'sections': sections,
        'response': response,
        'existing_responses': existing_responses,
        'page_title': f'Edit {checksheet.name}',
        'is_edit': True
    }
    return render(request, 'checksheets/checksheet_form.html', context)
# ============ EDIT CHECKSHEET RESPONSE ============

# @login_required
# def edit_checksheet_response(request, response_id):
#     """Edit an existing checksheet response"""
#     response = get_object_or_404(
#         ChecksheetResponse.objects.select_related('checksheet', 'filled_by')
#         .prefetch_related('field_responses__field__section'),
#         pk=response_id
#     )
    
#     # Check permissions
#     if response.filled_by != request.user and not request.user.is_superuser:
#         messages.error(request, 'You do not have permission to edit this response')
#         return redirect('checksheet_responses_list')
    
#     # Check if response can be edited
#     if not response.can_be_edited:
#         messages.error(request, f'Cannot edit response with status: {response.get_status_display()}')
#         return redirect('checksheet_response_detail', response_id=response.id)
    
#     checksheet = response.checksheet
    
#     if request.method == 'POST':
#         # Get existing field responses
#         existing_responses = {fr.field_id: fr for fr in response.field_responses.all()}
        
#         selected_model = None
#         errors = []
        
#         # First pass: collect selected model
#         for section in checksheet.sections.filter(is_active=True):
#             for field in section.fields.filter(is_active=True):
#                 if 'Program selection' in field.label or 'program' in field.label.lower():
#                     field_key = f'field_{field.id}'
#                     selected_model = request.POST.get(field_key, '')
#                     break
        
#         # Second pass: process all fields
#         for section in checksheet.sections.filter(is_active=True).order_by('order'):
#             for field in section.fields.filter(is_active=True).order_by('order'):
#                 field_key = f'field_{field.id}'
#                 status_key = f'status_{field.id}'
#                 comment_key = f'comment_{field.id}'
                
#                 value = request.POST.get(field_key, '').strip()
#                 status = request.POST.get(status_key, '')
#                 comment = request.POST.get(comment_key, '').strip()
                
#                 # Auto-fill based on model
#                 if field.auto_fill_based_on_model and selected_model and field.model_value_mapping:
#                     auto_value = field.get_value_for_model(selected_model)
#                     if auto_value:
#                         value = auto_value
                
#                 # Validation
#                 if field.is_required:
#                     if field.field_type == 'ok_nok':
#                         if not status:
#                             errors.append(f"{field.label} is required")
#                     elif not value:
#                         errors.append(f"{field.label} is required")
                
#                 # Check if comment is required
#                 if field.requires_comment_if_nok:
#                     if status == 'NOK' and not comment:
#                         errors.append(f"Comment is required for '{field.label}' when status is NOK")
                
#                 # Validate numeric ranges
#                 if field.field_type in ['number', 'decimal'] and value:
#                     try:
#                         numeric_value = float(value)
#                         if field.min_value is not None and numeric_value < field.min_value:
#                             errors.append(f"{field.label} must be at least {field.min_value} {field.unit}")
#                         if field.max_value is not None and numeric_value > field.max_value:
#                             errors.append(f"{field.label} must be at most {field.max_value} {field.unit}")
#                     except ValueError:
#                         errors.append(f"{field.label} must be a valid number")
                
#                 # Update or create field response
#                 if field.id in existing_responses:
#                     field_response = existing_responses[field.id]
#                     field_response.value = value
#                     field_response.status = status
#                     field_response.comment = comment
#                     field_response.filled_by = request.user
#                     field_response.save()
#                 else:
#                     ChecksheetFieldResponse.objects.create(
#                         response=response,
#                         field=field,
#                         value=value,
#                         status=status,
#                         comment=comment,
#                         filled_by=request.user
#                     )
        
#         if errors:
#             for error in errors:
#                 messages.error(request, error)
#             # Re-render form with data
#             context = {
#                 'checksheet': checksheet,
#                 'sections': checksheet.sections.filter(is_active=True).order_by('order'),
#                 'response': response,
#                 'page_title': f'Edit {checksheet.name}',
#                 'is_edit': True,
#                 'form_data': request.POST
#             }
#             return render(request, 'checksheets/checksheet_form.html', context)
        
#         # Update response metadata
#         response.updated_at = timezone.now()
        
#         # Check if user wants to submit or save as draft
#         action = request.POST.get('action')
#         if action == 'submit':
#             response.status = 'submitted'
#             response.submitted_at = timezone.now()
#             response.save()
#             messages.success(request, 'Checksheet submitted successfully! Waiting for supervisor approval.')
#             return redirect('checksheet_response_detail', response_id=response.id)
#         else:
#             response.save()
#             messages.success(request, 'Checksheet updated successfully!')
#             return redirect('edit_checksheet_response', response_id=response.id)
    
#     # GET request - show form with existing data
#     sections = checksheet.sections.filter(is_active=True).order_by('order')
    
#     # Get existing responses
#     existing_responses = {fr.field_id: fr for fr in response.field_responses.all()}
    
#     context = {
#         'checksheet': checksheet,
#         'sections': sections,
#         'response': response,
#         'existing_responses': existing_responses,
#         'page_title': f'Edit {checksheet.name}',
#         'is_edit': True
#     }
#     return render(request, 'checksheets/checksheet_form.html', context)


# ============ VIEW CHECKSHEET RESPONSE ============

@login_required
def checksheet_response_detail(request, response_id):
    """View checksheet response details"""
    response = get_object_or_404(
        ChecksheetResponse.objects.select_related(
            'checksheet', 'filled_by', 'supervisor_approved_by', 'quality_approved_by'
        ).prefetch_related(
            'field_responses__field__section'
        ),
        pk=response_id
    )
    
    # Check permissions
    if not request.user.is_superuser:
        if request.user.user_type == 'operator' and response.filled_by != request.user:
            messages.error(request, 'You do not have permission to view this response')
            return redirect('checksheet_responses_list')
    
    # Group responses by section
    sections_data = []
    for section in response.checksheet.sections.filter(is_active=True).order_by('order'):
        field_responses = response.field_responses.filter(
            field__section=section
        ).select_related('field').order_by('field__order')
        
        if field_responses.exists():
            sections_data.append({
                'section': section,
                'field_responses': field_responses
            })
    
    context = {
        'response': response,
        'sections_data': sections_data,
        'page_title': f'{response.checksheet.name} - Details',
        'can_approve_supervisor': request.user.user_type == 'shift_supervisor' and response.can_be_approved_by_supervisor,
        'can_approve_quality': request.user.user_type == 'quality_supervisor' and response.can_be_approved_by_quality,
        'can_edit': response.can_be_edited and response.filled_by == request.user
    }
    return render(request, 'checksheets/checksheet_response_detail.html', context)


# ============ SUPERVISOR APPROVAL ============

@login_required
def supervisor_approve_response(request, response_id):
    """Supervisor approves a checksheet response"""
    if request.user.user_type != 'shift_supervisor' and not request.user.is_superuser:
        messages.error(request, 'Only shift supervisors can approve checksheets')
        return redirect('checksheet_responses_list')
    
    response = get_object_or_404(ChecksheetResponse, pk=response_id)
    
    if not response.can_be_approved_by_supervisor:
        messages.error(request, f'Cannot approve response with status: {response.get_status_display()}')
        return redirect('checksheet_response_detail', response_id=response.id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        comments = request.POST.get('comments', '').strip()
        
        if action == 'approve':
            response.status = 'supervisor_approved'
            response.supervisor_approved_by = request.user
            response.supervisor_approved_at = timezone.now()
            response.supervisor_comments = comments
            response.save()
            messages.success(request, 'Checksheet approved by supervisor successfully!')
        elif action == 'reject':
            if not comments:
                messages.error(request, 'Please provide a reason for rejection')
                return redirect('checksheet_response_detail', response_id=response.id)
            
            response.status = 'rejected'
            response.rejection_reason = comments
            response.supervisor_comments = comments
            response.save()
            messages.warning(request, 'Checksheet rejected. Operator can now re-submit after corrections.')
        
        return redirect('checksheet_response_detail', response_id=response.id)
    
    return redirect('checksheet_response_detail', response_id=response.id)


# ============ QUALITY APPROVAL ============

@login_required
def quality_approve_response(request, response_id):
    """Quality supervisor approves a checksheet response"""
    if request.user.user_type != 'quality_supervisor' and not request.user.is_superuser:
        messages.error(request, 'Only quality supervisors can approve checksheets')
        return redirect('checksheet_responses_list')
    
    response = get_object_or_404(ChecksheetResponse, pk=response_id)
    
    if not response.can_be_approved_by_quality:
        messages.error(request, f'Cannot approve response with status: {response.get_status_display()}')
        return redirect('checksheet_response_detail', response_id=response.id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        comments = request.POST.get('comments', '').strip()
        
        if action == 'approve':
            response.status = 'quality_approved'
            response.quality_approved_by = request.user
            response.quality_approved_at = timezone.now()
            response.quality_comments = comments
            response.save()
            messages.success(request, 'Checksheet approved by quality successfully!')
        elif action == 'reject':
            if not comments:
                messages.error(request, 'Please provide a reason for rejection')
                return redirect('checksheet_response_detail', response_id=response.id)
            
            response.status = 'rejected'
            response.rejection_reason = comments
            response.quality_comments = comments
            response.save()
            messages.warning(request, 'Checksheet rejected. Operator can now re-submit after corrections.')
        
        return redirect('checksheet_response_detail', response_id=response.id)
    
    return redirect('checksheet_response_detail', response_id=response.id)


# ============ DELETE CHECKSHEET RESPONSE ============

@login_required
def delete_checksheet_response(request, response_id):
    """Delete a checksheet response"""
    response = get_object_or_404(ChecksheetResponse, pk=response_id)
    
    # Check permissions
    if response.filled_by != request.user and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to delete this response')
        return redirect('checksheet_responses_list')
    
    # Only allow deletion of drafts
    if response.status != 'draft':
        messages.error(request, 'Only draft responses can be deleted')
        return redirect('checksheet_response_detail', response_id=response.id)
    
    if request.method == 'POST':
        checksheet_name = response.checksheet.name
        response.delete()
        messages.success(request, f'{checksheet_name} response deleted successfully!')
        return redirect('checksheet_responses_list')
    
    return redirect('checksheet_response_detail', response_id=response.id)


# ============ AJAX: GET FIELD AUTO-FILL VALUE ============

@login_required
def get_autofill_value(request):
    """AJAX endpoint to get auto-fill value for a field based on model"""
    field_id = request.GET.get('field_id')
    model = request.GET.get('model')
    
    if not field_id or not model:
        return JsonResponse({'error': 'Missing parameters'}, status=400)
    
    try:
        field = ChecksheetField.objects.get(pk=field_id)
        value = field.get_value_for_model(model)
        return JsonResponse({'value': value})
    except ChecksheetField.DoesNotExist:
        return JsonResponse({'error': 'Field not found'}, status=404)